#PZone.py
#!/usr/bin/env python3
"""
Foodservice Zone Optimization Engine v9
NOW WITH: Flexible column configuration + Learning system
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Set
from collections import defaultdict
import warnings
import os
import json
from dataclasses import dataclass, asdict
warnings.filterwarnings('ignore')

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule, Rule
from openpyxl.styles.differential import DifferentialStyle

# ============================================================================
# ENHANCED CONFIGURATION
# ============================================================================

@dataclass
class DataConfiguration:
    """Configure which columns to use for grouping."""
    
    # Required columns (always needed)
    company_column: str = 'Company Name'
    customer_id_column: str = 'Company Customer Number'
    last_invoice_date_column: str = 'Last Invoice Date'
    fiscal_week_column: str = 'Fiscal Week Number'
    pounds_cy_column: str = 'Pounds CY'
    pounds_py_column: str = 'Pounds PY'
    zone_column: str = 'Zone_Suffix_Numeric'
    
    # Filtering columns
    company_number_column: str = 'Company Number'
    company_region_id_column: str = 'Company Region ID'
    
    # LAYER 1: Extended Totals (for volume calculations)
    margin_cy_column: str = 'Computer Margin Ext $ CY'
    margin_py_column: str = 'Computer Margin Ext $ PY'
    net_sales_cy_column: str = 'Net Sales Ext $ CY'
    net_sales_py_column: str = 'Net Sales Ext $ PY'
    has_margin_data: bool = True
    has_net_sales_data: bool = True
    
    # LAYER 2: Per-Pound Rates (for pricing sensitivity)
    margin_per_lb_cy_column: str = 'Computer Margin $ Per LB CY'
    margin_per_lb_py_column: str = 'Computer Margin $ Per LB PY'
    net_sales_per_lb_cy_column: str = 'Net Sales Ext $ Per LB CY'
    net_sales_per_lb_py_column: str = 'Net Sales Ext $ Per LB PY'
    has_per_lb_rates: bool = True
    
    # LAYER 3: Margin Percentage (for profitability analysis)
    margin_pct_cy_column: str = 'Computer Margin Ext $ % Net Sales CY'
    margin_pct_py_column: str = 'Computer Margin Ext $ % Net Sales PY'
    has_margin_pct: bool = True

    # Optional grouping columns (existing - keep as is)
    use_attribute_group: bool = True
    attribute_group_column: str = 'Attribute Group ID'
    
    use_business_center: bool = False
    business_center_column: str = 'Business Center ID'
    
    use_item_group: bool = False
    item_group_column: str = 'Item Group ID'
    
    use_cuisine: bool = True
    cuisine_column: str = 'NPD Cuisine Type'
    
    # Price source (usually CPA only)
    use_price_source: bool = True
    price_source_column: str = 'Price Source Type'
    
    def get_grouping_columns(self) -> List[str]:
        """Get list of columns to use for combo grouping."""
        cols = [self.company_column]
        
        if self.use_cuisine:
            cols.append(self.cuisine_column)
        if self.use_business_center:
            cols.append(self.business_center_column)
        if self.use_attribute_group:
            cols.append(self.attribute_group_column)
        if self.use_item_group:
            cols.append(self.item_group_column)
        if self.use_price_source:
            cols.append(self.price_source_column)
        
        return cols
    
    def validate_dataframe(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Check if dataframe has required columns."""
        missing = []
        
        # Check required columns
        required = [
            self.company_column,
            self.customer_id_column,
            self.last_invoice_date_column,
            self.pounds_cy_column
        ]
        
        for col in required:
            if col not in df.columns:
                missing.append(col)
        
        # Check enabled optional columns
        if self.use_attribute_group and self.attribute_group_column not in df.columns:
            missing.append(f"{self.attribute_group_column} (enabled but missing)")
        if self.use_business_center and self.business_center_column not in df.columns:
            missing.append(f"{self.business_center_column} (enabled but missing)")
        if self.use_item_group and self.item_group_column not in df.columns:
            missing.append(f"{self.item_group_column} (enabled but missing)")
        
        return (len(missing) == 0, missing)


@dataclass
class InputConfiguration:
    """Configure input file paths."""
    
    current_data_path: str = ""
    historical_data_paths: List[str] = None
    
    # Output settings
    output_directory: str = r"C:\Users\kmor6669\OneDrive - Sysco Corporation\Desktop\Pricing"
    output_name_prefix: str = "zone_optimization"
    
    # Learning file (where we save state)
    learning_file_path: str = None  # Auto-generated if None
    
    def __post_init__(self):
        if self.historical_data_paths is None:
            self.historical_data_paths = []
        
        # Auto-generate learning file path
        if self.learning_file_path is None:
            self.learning_file_path = os.path.join(
                self.output_directory,
                "zone_optimization_learning_state.json"
            )
    
    def validate_paths(self) -> Tuple[bool, List[str]]:
        """Check if paths exist."""
        issues = []
        
        if not self.current_data_path:
            issues.append("Current data path not specified")
        elif not os.path.exists(self.current_data_path):
            issues.append(f"Current data file not found: {self.current_data_path}")
        
        for path in self.historical_data_paths:
            if not os.path.exists(path):
                issues.append(f"Historical file not found: {path}")
        
        if not os.path.exists(self.output_directory):
            try:
                os.makedirs(self.output_directory)
            except Exception as e:
                issues.append(f"Cannot create output directory: {e}")
        
        return (len(issues) == 0, issues)


class FoodserviceConfig:
    """Main configuration class - combines all settings."""
    
    # Customer activity thresholds (in days)
    LAPSED_FROM_CATEGORY_DAYS = 45  # 6-7 weeks
    LOST_CUSTOMER_DAYS = 60  # 8-9 weeks
    
    # Analysis windows (in weeks)
    REACTIVE_LOOKBACK_WEEKS = 6
    BEHAVIOR_WINDOW_WEEKS = 12
    YOY_LOOKBACK_WEEKS = 8  # ← NEW: Configurable YoY comparison window
    
    # Recommendation thresholds
    MIN_VOLUME_FOR_ACTION = 1000
    HIGH_RECOVERY_THRESHOLD = 0.30
    
    # Filtering (NEW!)
    FILTER_BY_company_number: Optional[str] = None  # Set to specific Company Number or None for all
    FILTER_BY_company_region_id: Optional[str] = None   # Set to specific Company Region ID or None for all
    
    # ==========================================
    # LEADS GENERATION CONFIGURATION
    # ==========================================
    GENERATE_LEADS_FILE: bool = True  # Toggle to enable/disable
    LEADS_OUTPUT_PATH: str = r"C:\Users\kmor6669\OneDrive - Sysco Corporation\Desktop\Pricing\leads_guide.csv"
    
    # Revenue filtering
    LEADS_MIN_REVENUE_THRESHOLD: float = 10000  # Only customers with $10K+ annual revenue
    
    # Number of leads to generate
    LEADS_TOP_N_ALL_COMPANIES: int = 100  # When NO company filter: top 100 across all companies
    LEADS_N_FOR_FILTERED_COMPANY: int = 50  # When Company_Number filter IS active: how many for that company
    
    # ✅ ADD THESE 3 LINES HERE:
    has_per_lb_rates: bool = True
    has_margin_pct: bool = True
    has_margin_data: bool = True

    def __init__(self, 
                 data_config: Optional[DataConfiguration] = None,
                 input_config: Optional[InputConfiguration] = None,
                 yoy_lookback_weeks: int = 8,
                 filter_company_number: Optional[str] = None,
                 filter_company_region_id: Optional[str] = None):
        
        self.data_config = data_config or DataConfiguration()
        self.input_config = input_config or InputConfiguration()
        
        # Override defaults with parameters
        self.YOY_LOOKBACK_WEEKS = yoy_lookback_weeks
        self.FILTER_BY_company_number = filter_company_number
        self.FILTER_BY_company_region_id = filter_company_region_id
    
    @classmethod
    def get_timestamp(cls):
        return datetime.now().strftime("%Y%m%d_%H%M%S")

# ============================================================================
# LEARNING SYSTEM
# ============================================================================

@dataclass
class RecommendationRecord:
    """Track a single recommendation over time."""
    
    recommendation_id: str  # Unique ID
    combo_key: str
    company_name: str
    category_description: str
    
    # Recommendation details
    date_recommended: str
    from_zone: int
    to_zone: int
    recommendation_type: str
    
    # Predicted outcomes
    predicted_volume_lift: float
    predicted_customer_recovery: int
    predicted_timeline_weeks: int
    
    # Actual outcomes (filled in later)
    date_implemented: Optional[str] = None
    was_implemented: bool = False
    actual_volume_lift: Optional[float] = None
    actual_customer_recovery: Optional[int] = None
    weeks_to_result: Optional[int] = None
    
    # Learning
    outcome_vs_prediction: Optional[str] = None  # 'BETTER', 'AS_EXPECTED', 'WORSE'
    lessons_learned: List[str] = None
    
    def __post_init__(self):
        if self.lessons_learned is None:
            self.lessons_learned = []


class LearningEngine:
    """
    Tracks recommendations over time and learns from outcomes.
    
    Simple explanation:
    "This is the system's memory. Every time we make a recommendation, we write 
    it down. Every time we check results, we compare what happened vs what we 
    predicted. Over time, the system gets smarter about what works."
    """
    
    def __init__(self, learning_file_path: str):
        self.learning_file_path = learning_file_path
        self.recommendations: Dict[str, RecommendationRecord] = {}
        self.performance_history: List[Dict] = []
        self.zone_effectiveness_learnings: Dict[str, Dict] = defaultdict(dict)
        
        # Load existing learning if available
        self.load_state()
    
    def save_recommendation(self, rec: Dict, predicted_outcomes: Dict) -> str:
        """Save a new recommendation to learning system."""
        
        # Generate unique ID
        rec_id = f"{rec['company_combo']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        record = RecommendationRecord(
            recommendation_id=rec_id,
            combo_key=rec['company_combo'],
            company_name=rec['company_name'],
            category_description=f"{rec.get('cuisine', 'N/A')} - AG{rec.get('attribute_group', 'N/A')}",
            date_recommended=datetime.now().strftime('%Y-%m-%d'),
            from_zone=int(rec['current_zone']),
            to_zone=int(rec['recommended_zone']),
            recommendation_type=rec['recommendation_type'],
            predicted_volume_lift=float(predicted_outcomes.get('volume_lift', 0)),
            predicted_customer_recovery=int(predicted_outcomes.get('customer_recovery', 0)),
            predicted_timeline_weeks=int(predicted_outcomes.get('timeline_weeks', 6))
        )
        
        self.recommendations[rec_id] = record
        self.save_state()
        
        return rec_id
    
    def update_recommendation_outcome(self, 
                                     rec_id: str,
                                     actual_volume_lift: float,
                                     actual_customer_recovery: int,
                                     weeks_elapsed: int):
        """Update a recommendation with actual outcomes."""
        
        if rec_id not in self.recommendations:
            print(f"⚠️  Warning: Recommendation {rec_id} not found in learning system")
            return
        
        rec = self.recommendations[rec_id]
        rec.was_implemented = True
        rec.date_implemented = datetime.now().strftime('%Y-%m-%d')
        rec.actual_volume_lift = actual_volume_lift
        rec.actual_customer_recovery = actual_customer_recovery
        rec.weeks_to_result = weeks_elapsed
        
        # Assess outcome vs prediction
        volume_ratio = actual_volume_lift / rec.predicted_volume_lift if rec.predicted_volume_lift > 0 else 0
        
        if volume_ratio >= 1.2:
            rec.outcome_vs_prediction = 'BETTER_THAN_EXPECTED'
            rec.lessons_learned.append(f"🎉 Exceeded prediction by {(volume_ratio - 1) * 100:.0f}%")
        elif volume_ratio >= 0.8:
            rec.outcome_vs_prediction = 'AS_EXPECTED'
            rec.lessons_learned.append("✅ Performed as predicted")
        else:
            rec.outcome_vs_prediction = 'WORSE_THAN_EXPECTED'
            rec.lessons_learned.append(f"⚠️  Underperformed prediction by {(1 - volume_ratio) * 100:.0f}%")
        
        # Learn zone effectiveness
        self._update_zone_effectiveness(rec)
        
        self.save_state()
    
    def _update_zone_effectiveness(self, rec: RecommendationRecord):
        """Learn which zone changes work best."""
        
        # Key by recommendation type + zone change
        pattern_key = f"{rec.recommendation_type}_{rec.from_zone}_to_{rec.to_zone}"
        
        if pattern_key not in self.zone_effectiveness_learnings:
            self.zone_effectiveness_learnings[pattern_key] = {
                'attempts': 0,
                'successes': 0,
                'total_volume_lift': 0,
                'total_customer_recovery': 0,
                'avg_weeks_to_result': 0
            }
        
        learning = self.zone_effectiveness_learnings[pattern_key]
        learning['attempts'] += 1
        
        if rec.outcome_vs_prediction in ['BETTER_THAN_EXPECTED', 'AS_EXPECTED']:
            learning['successes'] += 1
        
        learning['total_volume_lift'] += rec.actual_volume_lift or 0
        learning['total_customer_recovery'] += rec.actual_customer_recovery or 0
        
        # Running average of weeks to result
        n = learning['attempts']
        learning['avg_weeks_to_result'] = (
            (learning['avg_weeks_to_result'] * (n - 1) + rec.weeks_to_result) / n
        )
    
    def get_pattern_confidence(self, recommendation_type: str, from_zone: int, to_zone: int) -> float:
        """Get confidence score for a recommendation pattern based on past results."""
        
        pattern_key = f"{recommendation_type}_{from_zone}_to_{to_zone}"
        
        if pattern_key not in self.zone_effectiveness_learnings:
            return 0.5  # Neutral confidence if no history
        
        learning = self.zone_effectiveness_learnings[pattern_key]
        
        if learning['attempts'] == 0:
            return 0.5
        
        # Confidence based on success rate and sample size
        success_rate = learning['successes'] / learning['attempts']
        sample_size_factor = min(learning['attempts'] / 10, 1.0)  # Max out at 10 attempts
        
        confidence = (success_rate * 0.7) + (sample_size_factor * 0.3)
        
        return confidence
    
    def get_expected_timeline(self, recommendation_type: str, from_zone: int, to_zone: int) -> int:
        """Get expected timeline based on past results."""
        
        pattern_key = f"{recommendation_type}_{from_zone}_to_{to_zone}"
        
        if pattern_key in self.zone_effectiveness_learnings:
            avg_weeks = self.zone_effectiveness_learnings[pattern_key]['avg_weeks_to_result']
            if avg_weeks > 0:
                return int(avg_weeks)
        
        # Default based on recommendation type
        defaults = {
            'HIGH_RECOVERY_POTENTIAL': 3,
            'REACTIVE_CORRECTION': 6,
            'PEER_CONSENSUS': 5,
            'NEEDS_FRACTIONAL_ZONES': 10
        }
        
        return defaults.get(recommendation_type, 6)
    
    def get_pending_recommendations(self) -> List[RecommendationRecord]:
        """Get recommendations that haven't been updated with outcomes."""
        return [rec for rec in self.recommendations.values() if not rec.was_implemented]
    
    def get_completed_recommendations(self) -> List[RecommendationRecord]:
        """Get recommendations with outcome data."""
        return [rec for rec in self.recommendations.values() if rec.was_implemented]
    
    def generate_learning_report(self) -> Dict:
        """Generate summary of what we've learned."""
        
        completed = self.get_completed_recommendations()
        
        if not completed:
            return {
                'total_recommendations_tracked': len(self.recommendations),
                'completed_recommendations': 0,
                'message': 'No completed recommendations yet - check back after implementation'
            }
        
        better_than_expected = sum(1 for r in completed if r.outcome_vs_prediction == 'BETTER_THAN_EXPECTED')
        as_expected = sum(1 for r in completed if r.outcome_vs_prediction == 'AS_EXPECTED')
        worse_than_expected = sum(1 for r in completed if r.outcome_vs_prediction == 'WORSE_THAN_EXPECTED')
        
        total_volume_lift = sum(r.actual_volume_lift for r in completed if r.actual_volume_lift)
        total_customers_recovered = sum(r.actual_customer_recovery for r in completed if r.actual_customer_recovery)
        
# ============================================================================
# CONTINUATION FROM WHERE IT CUT OFF
# ============================================================================

        # Best performing patterns
        pattern_success_rates = {}
        for pattern_key, learning in self.zone_effectiveness_learnings.items():
            if learning['attempts'] > 0:
                pattern_success_rates[pattern_key] = {
                    'pattern': pattern_key,
                    'success_rate': f"{(learning['successes'] / learning['attempts']) * 100:.0f}%",
                    'attempts': learning['attempts'],
                    'avg_volume_lift': learning['total_volume_lift'] / learning['attempts'],
                    'avg_weeks': learning['avg_weeks_to_result']
                }
        
        return {
            'total_recommendations_tracked': len(self.recommendations),
            'completed_recommendations': len(completed),
            'better_than_expected': better_than_expected,
            'as_expected': as_expected,
            'worse_than_expected': worse_than_expected,
            'total_volume_lift_achieved': total_volume_lift,
            'total_customers_recovered': total_customers_recovered,
            'success_rate': f"{((better_than_expected + as_expected) / len(completed)) * 100:.0f}%",
            'best_performing_patterns': sorted(
                pattern_success_rates.values(), 
                key=lambda x: x['avg_volume_lift'], 
                reverse=True
            )[:5]
        }
    
    def save_state(self):
        """Save learning state to JSON file."""
        try:
            state = {
                'last_updated': datetime.now().isoformat(),
                'recommendations': {
                    rec_id: asdict(rec) 
                    for rec_id, rec in self.recommendations.items()
                },
                'zone_effectiveness_learnings': dict(self.zone_effectiveness_learnings)
            }
            
            with open(self.learning_file_path, 'w') as f:
                json.dump(state, f, indent=2)
                
        except Exception as e:
            print(f"⚠️  Warning: Could not save learning state: {e}")
    
    def load_state(self):
        """Load learning state from JSON file."""
        if not os.path.exists(self.learning_file_path):
            print(f"   ℹ️  No previous learning state found (this is normal for first run)")
            return
        
        try:
            with open(self.learning_file_path, 'r') as f:
                state = json.load(f)
            
            # Reconstruct recommendation records
            for rec_id, rec_dict in state.get('recommendations', {}).items():
                self.recommendations[rec_id] = RecommendationRecord(**rec_dict)
            
            # Reconstruct zone effectiveness learnings
            self.zone_effectiveness_learnings = defaultdict(
                dict, 
                state.get('zone_effectiveness_learnings', {})
            )
            
            print(f"   ✅ Loaded learning state from {state.get('last_updated', 'unknown date')}")
            print(f"   📚 {len(self.recommendations)} recommendations in memory")
            
        except Exception as e:
            print(f"⚠️  Warning: Could not load learning state: {e}")

# ============================================================================
# CUSTOMER ACTIVITY CLASSIFICATION
# ============================================================================

def classify_customer_activity(df: pd.DataFrame, 
                               config: FoodserviceConfig = None) -> pd.DataFrame:
    """
    Classify customers into categories for recovery targeting.
    
    8th Grade Explanation:
    "We group customers into 3 buckets:
    1. ACTIVE: Still buying this category (yay!)
    2. LAPSED: Still our customer, but stopped buying THIS category (OPPORTUNITY!)
    3. LOST: Haven't bought anything in 2+ months (moved on)"
    """
    if config is None:
        config = FoodserviceConfig()
    
    df = df.copy()
    
    # Validate required columns
    if 'Fiscal_Week' not in df.columns:
        raise ValueError(f"❌ Missing 'Fiscal_Week' - CRITICAL!")
    
    if 'Pounds_CY' not in df.columns:
        raise ValueError(f"❌ Missing 'Pounds_CY' - CRITICAL!")
    
    # Ensure numeric
    df['Fiscal_Week'] = pd.to_numeric(df['Fiscal_Week'], errors='coerce')
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    
    # Get latest fiscal week
    max_week = df['Fiscal_Week'].max()
    
    if pd.isna(max_week):
        raise ValueError("❌ No valid fiscal weeks found")
    
    # For each customer, find their last week of buying THIS category (where Pounds_CY > 0)
    category_purchases = df[df['Pounds_CY'] > 0].copy()
    
    last_category_purchase = category_purchases.groupby(
        config.data_config.customer_id_column
    )['Fiscal_Week'].max().reset_index()
    last_category_purchase.columns = ['customer', 'last_category_week']
    
    # Merge back to main df
    df = df.merge(
        last_category_purchase,
        left_on=config.data_config.customer_id_column,
        right_on='customer',
        how='left'
    )
    
    # Calculate weeks since last category purchase
    df['Weeks_Since_Category_Purchase'] = max_week - df['last_category_week']
    
    # For customers who never bought (NaN), set to a high number
    df['Weeks_Since_Category_Purchase'] = df['Weeks_Since_Category_Purchase'].fillna(999)
    
    # Get Last Invoice Date to see if they're still active buying OTHER stuff
    if config.data_config.last_invoice_date_column in df.columns:
        df['Last_Invoice_Date'] = pd.to_datetime(
            df[config.data_config.last_invoice_date_column],
            errors='coerce'
        )
        current_date = pd.Timestamp.now()
        df['Days_Since_Any_Purchase'] = (current_date - df['Last_Invoice_Date']).dt.days
    else:
        # Fallback if no Last Invoice Date
        df['Days_Since_Any_Purchase'] = df['Weeks_Since_Category_Purchase'] * 7
    
    # Classification logic
    def classify(row):
        weeks_since_category = row['Weeks_Since_Category_Purchase']
        days_since_any = row['Days_Since_Any_Purchase']
        has_current_volume = row['Pounds_CY'] > 0
        
        if has_current_volume:
            return 'ACTIVE_BUYER'
        elif weeks_since_category <= 8:  # Within 8 weeks (roughly config.LAPSED_FROM_CATEGORY_DAYS / 7)
            # Stopped buying category recently
            if days_since_any <= 30:
                return 'LAPSED_FROM_CATEGORY'  # ← THE GOLD MINE! Still buying other stuff
            else:
                return 'RECENTLY_LOST'  # Not buying anything
        elif weeks_since_category <= 12:  # 8-12 weeks (roughly config.LOST_CUSTOMER_DAYS / 7)
            return 'RECENTLY_LOST'
        else:
            return 'LOST_CUSTOMER'
    
    df['Customer_Status'] = df.apply(classify, axis=1)
    
    # Add recovery potential score
    df['Recovery_Potential'] = df['Customer_Status'].map({
        'ACTIVE_BUYER': 0,
        'LAPSED_FROM_CATEGORY': 10,  # HIGH opportunity - still active, just not buying this category
        'RECENTLY_LOST': 5,
        'LOST_CUSTOMER': 0
    })
    
    return df

def calculate_purchase_consistency(df: pd.DataFrame, 
                                   config: FoodserviceConfig) -> pd.DataFrame:
    """
    Measure week-over-week purchase consistency - THE GREEN FLAG!
    
    8th Grade Explanation:
    "If a customer buys from us EVERY WEEK for 8 weeks straight at Zone 3, 
    that's a GREEN FLAG that Zone 3 works! If they buy once then disappear, 
    that's a RED FLAG."
    
    This catches the 'stickiness' of a zone.
    """
    
    df = df.copy()
    
    # Ensure we have what we need
    if 'Pounds_CY' not in df.columns:
        df['Pounds_CY'] = 0
    
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    
    # Group by combo + zone + customer + week
    customer_weeks = df.groupby([
        'Company_Combo_Key',
        config.data_config.zone_column,
        config.data_config.customer_id_column,
        config.data_config.fiscal_week_column
    ], dropna=False).agg({
        'Pounds_CY': 'sum'
    }).reset_index()
    
    # Mark active weeks (any purchase > 0)
    customer_weeks['active_week'] = (customer_weeks['Pounds_CY'] > 0).astype(int)
    
    # Calculate consistency metrics per customer per combo/zone
    consistency = customer_weeks.groupby([
        'Company_Combo_Key',
        config.data_config.zone_column,
        config.data_config.customer_id_column
    ], dropna=False).agg({
        config.data_config.fiscal_week_column: 'nunique',  # Total weeks present
        'active_week': 'sum'  # Weeks with purchases
    }).reset_index()
    
    consistency.columns = [
        'Company_Combo_Key', 'Zone', 'Customer_ID', 
        'weeks_present', 'weeks_active'
    ]
    
    # Calculate consistency rate (% of weeks they bought when present)
    consistency['consistency_rate'] = (
        consistency['weeks_active'] / consistency['weeks_present']
    ).fillna(0)
    
    # GREEN FLAG: Consistent buyers (bought in 75%+ of weeks present)
    consistency['consistent_buyer'] = (consistency['consistency_rate'] >= 0.75).astype(int)
    
    # Aggregate to zone level
    zone_consistency = consistency.groupby([
        'Company_Combo_Key', 'Zone'
    ], dropna=False).agg({
        'Customer_ID': 'nunique',  # DISTINCT customers
        'consistent_buyer': 'sum',  # How many are consistent
        'consistency_rate': 'mean',  # Average consistency
        'weeks_active': 'mean'  # Average active weeks per customer
    }).reset_index()
    
    zone_consistency.columns = [
        'Company_Combo_Key', 'Zone',
        'distinct_customers',
        'consistent_buyers',
        'avg_consistency_rate',
        'avg_active_weeks'
    ]
    
    # Calculate GREEN FLAG rate
    zone_consistency['green_flag_rate'] = (
        zone_consistency['consistent_buyers'] / 
        zone_consistency['distinct_customers'].replace(0, 1)
    )
    
    return zone_consistency

def calculate_simple_zone_scores(historical_df: pd.DataFrame, 
                                config: FoodserviceConfig) -> pd.DataFrame:
    """
    Simple zone scoring using ONLY the data you have.
    No fancy customer journeys - just straightforward analysis.
    
    8th Grade Explanation:
    "For each combo and zone, calculate:
    - How much volume did we sell? (more = better)
    - How many distinct customers bought? (more = better)  
    - How much margin did we make? (more = better, if available)
    Then pick the zone with the best score."
    """
    
    print("\n   📊 Calculating zone performance scores...")
    
    df = historical_df.copy()
    
    # Ensure we have the basics
    required = ['Company_Combo_Key', 'Zone_Suffix_Numeric', 'Pounds_CY']
    missing = [c for c in required if c not in df.columns]
    
    if missing:
        print(f"      ❌ Missing columns: {missing}")
        return pd.DataFrame()
    
    # Ensure numeric
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    df['Zone_Suffix_Numeric'] = pd.to_numeric(df['Zone_Suffix_Numeric'], errors='coerce')
    
    # Keep only valid zones (0-5)
    df = df[df['Zone_Suffix_Numeric'].between(0, 5)]
    
    # Get customer ID column
    cust_col = config.data_config.customer_id_column
    
    # Check if we have margin data
    has_margin = False
    if config.data_config.has_margin_data and 'Margin_CY' in df.columns:
        df['Margin_CY'] = pd.to_numeric(df['Margin_CY'], errors='coerce').fillna(0)
        has_margin = True
    
    # Aggregate by combo + zone
    agg_dict = {
        'Pounds_CY': 'sum',
        cust_col: 'nunique'
    }
    
    if has_margin:
        agg_dict['Margin_CY'] = 'sum'
    
    zone_perf = df.groupby(
        ['Company_Combo_Key', 'Zone_Suffix_Numeric'], 
        dropna=False
    ).agg(agg_dict).reset_index()
    
    zone_perf.columns = ['combo', 'zone', 'total_volume', 'distinct_customers']
    if has_margin:
        zone_perf['total_margin'] = df.groupby(
            ['Company_Combo_Key', 'Zone_Suffix_Numeric']
        )['Margin_CY'].sum().values
    
    print(f"      ✅ Analyzed {len(zone_perf)} zone combinations")
    
    # Calculate scores
    if has_margin:
        # WITH MARGIN: Weight profit heavily
        print(f"      💰 Scoring WITH margin data")
        zone_perf['Zone_Score'] = (
            zone_perf['total_volume'] * 0.3 +           # 30% volume
            zone_perf['distinct_customers'] * 100 * 0.3 +  # 30% customers
            zone_perf['total_margin'] * 0.4             # 40% profit
        )
    else:
        # WITHOUT MARGIN: Weight volume and customers equally
        print(f"      📦 Scoring WITHOUT margin data")
        zone_perf['Zone_Score'] = (
            zone_perf['total_volume'] * 0.5 +           # 50% volume
            zone_perf['distinct_customers'] * 100 * 0.5    # 50% customers
        )
    
    # Find optimal zone per combo (highest score)
    idx = zone_perf.groupby('combo')['Zone_Score'].idxmax()
    optimal = zone_perf.loc[idx, ['combo', 'zone', 'Zone_Score']].copy()
    optimal.columns = ['combo', 'optimal_zone', 'optimal_score']
    
    # Merge back to get full details
    zone_perf = zone_perf.merge(optimal, on='combo', how='left')
    
    # Add performance flags
    zone_perf['is_optimal'] = zone_perf['zone'] == zone_perf['optimal_zone']
    
    # Calculate how much worse non-optimal zones are
    zone_perf['score_gap'] = zone_perf['optimal_score'] - zone_perf['Zone_Score']
    zone_perf['score_gap_pct'] = zone_perf['score_gap'] / zone_perf['optimal_score']
    
    print(f"      ✅ Found optimal zones for {zone_perf['combo'].nunique()} combos")
    
    return zone_perf


def detect_reactive_pricing_simple(historical_df: pd.DataFrame,
                                   config: FoodserviceConfig) -> Dict:
    """
    Detect reactive pricing WITHOUT needing Last Invoice Date.
    Use fiscal weeks and volume trends instead.
    
    8th Grade Explanation:
    "Look for combos where:
    1. Volume was declining at a high zone
    2. We dropped the zone
    3. Volume recovered
    This means we were probably overpriced at the high zone."
    """
    
    print("\n   🔍 Checking for reactive pricing patterns...")
    
    df = historical_df.copy()
    
    # Need fiscal week and zone
    if 'Fiscal_Week' not in df.columns or 'Zone_Suffix_Numeric' not in df.columns:
        print("      ⚠️  Missing fiscal week or zone data")
        return {}
    
    # Ensure numeric
    df['Fiscal_Week'] = pd.to_numeric(df['Fiscal_Week'], errors='coerce')
    df['Zone_Suffix_Numeric'] = pd.to_numeric(df['Zone_Suffix_Numeric'], errors='coerce')
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    
    # Drop nulls
    df = df.dropna(subset=['Fiscal_Week', 'Zone_Suffix_Numeric'])
    
    # Sort by combo and week
    df = df.sort_values(['Company_Combo_Key', 'Fiscal_Week'])
    
    reactive_flags = {}
    
    for combo, g in df.groupby('Company_Combo_Key'):
        
        if len(g) < 12:  # Need at least 12 weeks of data
            continue
        
        # Get weekly aggregates
        weekly = g.groupby('Fiscal_Week').agg({
            'Zone_Suffix_Numeric': 'first',  # Assume one zone per week
            'Pounds_CY': 'sum'
        }).reset_index()
        
        # Look for zone drops
        weekly['zone_change'] = weekly['Zone_Suffix_Numeric'].diff()
        
        zone_drops = weekly[weekly['zone_change'] < 0]
        
        if zone_drops.empty:
            continue
        
        # Analyze first drop
        drop_idx = zone_drops.index[0]
        drop_week = weekly.loc[drop_idx, 'Fiscal_Week']
        
        # Get 6 weeks before drop
        pre_drop = weekly[weekly['Fiscal_Week'].between(drop_week - 6, drop_week - 1)]
        
        # Get 6 weeks after drop
        post_drop = weekly[weekly['Fiscal_Week'].between(drop_week, drop_week + 6)]
        
        if len(pre_drop) < 3 or len(post_drop) < 3:
            continue
        
        # Check if volume was declining before drop
        pre_trend = pre_drop['Pounds_CY'].iloc[-1] - pre_drop['Pounds_CY'].iloc[0]
        pre_trend_pct = pre_trend / pre_drop['Pounds_CY'].iloc[0] if pre_drop['Pounds_CY'].iloc[0] > 0 else 0
        
        # Check if volume recovered after drop
        post_trend = post_drop['Pounds_CY'].iloc[-1] - post_drop['Pounds_CY'].iloc[0]
        post_trend_pct = post_trend / post_drop['Pounds_CY'].iloc[0] if post_drop['Pounds_CY'].iloc[0] > 0 else 0
        
        # Reactive pattern: declining before, recovering after
        if pre_trend_pct < -0.15 and post_trend_pct > 0:
            from_zone = weekly.loc[drop_idx - 1, 'Zone_Suffix_Numeric']
            to_zone = weekly.loc[drop_idx, 'Zone_Suffix_Numeric']
            
            reactive_flags[combo] = {
                'reactive_downzone': True,
                'from_zone': int(from_zone),
                'to_zone': int(to_zone),
                'pre_decline_pct': pre_trend_pct,
                'post_recovery_pct': post_trend_pct,
                'likely_true_optimal': int(from_zone) - 1 if from_zone > 1 else 1,
                'stakeholder_message': (
                    f"⚠️ REACTIVE: Volume declined {abs(pre_trend_pct):.0%} at Zone {int(from_zone)}, "
                    f"then we dropped to Zone {int(to_zone)} and it recovered {post_trend_pct:.0%}. "
                    f"Real optimal is likely Zone {int(from_zone) - 1}."
                )
            }
    
    print(f"      ✅ Found {len(reactive_flags)} reactive pricing patterns")
    
    return reactive_flags

def create_last_purchase_week(df: pd.DataFrame, config: FoodserviceConfig) -> pd.DataFrame:
    """
    Calculate each customer's last purchase week for the target category.
    
    Simple explanation:
    "Find the most recent week each customer bought this category (Pounds CY > 0).
    That's their 'Last Purchase Week' for this Attribute Group."
    """
    
    print("\n   📅 Creating Last Purchase Week from fiscal week data...")
    
    df = df.copy()
    
    # Use existing Fiscal_Week_Combined if available, otherwise create Fiscal_Week_Numeric
    if 'Fiscal_Week_Combined' in df.columns:
        week_column = 'Fiscal_Week_Combined'
    elif 'Fiscal_Week' in df.columns:
        week_column = 'Fiscal_Week'
    else:
        df['Fiscal_Week_Numeric'] = pd.to_numeric(
            df[config.data_config.fiscal_week_column], 
            errors='coerce'
        )
        week_column = 'Fiscal_Week_Numeric'
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    
    # Get latest week in dataset
    max_week = df[week_column].max()
    if max_week > 10000:  # Combined format (YYYYWW)
        fy = int(max_week // 100)
        fw = int(max_week % 100)
        print(f"      Latest week in data: FY{fy} Week {fw} (YYYYWW: {int(max_week)})")
    else:  # Just week number
        print(f"      Latest week in data: Week {int(max_week)}")
    
    # For each customer, find their LAST purchase week (where Pounds CY > 0)
    active_purchases = df[df['Pounds_CY'] > 0].copy()
    
    last_purchase = active_purchases.groupby(
        config.data_config.customer_id_column
    )[week_column].max().reset_index()
    
    last_purchase.columns = [config.data_config.customer_id_column, 'Last_Purchase_Week']
    
    print(f"      ✅ Calculated last purchase week for {len(last_purchase):,} customers")
    
    # Merge back to main dataframe
    df = df.merge(
        last_purchase, 
        on=config.data_config.customer_id_column, 
        how='left'
    )
    
    # Calculate weeks since last purchase
    df['Weeks_Since_Last_Purchase'] = max_week - df['Last_Purchase_Week']
    
    # Fill NaN (customers with no purchases) with high number
    df['Weeks_Since_Last_Purchase'] = df['Weeks_Since_Last_Purchase'].fillna(999)
    
    print(f"      ✅ Weeks since last purchase calculated")
    
    return df


def classify_customer_activity_from_weeks(df: pd.DataFrame, 
                                          config: FoodserviceConfig) -> pd.DataFrame:
    """
    Classify customers using weeks since last purchase.
    
    Simple explanation:
    "Based on how many weeks since their last purchase:
    - 0-6 weeks ago = ACTIVE (still buying)
    - 7-8 weeks ago = LAPSED (stopped buying THIS category)
    - 9+ weeks ago = LOST (gone too long)"
    """
    
    print("\n   🏷️  Classifying customers by purchase recency...")
    
    df = df.copy()
    
    # Ensure we have the weeks column
    if 'Weeks_Since_Last_Purchase' not in df.columns:
        df = create_last_purchase_week(df, config)
    
    # Also check if they have ANY current year volume
    df['Has_CY_Volume'] = df['Pounds_CY'] > 0
    
    # Classification logic
    def classify(row):
        weeks_since = row['Weeks_Since_Last_Purchase']
        has_volume = row['Has_CY_Volume']
        
        # If they have current volume, they're active
        if has_volume:
            return 'ACTIVE_BUYER'
        
        # Based on weeks since last purchase
        if weeks_since <= 6:
            return 'ACTIVE_BUYER'  # Recent purchase
        elif weeks_since <= 8:
            return 'LAPSED_FROM_CATEGORY'  # 6-8 weeks = THE OPPORTUNITY!
        else:
            return 'LOST_CUSTOMER'  # 9+ weeks
    
    df['Customer_Status'] = df.apply(classify, axis=1)
    
    # Recovery potential
    df['Recovery_Potential'] = df['Customer_Status'].map({
        'ACTIVE_BUYER': 0,
        'LAPSED_FROM_CATEGORY': 10,  # HIGH PRIORITY
        'LOST_CUSTOMER': 2
    })
    
    # Count by status
    status_counts = df['Customer_Status'].value_counts()
    print(f"      Customer classification:")
    for status, count in status_counts.items():
        print(f"         • {status}: {count:,}")

    # Check if Last Invoice mapping worked
    if 'Weeks_Since_Any_Purchase' in df.columns:
        active_but_lapsed = df[
            (df['Customer_Status'] == 'LAPSED_FROM_CATEGORY') & 
            (df['Weeks_Since_Any_Purchase'] <= 8)
        ]
        print(f"      ✅ Last Invoice mapping worked")
        print(f"      🎯 Found {len(active_but_lapsed):,} rows where customer lapsed from category but still active")
    else:
        print(f"      ❌ Last Invoice Date NOT mapped - all customers may look lost!")

    return df
    

def track_customer_zone_journeys(df: pd.DataFrame, config: FoodserviceConfig) -> pd.DataFrame:
    """
    Track each customer's journey through zones over time.
    Uses FISCAL WEEKS instead of requiring Last Invoice Date column.
    """
    
    print("   🔄 Tracking customer zone journeys over historical data...")
    
    df = df.copy()
    
    # Required columns check (NO Last Invoice Date needed!)
    required = [
        config.data_config.customer_id_column,
        config.data_config.fiscal_week_column,
        'Zone_Suffix_Numeric',
        'Pounds_CY',
        'Company_Combo_Key'
    ]
    
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"      ❌ Missing required columns: {missing}")
        return pd.DataFrame()
    
    # Create Last Purchase Week if not already done
    if 'Last_Purchase_Week' not in df.columns:
        df = create_last_purchase_week(df, config)
    
    # Ensure numeric and sorted
    df['Fiscal_Week_Numeric'] = pd.to_numeric(df[config.data_config.fiscal_week_column], errors='coerce')
    df['Zone_Suffix_Numeric'] = pd.to_numeric(df['Zone_Suffix_Numeric'], errors='coerce')
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    
    df = df.dropna(subset=['Fiscal_Week_Numeric', 'Zone_Suffix_Numeric'])
    df = df.sort_values(['Company_Combo_Key', config.data_config.customer_id_column, 'Fiscal_Week_Numeric'])
    
    # Get latest week for recency calculations
    max_week = df['Fiscal_Week_Numeric'].max()
    
    # Track journeys
    journey_records = []
    
    grouped = df.groupby(['Company_Combo_Key', config.data_config.customer_id_column], dropna=False)
    total_customers = len(grouped)
    
    print(f"      • Analyzing {total_customers:,} unique customer-combo pairs...")
    
    for (combo, cust_id), cust_data in grouped:
        
        cust_data = cust_data.sort_values('Fiscal_Week_Numeric').reset_index(drop=True)
        
        if len(cust_data) < 2:
            # Single purchase - classify as stable
            journey_records.append({
                'combo': combo,
                'customer': cust_id,
                'customer_type': 'GREEN_FLAG_STABLE',
                'primary_zone': cust_data.iloc[0]['Zone_Suffix_Numeric'],
                'total_volume': cust_data.iloc[0]['Pounds_CY'],
                'weeks_active': 1,
                'weight': 1.0,
                'evidence': 'Single purchase - treated as stable'
            })
            continue
        
        # Extract arrays
        weeks = cust_data['Fiscal_Week_Numeric'].values
        zones = cust_data['Zone_Suffix_Numeric'].values
        pounds = cust_data['Pounds_CY'].values
        last_purchase_week = cust_data['Last_Purchase_Week'].max()
        
        # Check if customer ever changed zones
        unique_zones = set(zones)
        
        if len(unique_zones) == 1:
            # GREEN FLAG - STABLE: Never changed zones
            total_vol = pounds.sum()
            weeks_active = len(weeks)
            
            # Check consistency (buying regularly?)
            max_gap = max(np.diff(weeks)) if len(weeks) > 1 else 0
            is_consistent = max_gap <= 3  # No gap longer than 3 weeks
            
            journey_records.append({
                'combo': combo,
                'customer': cust_id,
                'customer_type': 'GREEN_FLAG_STABLE',
                'primary_zone': zones[0],
                'total_volume': total_vol,
                'weeks_active': weeks_active,
                'weight': 1.0,
                'consistency': 'HIGH' if is_consistent else 'MODERATE',
                'evidence': f'Stable at Zone {int(zones[0])} for {weeks_active} weeks'
            })
            
        else:
            # Customer experienced multiple zones - analyze transitions
            # (Keep existing zone change detection logic)
            zone_changes = []
            
            for i in range(1, len(zones)):
                if zones[i] != zones[i-1]:
                    # ZONE CHANGE DETECTED
                    change_week = weeks[i]
                    from_zone = zones[i-1]
                    to_zone = zones[i]
                    
                    # Look back 6 weeks before change
                    lookback_start = max(0, i - 6)
                    pre_change_weeks = weeks[lookback_start:i]
                    pre_change_pounds = pounds[lookback_start:i]
                    
                    # Were they lapsed? (8+ weeks gap OR 6 weeks of zero pounds)
                    weeks_since_last = weeks[i] - weeks[i-1] if i > 0 else 0
                    
                    was_lapsed = (
                        weeks_since_last >= 8 or
                        (len(pre_change_pounds[-3:]) >= 3 and all(p == 0 for p in pre_change_pounds[-3:]))
                    )
                    
                    # Calculate pre-change average volume
                    pre_avg_volume = pre_change_pounds[pre_change_pounds > 0].mean() if len(pre_change_pounds[pre_change_pounds > 0]) > 0 else 0
                    
                    # Look forward 4 weeks after change
                    lookforward_end = min(len(pounds), i + 4)
                    post_change_pounds = pounds[i:lookforward_end]
                    
                    came_back = any(p > 0 for p in post_change_pounds)
                    post_avg_volume = post_change_pounds[post_change_pounds > 0].mean() if len(post_change_pounds[post_change_pounds > 0]) > 0 else 0
                    
                    # Classify the transition
                    if to_zone < from_zone:  # Price went DOWN
                        if was_lapsed and came_back:
                            classification = 'GREEN_FLAG_REACTIVE_RECOVERY'
                            weight = 0.90
                            evidence = f"Lapsed at Zone {int(from_zone)}, moved to Zone {int(to_zone)}, came back"
                        
                        elif not was_lapsed and came_back:
                            if post_avg_volume > pre_avg_volume * 1.1:
                                classification = 'YELLOW_FLAG_PROACTIVE_GAINER'
                                weight = 0.75
                                evidence = f"Active at Zone {int(from_zone)}, moved to Zone {int(to_zone)}, volume increased"
                            else:
                                classification = 'RED_FLAG_PROACTIVE_FLAT'
                                weight = 0.0
                                evidence = f"Active at Zone {int(from_zone)}, moved to Zone {int(to_zone)}, volume stayed flat - MARGIN GIVEAWAY"
                        else:
                            classification = 'UNCLEAR'
                            weight = 0.5
                            evidence = f"Zone {int(from_zone)}→{int(to_zone)}, unclear pattern"
                    
                    else:  # Price went UP
                        classification = 'PRICE_INCREASE'
                        weight = 0.25
                        evidence = f"Zone increased {int(from_zone)}→{int(to_zone)}"
                    
                    zone_changes.append({
                        'change_week': change_week,
                        'from_zone': from_zone,
                        'to_zone': to_zone,
                        'classification': classification,
                        'weight': weight,
                        'evidence': evidence
                    })
            
            # Determine primary classification
            if zone_changes:
                best_change = max(zone_changes, key=lambda x: x['weight'])
                
                journey_records.append({
                    'combo': combo,
                    'customer': cust_id,
                    'customer_type': best_change['classification'],
                    'primary_zone': best_change['to_zone'],
                    'total_volume': pounds.sum(),
                    'weeks_active': len(weeks),
                    'weight': best_change['weight'],
                    'evidence': best_change['evidence'],
                    'zone_transitions': len(zone_changes)
                })
        
        # Check if currently lapsed (using weeks)
        weeks_since_purchase = max_week - last_purchase_week
        
        if weeks_since_purchase >= 8:
            # Still active buyer elsewhere? (they bought something recently overall)
            # This would need overall Last Invoice Date across all categories
            # For now, just flag as lapsed
            if journey_records:  # If we have records for this customer
                journey_records[-1]['customer_type'] = 'RED_FLAG_LAPSED_ACTIVE_ELSEWHERE'
                journey_records[-1]['alert'] = f"Lapsed from category {int(weeks_since_purchase)} weeks ago"
    
    journey_df = pd.DataFrame(journey_records)
    
    if not journey_df.empty:
        # Summarize patterns
        pattern_counts = journey_df['customer_type'].value_counts()
        print(f"      ✅ Journey analysis complete:")
        for pattern, count in pattern_counts.items():
            pct = (count / len(journey_df)) * 100
            print(f"         • {pattern}: {count:,} ({pct:.1f}%)")
    
    return journey_df

def calculate_lapse_penalty(df: pd.DataFrame, config: FoodserviceConfig) -> pd.DataFrame:
    """
    Calculate lapse rate penalty for each combo + zone.
    
    Lapse rate = % of customers who:
    - Stopped buying THIS CATEGORY (max fiscal week with Pounds CY > 0, 8+ weeks ago)
    - BUT still active with company (Last Invoice Date is recent - buying OTHER stuff)
    
    This is a SIGNAL OF OVERPRICING.
    
    Returns: DataFrame with combo, zone, lapse_rate, lapse_penalty
    """
    
    print("   🚨 Calculating lapse rate penalties...")
    
    df = df.copy()
    
    # Ensure we have required columns
    if 'Fiscal_Week' not in df.columns:
        print("      ⚠️  Fiscal_Week not found - skipping lapse penalty")
        return pd.DataFrame()
    
    if 'Pounds_CY' not in df.columns:
        print("      ⚠️  Pounds_CY not found - skipping lapse penalty")
        return pd.DataFrame()
    
    # Get latest week overall
    max_week = df['Fiscal_Week'].max()
    
    # For each customer + combo + zone, find their last purchase week in THIS category
    # Only count weeks where they actually bought something (Pounds CY > 0)
    category_purchases = df[df['Pounds_CY'] > 0].copy()
    
    if category_purchases.empty:
        print("      ⚠️  No purchases with Pounds_CY > 0 - skipping lapse penalty")
        return pd.DataFrame()
    
    customer_zones = category_purchases.groupby([
        'Company_Combo_Key', 
        config.data_config.zone_column,
        config.data_config.customer_id_column
    ], dropna=False).agg({
        'Fiscal_Week': 'max',  # Last week they bought THIS category (with volume > 0)
    }).reset_index()
    
    customer_zones.columns = ['combo', 'zone', 'customer', 'last_category_week']
    
    # Get Last Invoice Date for each customer (from full df, not filtered)
    if 'Last_Invoice_Date' in df.columns:
        last_invoice = df.groupby(config.data_config.customer_id_column)[
            'Last_Invoice_Date'
        ].max().reset_index()
        last_invoice.columns = ['customer', 'last_any_purchase_date']
        
        customer_zones = customer_zones.merge(last_invoice, on='customer', how='left')
        has_invoice_date = True
    else:
        print("      ⚠️  Last Invoice Date column not found - using fiscal week only")
        customer_zones['last_any_purchase_date'] = None
        has_invoice_date = False
    
    # Calculate weeks since category purchase
    customer_zones['weeks_since_category'] = max_week - customer_zones['last_category_week']
    
    # Calculate days since ANY purchase (if we have Last Invoice Date)
    if has_invoice_date:
        customer_zones['last_any_purchase_date'] = pd.to_datetime(
            customer_zones['last_any_purchase_date'], 
            errors='coerce'
        )
        
        # Use current date as reference
        current_date = pd.Timestamp.now()
        customer_zones['days_since_any_purchase'] = (
            current_date - customer_zones['last_any_purchase_date']
        ).dt.days
    else:
        # Fallback: use weeks since category purchase
        customer_zones['days_since_any_purchase'] = customer_zones['weeks_since_category'] * 7
    
    # Classify: LAPSED FROM CATEGORY but STILL ACTIVE buying other stuff
    customer_zones['is_lapsed_but_active'] = (
        (customer_zones['weeks_since_category'] >= 8) &  # Haven't bought THIS category in 8+ weeks
        (customer_zones['days_since_any_purchase'] <= 30)  # BUT bought OTHER stuff recently (within 30 days)
    )
    
    # Calculate lapse rate per combo + zone
    lapse_rates = customer_zones.groupby(['combo', 'zone']).agg({
        'customer': 'count',
        'is_lapsed_but_active': 'sum'
    }).reset_index()
    
    lapse_rates.columns = ['combo', 'zone', 'total_customers', 'lapsed_but_active']
    
    lapse_rates['lapse_rate'] = (
        lapse_rates['lapsed_but_active'] / lapse_rates['total_customers'].replace(0, 1)
    )
    
    # Calculate penalty (scaled by severity)
    lapse_rates['lapse_penalty'] = lapse_rates['lapse_rate'] * 50000
    
    print(f"      ✅ Calculated lapse penalties for {len(lapse_rates):,} combo-zone pairs")
    
    # Show worst offenders
    worst = lapse_rates.nlargest(5, 'lapse_rate')
    if not worst.empty:
        print(f"      ⚠️  Highest lapse rates:")
        for _, row in worst.iterrows():
            if row['lapse_rate'] > 0:
                print(f"         • {row['combo']} Zone {int(row['zone'])}: {row['lapse_rate']:.1%} lapsed but active elsewhere")
    
    return lapse_rates[['combo', 'zone', 'lapse_rate', 'lapse_penalty']]

def calculate_weighted_zone_scores(journey_df: pd.DataFrame, 
                                   config: FoodserviceConfig,
                                   lapse_penalty_df: pd.DataFrame = None) -> pd.DataFrame:  # ✅ Add parameter
    """
    Calculate zone performance scores using weighted customer types.
    
    WITH MARGIN (has_margin_data=True):
    - Profit dollars (volume × margin): 50%
    - Customer count: 30%
    - Consistency: 20%
    
    WITHOUT MARGIN (has_margin_data=False):
    - Weighted volume: 60%
    - Customer count: 30%
    - Consistency: 10%
    """
    
    print("   📊 Calculating weighted zone scores...")
    
    if journey_df.empty:
        return pd.DataFrame()
    
    # Group by combo and primary zone
    zone_scores = journey_df.groupby(['combo', 'primary_zone'], dropna=False).agg({
        'customer': 'nunique',
        'total_volume': 'sum',
        'weight': 'mean',
        'weeks_active': 'sum'
    }).reset_index()
    
    zone_scores.columns = ['combo', 'zone', 'customer_count', 'total_volume', 
                           'avg_weight', 'total_weeks_active']
    
    # Calculate weighted volume
    zone_scores['weighted_volume'] = zone_scores['total_volume'] * zone_scores['avg_weight']

    # ADD LAPSE RATE PENALTY (NEW!)
    if lapse_penalty_df is not None and not lapse_penalty_df.empty:
        zone_scores = zone_scores.merge(
            lapse_penalty_df,
            on=['combo', 'zone'],
            how='left'
        )
        zone_scores['lapse_penalty'] = zone_scores['lapse_penalty'].fillna(0)
        zone_scores['lapse_rate'] = zone_scores['lapse_rate'].fillna(0)
    else:
        zone_scores['lapse_penalty'] = 0
        zone_scores['lapse_rate'] = 0
 
    # Add margin data if available
    if config.data_config.has_margin_data and 'avg_margin_per_lb' in journey_df.columns:
        # Get average margin for this zone from journey data
        margin_data = journey_df.groupby(['combo', 'primary_zone'])['avg_margin_per_lb'].mean().reset_index()
        margin_data.columns = ['combo', 'zone', 'avg_margin_per_lb']
        
        zone_scores = zone_scores.merge(margin_data, on=['combo', 'zone'], how='left')
        zone_scores['avg_margin_per_lb'] = zone_scores['avg_margin_per_lb'].fillna(0)
        
        # Calculate weighted profit
        zone_scores['weighted_profit'] = zone_scores['weighted_volume'] * zone_scores['avg_margin_per_lb']
        
        # SCORE WITHOUT MARGIN
        zone_scores['Zone_Score'] = (
            zone_scores['weighted_volume'] * 0.6 +
            zone_scores['customer_count'] * 100 * 0.3 +
            (zone_scores['total_weeks_active'] / zone_scores['customer_count'].replace(0, 1)) * 10 * 0.1
            - zone_scores['lapse_penalty']  # ✅ SUBTRACT PENALTY!
        )
        
        print(f"      ✅ Scored {zone_scores['combo'].nunique():,} combos WITH margin data")
        
    else:
        # SCORE WITHOUT MARGIN
        zone_scores['Zone_Score'] = (
            zone_scores['weighted_volume'] * 0.6 +
            zone_scores['customer_count'] * 100 * 0.3 +
            (zone_scores['total_weeks_active'] / zone_scores['customer_count'].replace(0, 1)) * 10 * 0.1
        )
        
        zone_scores['avg_margin_per_lb'] = None
        zone_scores['weighted_profit'] = None
        
        print(f"      ⚠️  Scored {zone_scores['combo'].nunique():,} combos WITHOUT margin")
    
    # Get customer type breakdown per zone
    type_breakdown = journey_df.groupby(['combo', 'primary_zone', 'customer_type'], dropna=False).size().reset_index(name='count')
    
    type_pivot = type_breakdown.pivot_table(
        index=['combo', 'primary_zone'],
        columns='customer_type',
        values='count',
        fill_value=0
    ).reset_index()
    
    zone_scores = zone_scores.merge(
        type_pivot,
        left_on=['combo', 'zone'],
        right_on=['combo', 'primary_zone'],
        how='left'
    )
    
    # Find optimal zone per combo
    idx = zone_scores.groupby('combo')['Zone_Score'].idxmax()
    optimal_zones = zone_scores.loc[idx, ['combo', 'zone', 'Zone_Score']].rename(columns={'zone': 'optimal_zone'})
    
    zone_scores = zone_scores.merge(optimal_zones[['combo', 'optimal_zone']], on='combo', how='left')
    
    return zone_scores

def calculate_yoy_customer_metrics(df: pd.DataFrame, 
                                   config: FoodserviceConfig) -> Dict:
    """
    Calculate year-over-year distinct customer metrics using configurable lookback window.
    
    Simple explanation:
    "Compare the last X weeks this year vs. the same X weeks last year. 
    How many distinct customers are we keeping?"
    
    Example: If lookback = 8 weeks and we're at Week 52:
    - This Year: Weeks 45-52 (last 8 weeks)
    - Last Year: Weeks 45-52 (same period last year)
    """
    
    df = df.copy()

    # DEBUG: Check what we received
    print(f"   🔍 DEBUG: DataFrame has {len(df):,} rows")
    print(f"   🔍 DEBUG: Columns available: {list(df.columns[:10])}...")  # First 10 columns
    print(f"   🔍 DEBUG: Has Fiscal_Week_Combined? {'Fiscal_Week_Combined' in df.columns}")
    if 'Fiscal_Week_Combined' in df.columns:
        print(f"   🔍 DEBUG: Week range: {df['Fiscal_Week_Combined'].min()}-{df['Fiscal_Week_Combined'].max()}")
        
    # Ensure we have fiscal week data
    if 'Fiscal_Week' not in df.columns:
        print("   ⚠️  Warning: No fiscal week column, can't calculate YoY with lookback")
        return {}
    
    # Ensure numeric
    df['Fiscal_Week'] = pd.to_numeric(df['Fiscal_Week'], errors='coerce')
    df = df.dropna(subset=['Fiscal_Week'])
    df['Fiscal_Week'] = df['Fiscal_Week'].astype(int)
    
    # Ensure we have Pounds columns
    if 'Pounds_CY' not in df.columns or 'Pounds_PY' not in df.columns:
        print("   ⚠️  Warning: Missing Pounds_CY or Pounds_PY column")
        return {}
    
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    df['Pounds_PY'] = pd.to_numeric(df['Pounds_PY'], errors='coerce').fillna(0)
    
    # Use COMBINED fiscal week (YYYYWW)
    if 'Fiscal_Week_Combined' in df.columns:
        latest_week_combined = df['Fiscal_Week_Combined'].max()  # Gets 202613
        lookback_weeks = config.YOY_LOOKBACK_WEEKS
        cutoff_week = latest_week_combined - lookback_weeks + 1  # 202613 - 8 + 1 = 202606

        print(f"   🔍 DEBUG: cutoff_week = {cutoff_week}, latest = {latest_week_combined}")
        print(f"   🔍 DEBUG: Fiscal_Week_Combined dtype = {df['Fiscal_Week_Combined'].dtype}")
        print(f"   📅 Comparing Weeks {cutoff_week}-{latest_week_combined} (last {lookback_weeks} weeks)")
        # Filter to lookback window
        recent_data = df[df['Fiscal_Week_Combined'] >= cutoff_week].copy()
    else:
        # Fallback if combined column doesn't exist
        latest_week = df['Fiscal_Week'].max()
        lookback_weeks = config.YOY_LOOKBACK_WEEKS
        cutoff_week = latest_week - lookback_weeks + 1
        recent_data = df[df['Fiscal_Week'] >= cutoff_week].copy()                 
    if recent_data.empty:
        print("   ⚠️  No data in lookback window")
        return {}
    
    # Identify distinct customers by year (in the lookback window)
    cy_customers = recent_data[recent_data['Pounds_CY'] > 0][config.data_config.customer_id_column].unique()
    py_customers = recent_data[recent_data['Pounds_PY'] > 0][config.data_config.customer_id_column].unique()
    
    # Calculate overlaps
    retained_customers = set(cy_customers) & set(py_customers)
    new_customers = set(cy_customers) - set(py_customers)
    lost_customers = set(py_customers) - set(cy_customers)
    
    # Overall metrics
    overall = {
        'lookback_weeks': lookback_weeks,
        'week_range': f"{cutoff_week}-{cutoff_week + lookback_weeks - 1}",
        'distinct_customers_cy': len(cy_customers),
        'distinct_customers_py': len(py_customers),
        'customer_change': len(cy_customers) - len(py_customers),
        'customer_change_pct': ((len(cy_customers) - len(py_customers)) / len(py_customers) * 100) if len(py_customers) > 0 else 0,
        'retained_customers': len(retained_customers),
        'retention_rate': (len(retained_customers) / len(py_customers) * 100) if len(py_customers) > 0 else 0,
        'new_customers': len(new_customers),
        'lost_customers': len(lost_customers),
        'customer_status_summary': {
            'RETAINED': len(retained_customers),
            'NEW': len(new_customers),
            'LOST': len(lost_customers)
        }
    }
    
    # By combo breakdown
    combo_metrics = []
    
    for combo in recent_data['Company_Combo_Key'].unique():
        combo_data = recent_data[recent_data['Company_Combo_Key'] == combo]
        
        combo_cy = combo_data[combo_data['Pounds_CY'] > 0][config.data_config.customer_id_column].unique()
        combo_py = combo_data[combo_data['Pounds_PY'] > 0][config.data_config.customer_id_column].unique()
        
        combo_retained = set(combo_cy) & set(combo_py)
        combo_new = set(combo_cy) - set(combo_py)
        combo_lost = set(combo_py) - set(combo_cy)
        
        combo_metrics.append({
            'Company_Combo_Key': combo,
            'distinct_customers_cy': len(combo_cy),
            'distinct_customers_py': len(combo_py),
            'customer_change': len(combo_cy) - len(combo_py),
            'customer_change_pct': ((len(combo_cy) - len(combo_py)) / len(combo_py) * 100) if len(combo_py) > 0 else 0,
            'retained_customers': len(combo_retained),
            'retention_rate': (len(combo_retained) / len(combo_py) * 100) if len(combo_py) > 0 else 0,
            'new_customers': len(combo_new),
            'lost_customers': len(combo_lost)
        })
    
    overall['by_combo'] = pd.DataFrame(combo_metrics)
    
    return overall

def apply_filters(df: pd.DataFrame, config: FoodserviceConfig) -> pd.DataFrame:
    """Apply Company Number and/or Region ID filters."""
    
    df = df.copy()
    original_count = len(df)
    
    # Filter by Company Number
    if config.FILTER_BY_company_number:  # ✅ Match your class attribute
        if config.data_config.company_number_column in df.columns:
            # Strip leading zeros for comparison
            df_stripped = df[config.data_config.company_number_column].astype(str).str.strip().str.lstrip('0')
            filter_stripped = str(config.FILTER_BY_company_number).lstrip('0')
            df = df[df_stripped == filter_stripped]
            print(f"   🔍 Filtered to Company Number '{config.FILTER_BY_company_number}': {len(df):,} rows (was {original_count:,})")
        else:
            print(f"   ⚠️  Warning: Company Number column not found")
    
    # Filter by Region ID
    if config.FILTER_BY_company_region_id:  # ✅ Match your class attribute
        if config.data_config.company_region_id_column in df.columns:
            df = df[df[config.data_config.company_region_id_column].astype(str) == str(config.FILTER_BY_company_region_id)]
            print(f"   🔍 Filtered to Region ID '{config.FILTER_BY_company_region_id}': {len(df):,} rows")
        else:
            print(f"   ⚠️  Warning: Region ID column not found")
    
    if df.empty:
        print(f"   ❌ WARNING: Filters removed all data!")
    
    return df
# ============================================================================
# REACTIVE PRICING DETECTION
# ============================================================================

def detect_reactive_pricing_failures(df: pd.DataFrame, 
                                    config: FoodserviceConfig = None) -> Dict:
    """
    Find combos where we overpriced, lost customers, then dropped zone reactively.
    
    8th Grade Explanation:
    "Sometimes we charge too much, restaurants stop buying, then we panic and drop 
    the price. The lower zone looks good, but only because we already killed the 
    business at the higher zone."
    """
    if config is None:
        config = FoodserviceConfig()
    
    # Need customer status first
    if 'Customer_Status' not in df.columns:
        df = classify_customer_activity(df, config)
    
    flags = {}
    
    for combo, g in df.groupby('Company_Combo_Key', dropna=False):
        g = g.sort_values(config.data_config.fiscal_week_column)
        
        # Must have zone info
        if config.data_config.zone_column not in g.columns:
            continue
            
        g[config.data_config.zone_column] = pd.to_numeric(
            g[config.data_config.zone_column], 
            errors='coerce'
        )
        g = g.dropna(subset=[config.data_config.zone_column])
        
        if len(g) < 2:
            continue
        
        # Find first zone drop
        zone_diff = g[config.data_config.zone_column].diff()
        zone_drops = g[zone_diff < 0]
        
        if zone_drops.empty:
            continue
        
        # Analyze first drop
        first_drop = zone_drops.iloc[0]
        drop_week = first_drop[config.data_config.fiscal_week_column]
        drop_idx = g[g[config.data_config.fiscal_week_column] == drop_week].index[0]
        
        # Get zones before and after
        pre_drop_data = g.loc[g.index < drop_idx]
        if len(pre_drop_data) == 0:
            continue
            
        pre_drop_zone = pre_drop_data[config.data_config.zone_column].iloc[-1]
        post_drop_zone = first_drop[config.data_config.zone_column]
        
        # Get pre-drop window
        lookback = config.REACTIVE_LOOKBACK_WEEKS
        pre_window = g[
            g[config.data_config.fiscal_week_column].between(
                drop_week - lookback, 
                drop_week - 1
            )
        ]
        
        if pre_window.empty:
            continue
        
        # KEY METRIC: How many customers lapsed BEFORE the zone drop?
        status_counts = pre_window['Customer_Status'].value_counts()
        total_customers = pre_window[config.data_config.customer_id_column].nunique()
        
        lapsed_count = status_counts.get('LAPSED_FROM_CATEGORY', 0)
        lapsed_pct = (lapsed_count / total_customers) if total_customers > 0 else 0
        
        # If 30%+ customers lapsed before drop, it's reactive!
        if lapsed_pct >= 0.30:
            
            # Measure recovery after drop
            post_window = g[
                g[config.data_config.fiscal_week_column].between(
                    drop_week, 
                    drop_week + lookback
                )
            ]
            
            pre_volume = pre_window['Pounds_CY'].sum()
            post_volume = post_window['Pounds_CY'].sum()
            recovery_pct = ((post_volume - pre_volume) / pre_volume) if pre_volume > 0 else 0
            
            # Calculate likely true optimal
            likely_optimal = int(pre_drop_zone) - 1
            if likely_optimal < 1:
                likely_optimal = 1
            
            flags[combo] = {
                'reactive_downzone': True,
                'from_zone': int(pre_drop_zone),
                'to_zone': int(post_drop_zone),
                'customers_lapsed_before_drop': int(lapsed_count),
                'lapsed_percentage': lapsed_pct,
                'volume_recovery_rate': recovery_pct,
                'likely_true_optimal': likely_optimal,
                'stakeholder_message': (
                    f"⚠️ REACTIVE: {int(lapsed_count)} customers stopped buying "
                    f"at Zone {int(pre_drop_zone)}. We dropped to Zone {int(post_drop_zone)} "
                    f"to recover them. Real optimal is likely Zone {likely_optimal}."
                ),
                'trust_level': 'MEDIUM'
            }
    
    return flags

def load_fiscal_calendar(calendar_path: str) -> pd.DataFrame:
    """Load and prepare fiscal calendar for date-to-week conversion."""
    cal = pd.read_csv(calendar_path)
    
    # Parse dates
    cal['week_start_date'] = pd.to_datetime(cal['week start date'], errors='coerce')
    cal['week_end_date'] = pd.to_datetime(cal['week end date'], errors='coerce')
    
    # Create combined fiscal week ID (YYYYWW)
    cal['fiscal_week_combined'] = cal['fiscal year'] * 100 + cal['fiscal week number']
    
    return cal[['fiscal_week_combined', 'fiscal year', 'fiscal week number', 
                'week_start_date', 'week_end_date']]


def map_date_to_fiscal_week(dates: pd.Series, fiscal_calendar: pd.DataFrame) -> pd.Series:
    """
    Convert calendar dates to fiscal week numbers.
    
    Args:
        dates: Series of datetime dates
        fiscal_calendar: DataFrame from load_fiscal_calendar()
    
    Returns:
        Series of fiscal week numbers (1-52)
    """
    
    result = pd.Series(index=dates.index, dtype='Int64')
    
    for idx, date in dates.items():
        if pd.isna(date):
            result[idx] = pd.NA
            continue
        
        # Find which fiscal week this date falls into
        mask = (fiscal_calendar['week_start_date'] <= date) & (date <= fiscal_calendar['week_end_date'])
        matching_weeks = fiscal_calendar[mask]
        
        if len(matching_weeks) > 0:
            result[idx] = matching_weeks.iloc[0]['fiscal week number']
        else:
            result[idx] = pd.NA
    
    return result


def classify_customer_activity_from_weeks(df: pd.DataFrame, 
                                          config: FoodserviceConfig,
                                          fiscal_calendar: pd.DataFrame = None) -> pd.DataFrame:
    """
    Classify customers using fiscal weeks consistently.
    
    Simple explanation:
    "Based on fiscal weeks since their last purchase:
    - 0-6 weeks ago = ACTIVE (still buying)
    - 7-8 weeks ago but bought OTHER stuff = LAPSED (THE OPPORTUNITY!)
    - 9+ weeks ago = LOST (gone too long)"
    """
    
    print("\n   🏷️  Classifying customers by purchase recency...")
    
    df = df.copy()
    
    # Ensure we have the weeks column for THIS category
    if 'Weeks_Since_Last_Purchase' not in df.columns:
        df = create_last_purchase_week(df, config)
    
    df['Has_CY_Volume'] = df['Pounds_CY'] > 0
    
    # Check if Last Invoice Date exists and convert to fiscal week
    has_last_invoice = 'Last_Invoice_Date' in df.columns
    
    if has_last_invoice and fiscal_calendar is not None:
        print("      📅 Converting Last Invoice Date to fiscal weeks...")
        
        df['Last_Invoice_Date'] = pd.to_datetime(
            df['Last_Invoice_Date'],
            errors='coerce'
        )
        
        # Map Last Invoice Date to fiscal week
        df['Last_Invoice_Fiscal_Week'] = map_date_to_fiscal_week(
            df['Last_Invoice_Date'], 
            fiscal_calendar
        )
        
        # Calculate weeks since ANY purchase (using fiscal weeks)
        max_week = df['Fiscal_Week'].max()
        df['Weeks_Since_Any_Purchase'] = max_week - df['Last_Invoice_Fiscal_Week']
        df['Weeks_Since_Any_Purchase'] = df['Weeks_Since_Any_Purchase'].fillna(999)
        
        print(f"      ✅ Mapped Last Invoice Date to fiscal weeks")
    
    # Classification logic
    def classify(row):
        weeks_since_category = row['Weeks_Since_Last_Purchase']
        has_volume = row['Has_CY_Volume']
        
        # If they have current volume, they're active
        if has_volume:
            return 'ACTIVE_BUYER'
        
        # Use Last Invoice fiscal week if available
        if has_last_invoice and fiscal_calendar is not None:
            weeks_since_any = row.get('Weeks_Since_Any_Purchase', 999)
            
            if weeks_since_category >= 8:  # 8+ weeks since THIS category
                if weeks_since_any <= 8:  # But bought OTHER stuff within 8 weeks
                    return 'LAPSED_FROM_CATEGORY'  # THE OPPORTUNITY!
                else:
                    return 'LOST_CUSTOMER'  # Gone entirely
            else:
                return 'ACTIVE_BUYER'
        else:
            # Fallback if no Last Invoice Date
            if weeks_since_category <= 6:
                return 'ACTIVE_BUYER'
            elif weeks_since_category <= 8:
                return 'LAPSED_FROM_CATEGORY'
            else:
                return 'LOST_CUSTOMER'
    
    df['Customer_Status'] = df.apply(classify, axis=1)
    
    # Recovery potential
    df['Recovery_Potential'] = df['Customer_Status'].map({
        'ACTIVE_BUYER': 0,
        'LAPSED_FROM_CATEGORY': 10,  # HIGH PRIORITY
        'LOST_CUSTOMER': 2
    })
    
    # Count by status
    status_counts = df['Customer_Status'].value_counts()
    print(f"      Customer classification:")
    for status, count in status_counts.items():
        print(f"         • {status}: {count:,}")
    
    return df

# ============================================================================
# ZONE OPTIMIZATION ENGINE
# ============================================================================

class FoodserviceZoneEngine:
    """Enhanced with purchase consistency tracking."""
    
    def __init__(self, 
                 config: FoodserviceConfig,
                 learning_engine: Optional[LearningEngine] = None):
        self.config = config
        self.learning_engine = learning_engine
        self.reactive_flags = {}
        self.customer_analysis = None
        self.consistency_analysis = None 
        self.yoy_customer_metrics = None
    
    def analyze_current_state(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare current state data with all needed features."""
        
        df = df.copy()
        
        # STEP 0: Load fiscal calendar
        calendar_path = os.path.join(
            r"C:\Users\kmor6669\OneDrive - Sysco Corporation\Desktop\Pricing",
            "calendar.csv"
        )
        fiscal_calendar = load_fiscal_calendar(calendar_path)
        
        # STEP 1: Create Last Purchase Week from fiscal week data
        df = create_last_purchase_week(df, self.config)
        
        # STEP 2: Classify customers by recency (NOW WITH FISCAL CALENDAR)
        df = classify_customer_activity_from_weeks(df, self.config, fiscal_calendar)
        
        # STEP 3: Calculate purchase consistency (THE GREEN FLAG!)
        self.consistency_analysis = calculate_purchase_consistency(df, self.config)
        print(f"   🟢 Calculated consistency for {len(self.consistency_analysis)} zone combinations")
        
        # STEP 4: Detect reactive pricing patterns (will use historical in generate_recommendations)
        self.reactive_flags = {}
        
        # STEP 5: Store for later use
        self.customer_analysis = self._build_customer_summary(df)
        
        return df  # ← CRITICAL: Must return the dataframe!
        
    
    def _build_customer_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aggregate customer metrics by combo and zone."""
        
        grouping_cols = self.config.data_config.get_grouping_columns()
        grouping_cols.append(self.config.data_config.zone_column)
        
        # ✅ FILTER to only columns that exist in the DataFrame
        grouping_cols = [col for col in grouping_cols if col in df.columns]
        
        # Check if we have required columns
        agg_dict = {
            self.config.data_config.customer_id_column: 'nunique',
        }
        
        # Only add columns that exist
        if 'Customer_Status' in df.columns:
            agg_dict['Customer_Status'] = lambda x: (x == 'ACTIVE_BUYER').sum()
        if 'Recovery_Potential' in df.columns:
            agg_dict['Recovery_Potential'] = 'sum'
        if 'Pounds_CY' in df.columns:
            agg_dict['Pounds_CY'] = 'sum'
        if 'Pounds_PY' in df.columns:
            agg_dict['Pounds_PY'] = 'sum'
        
        summary = df.groupby(grouping_cols, dropna=False).agg(agg_dict).reset_index()
        
        # Build column names dynamically based on what we actually aggregated
        new_columns = list(grouping_cols)
        new_columns.append('Total_Customers')
        
        if 'Customer_Status' in df.columns:
            new_columns.append('Active_Customers')
        if 'Recovery_Potential' in df.columns:
            new_columns.append('Recovery_Potential_Score')
        if 'Pounds_CY' in df.columns:
            new_columns.append('Volume_CY')
        if 'Pounds_PY' in df.columns:
            new_columns.append('Volume_PY')
        
        summary.columns = new_columns
        
        # Calculate derived metrics only if we have the required columns
        if 'Active_Customers' in summary.columns:
            summary['Lapsed_Customers'] = summary['Total_Customers'] - summary['Active_Customers']
            summary['Lapsed_Pct'] = summary['Lapsed_Customers'] / summary['Total_Customers']
        
        if 'Volume_CY' in summary.columns and 'Volume_PY' in summary.columns:
            summary['Volume_Change'] = summary['Volume_CY'] - summary['Volume_PY']
            summary['Volume_Change_Pct'] = summary['Volume_Change'] / summary['Volume_PY'].replace(0, 1)
        
        return summary
        
    def generate_recommendations(self, current_df: pd.DataFrame, 
                                historical_df: Optional[pd.DataFrame] = None) -> List[Dict]:
        """Generate recommendations using historical analysis."""
        # ✅ ADD THIS:
        zone_1_to_0_flags = []  # Track combos that want to move from 1→0
        # Analyze current state
        current_df = self.analyze_current_state(current_df)
        
        # Track journeys and get zone scores from historical data
        journey_df = pd.DataFrame()
        zone_scores = pd.DataFrame()
        
        if historical_df is not None:
            print("\n🔄 Analyzing historical zone performance...")
            
            # Track customer journeys
            journey_df = track_customer_zone_journeys(historical_df, self.config)
            
            if not journey_df.empty:
                # Calculate lapse penalties
                lapse_penalty_df = calculate_lapse_penalty(historical_df, self.config)
                # Calculate weighted zone scores WITH penalties
                zone_scores = calculate_weighted_zone_scores(journey_df, self.config, lapse_penalty_df)
            else:
                # Fallback to simple scoring if journey tracking fails
                print("      ⚠️  Journey tracking empty, using simple scoring")
                zone_scores = calculate_simple_zone_scores(historical_df, self.config)
            
            # Calculate YoY customer metrics (MOVED OUTSIDE if/else!)
            print("\n📊 Calculating YoY customer metrics...")
            self.yoy_customer_metrics = calculate_yoy_customer_metrics(historical_df, self.config)
            
            # Detect reactive pricing
            self.reactive_flags = detect_reactive_pricing_simple(historical_df, self.config)
        
        recommendations = []
        
        # Process each combo
        for combo in current_df['Company_Combo_Key'].unique():
            
            combo_data = current_df[current_df['Company_Combo_Key'] == combo]
            
            # Get journey data for THIS combo
            combo_journey = journey_df[journey_df['combo'] == combo] if not journey_df.empty else pd.DataFrame()
            
            # 🔍 DEBUG
            print(f"\n🔍 Combo: {combo}")
            
            # Get current zone
            zone_mode = combo_data['Zone_Suffix_Numeric'].mode()
            current_zone = zone_mode.iloc[0] if len(zone_mode) > 0 else 5
            print(f"   Current Zone: {int(current_zone)}")

            # ✅ ADD THIS: Look up consistency metrics for current zone
            consistent_buyers_count = 0
            green_flag_rate = 0.0

            if self.consistency_analysis is not None and not self.consistency_analysis.empty:
                consistency_row = self.consistency_analysis[
                    (self.consistency_analysis['Company_Combo_Key'] == combo) & 
                    (self.consistency_analysis['Zone'] == current_zone)
                ]
                
                if not consistency_row.empty:
                    consistent_buyers_count = int(consistency_row.iloc[0]['consistent_buyers'])
                    green_flag_rate = float(consistency_row.iloc[0]['green_flag_rate'])
                    print(f"   Consistency: {consistent_buyers_count} consistent buyers ({green_flag_rate:.1%} green flag rate)")

            # ✅ NEW SECTION (add after line ~2037):
            # High consistency = sticky customers = be MORE cautious about changing zone
            if consistent_buyers_count >= 10 and green_flag_rate >= 0.60:
                print(f"   🔒 HIGH CONSISTENCY ZONE: {consistent_buyers_count} sticky buyers ({green_flag_rate:.1%}) - raising bar for moves")
                required_lapse_rate = 0.40  # Need 40% lapsed before we'll move a sticky zone
                high_consistency_zone = True
            else:
                required_lapse_rate = 0.30  # Normal 30% threshold
                high_consistency_zone = False

            # Line ~2040: Get optimal zone from scoring
            optimal_zone = current_zone  # Default
            zone_score_value = 0

            if not zone_scores.empty:
                combo_score = zone_scores[zone_scores['combo'] == combo]
                if not combo_score.empty:
                    optimal_zone = combo_score.iloc[0]['optimal_zone']
                    zone_score_value = combo_score.iloc[0]['Zone_Score']
                    print(f"   Optimal Zone: {int(optimal_zone)} (Score: {zone_score_value:.2f})")
                else:
                    print(f"   ⚠️  WARNING: No score for this combo in zone_scores")
            else:
                print(f"   ⚠️  WARNING: No zone scores available!")

            # Check for Zone 1→0 moves (FLAG but don't recommend)
            if current_zone == 1 and optimal_zone == 0:
                print(f"   🚩 FLAGGED: Zone 1→0 move detected (will not recommend)")
                
                # ✅ ADD TO FLAGS:
                zone_1_to_0_flags.append({
                    'company_combo': combo,
                    'company_name': combo_data[self.config.data_config.company_column].iloc[0],
                    'attribute_group': combo_data.get(self.config.data_config.attribute_group_column, pd.Series(['N/A'])).iloc[0],
                    'current_zone': 1,
                    'suggested_zone': 0,
                    'total_customers': combo_data[self.config.data_config.customer_id_column].nunique(),
                    'total_volume': combo_data['Pounds_CY'].sum(),
                    'reason': 'Historical data suggests Zone 0, but policy prevents 1→0 moves'
                })
                
                continue  # Skip to next combo
            
            # Get customer metrics
            total_customers = combo_data[self.config.data_config.customer_id_column].nunique()
            customer_status = combo_data.groupby(
                self.config.data_config.customer_id_column
            )['Customer_Status'].first()
            
            active_customers = (customer_status == 'ACTIVE_BUYER').sum()
            lapsed_customers = (customer_status == 'LAPSED_FROM_CATEGORY').sum()
            lapsed_pct = lapsed_customers / total_customers if total_customers > 0 else 0
            
            total_volume = combo_data['Pounds_CY'].sum()

            # Skip if no volume AND no lapsed customers to recover
            if total_volume == 0 and lapsed_customers == 0:
                print(f"   ⏭️  SKIPPED: No volume and no customers to recover")
                continue

            print(f"   Customers: {total_customers} total, {active_customers} active, {lapsed_customers} lapsed ({lapsed_pct:.1%})")
            print(f"   Volume: {total_volume:,.0f} lbs")
            
            # ✅ IMPROVED: Track BOTH customer counts AND volume changes
            customers_growing = 0
            customers_declining = 0
            customers_active = 0
            total_growth_volume = 0  # ✅ NEW: Track pounds growing
            total_decline_volume = 0  # ✅ NEW: Track pounds declining

            for customer_id in combo_data[self.config.data_config.customer_id_column].unique():
                cust_data = combo_data[combo_data[self.config.data_config.customer_id_column] == customer_id]
                pounds_cy = cust_data['Pounds_CY'].sum()
                pounds_py = cust_data['Pounds_PY'].sum()
                
                if pounds_cy > 0:  # Active this year
                    customers_active += 1
                    volume_change = pounds_cy - pounds_py  # ✅ Calculate change in pounds
                    
                    if volume_change > 0:
                        customers_growing += 1
                        total_growth_volume += volume_change  # ✅ Track actual pounds growing
                    elif volume_change < 0:
                        customers_declining += 1
                        total_decline_volume += abs(volume_change)  # ✅ Track actual pounds declining

            if customers_active > 0:
                # Customer-based rates (what % of customers are growing?)
                growth_rate = customers_growing / customers_active
                decline_rate = customers_declining / customers_active
                
                # ✅ NEW: Volume-based analysis (what % of total volume is growing?)
                net_volume_change = total_growth_volume - total_decline_volume
                net_change_pct = net_volume_change / total_volume if total_volume > 0 else 0
                
                print(f"   📊 Customer Analysis: {customers_growing} growing, {customers_declining} declining out of {customers_active} active")
                print(f"   📈 Customer Rates: {growth_rate:.1%} growing | {decline_rate:.1%} declining")
                print(f"   💰 Volume Analysis: +{total_growth_volume:,.0f} lbs growing, -{total_decline_volume:,.0f} lbs declining")
                print(f"   💰 Net Volume Change: {net_volume_change:+,.0f} lbs ({net_change_pct:+.1%})")
                
                # ✅ IMPROVED LOGIC: Use NET VOLUME CHANGE instead of customer count
                # If the VOLUME is growing by 10%+, the zone is working - don't touch it!
                if net_change_pct > 0.10:
                    print(f"   🎉 KEEPING ZONE: Net volume growing by {net_change_pct:+.1%} at Zone {int(current_zone)} - ZONE IS WORKING!")
                    continue  # Skip to next combo - no recommendation
                
                # ✅ NEW: If volume is slightly declining but customer count is growing, investigate further
                if net_change_pct > -0.05 and growth_rate > 0.50:
                    print(f"   🤔 STABLE ZONE: Volume nearly flat ({net_change_pct:+.1%}) but {growth_rate:.1%} of customers growing - ZONE IS ACCEPTABLE")
                    continue  # Zone is stable enough
                
                # ✅ IMPROVED: If <30% lapsed AND volume is stable/growing, KEEP THE ZONE
                if lapsed_pct < required_lapse_rate and net_change_pct >= 0:
                    print(f"   ✅ KEEPING ZONE: Low lapse rate ({lapsed_pct:.1%}) + stable/growing volume ({net_change_pct:+.1%})")
                    continue  # Skip to next combo

            # Get journey evidence
            green_flag_count = 0
            reactive_recovery_count = 0
            margin_giveaway_count = 0
            
            if not combo_journey.empty:
                green_flag_count = len(combo_journey[combo_journey['customer_type'].str.contains('GREEN_FLAG', na=False)])
                reactive_recovery_count = len(combo_journey[combo_journey['customer_type'] == 'GREEN_FLAG_REACTIVE_RECOVERY'])
                margin_giveaway_count = len(combo_journey[combo_journey['customer_type'] == 'RED_FLAG_PROACTIVE_FLAT'])
                print(f"   Journey: {green_flag_count} green flags, {reactive_recovery_count} reactive recoveries")
            
            # Never recommend going UP
            if optimal_zone > current_zone:
                optimal_zone = current_zone
            
            # Skip if no change needed
            if current_zone == optimal_zone:
                print(f"   ⏭️  SKIPPED: Already at optimal zone")
                continue
            
            print(f"   ✅ WILL RECOMMEND: {int(current_zone)} → {int(optimal_zone)}")
            
            # Check if reactive
            is_reactive = combo in self.reactive_flags
            # Build recommendation
            rec = {
                'company_combo': combo,
                'company_name': combo_data[self.config.data_config.company_column].iloc[0],
                'cuisine': combo_data.get(self.config.data_config.cuisine_column, pd.Series(['N/A'])).iloc[0] if self.config.data_config.use_cuisine else 'N/A',
                'attribute_group': combo_data.get(self.config.data_config.attribute_group_column, pd.Series(['N/A'])).iloc[0] if self.config.data_config.use_attribute_group else 'N/A',
                'current_zone': int(current_zone),
                'recommended_zone': int(optimal_zone),
                'total_customers': int(total_customers),
                'active_customers': int(active_customers),
                'lapsed_customers': int(lapsed_customers),
                'lapsed_pct': lapsed_pct,
                'total_volume': float(total_volume),
                'green_flag_customers': int(green_flag_count),
                'reactive_recovery_customers': int(reactive_recovery_count),
                'margin_giveaway_count': int(margin_giveaway_count),
                'consistent_buyers': consistent_buyers_count, 
                'green_flag_rate': green_flag_rate,  
                'implementation_priority': 0
            }

            # ✅ CALCULATE RATES FIRST (before if-elif chain)
            reactive_rate = reactive_recovery_count / total_customers if total_customers > 0 else 0
            green_flag_rate_journey = green_flag_count / total_customers if total_customers > 0 else 0

            # NOW start the if-elif chain
            # Determine recommendation type
            if lapsed_pct >= required_lapse_rate and current_zone > optimal_zone:
                # Check if we have PROOF that price was the issue
                has_reactive_evidence = reactive_recovery_count >= max(3, int(lapsed_customers * 0.10))
                
                if has_reactive_evidence:
                    rec.update({
                        'recommendation_type': 'HIGH_RECOVERY_POTENTIAL',
                        'stakeholder_message': (
                            f"🎯 PROVEN WIN: {reactive_recovery_count} similar customers came back at Zone {int(optimal_zone)}. "
                            f"{lapsed_customers} customers ({lapsed_pct:.0%}) at risk - strong evidence price is the issue."
                        ),
                        'expected_result': f"Win back {int(lapsed_customers * 0.6)} customers = +{int(total_volume * 0.3):,} lbs",
                        'timeline': '⚡ 2-4 weeks',
                        'risk_level': '✅ LOW',
                        'implementation_priority': 95
                    })
                else:
                    rec.update({
                        'recommendation_type': 'HIGH_RECOVERY_POTENTIAL',
                        'stakeholder_message': (
                            f"🤔 TEST OPPORTUNITY: {lapsed_customers} customers ({lapsed_pct:.0%}) stopped buying. "
                            f"Testing Zone {int(optimal_zone)} to see if price is the issue."
                        ),
                        'expected_result': f"Win back {int(lapsed_customers * 0.4)} customers = +{int(total_volume * 0.2):,} lbs",
                        'timeline': '⚡ 2-4 weeks',
                        'risk_level': '✅ LOW',
                        'implementation_priority': 75
                    })

            elif is_reactive:
                reactive_info = self.reactive_flags[combo]
                rec.update({
                    'recommendation_type': 'REACTIVE_CORRECTION',
                    'stakeholder_message': reactive_info['stakeholder_message'],
                    'expected_result': f"Return to sustainable pricing = +{int(total_volume * 0.20):,} lbs",
                    'timeline': '📅 4-6 weeks',
                    'risk_level': '✅ LOW',
                    'implementation_priority': 85
                })

            elif reactive_rate >= 0.20 and current_zone > optimal_zone:
                # ✅ ADD: For high consistency zones, require STRONGER reactive evidence
                min_reactive_threshold = 0.30 if high_consistency_zone else 0.20
                
                if reactive_rate >= min_reactive_threshold:
                    rec.update({
                        'recommendation_type': 'REACTIVE_CORRECTION',
                        'stakeholder_message': (
                            f"⚠️ {reactive_recovery_count} customers ({reactive_rate:.0%}) lapsed at Zone {int(current_zone)}, "
                            f"then came back at Zone {int(optimal_zone)}. This proves Zone {int(current_zone)} was wrong."
                        ),
                        'expected_result': f"Recover lapsed customers = +{int(total_volume * 0.2):,} lbs",
                        'timeline': '📅 4-6 weeks',
                        'risk_level': '✅ LOW',
                        'implementation_priority': 85
                    })
                    
            elif green_flag_rate_journey >= 0.30 and current_zone > optimal_zone:  # ✅ NOW this works!
                rec.update({
                    'recommendation_type': 'GREEN_FLAG_ZONE',
                    'stakeholder_message': (
                        f"🟢 {green_flag_count} customers ({green_flag_rate_journey:.0%}) are stable buyers at Zone {int(optimal_zone)}. "
                        f"Move Zone {int(current_zone)} customers down to match them."
                    ),
                    'expected_result': f"Match stable performers = +{int(total_volume * 0.15):,} lbs",
                    'timeline': '📅 4-6 weeks',
                    'risk_level': '✅ LOW',
                    'implementation_priority': 80
                })

            elif current_zone > optimal_zone:
                rec.update({
                    'recommendation_type': 'PRICE_ADJUSTMENT',
                    'stakeholder_message': (
                        f"Historical data suggests Zone {int(optimal_zone)} performs better than Zone {int(current_zone)}."
                    ),
                    'expected_result': f"Modest volume increase = +{int(total_volume * 0.10):,} lbs",
                    'timeline': '📅 4-6 weeks',
                    'risk_level': '✅ LOW',
                    'implementation_priority': 60
                })

            else:
                rec.update({
                    'recommendation_type': 'MONITOR',
                    'stakeholder_message': f"Zone {int(current_zone)} performing adequately",
                    'expected_result': 'Continue monitoring',
                    'timeline': 'Review quarterly',
                    'risk_level': '✅ LOW',
                    'implementation_priority': 30
                })

            # Add warning if applicable
            if margin_giveaway_count > 0:
                rec['warning'] = f"⚠️ {margin_giveaway_count} customers were moved down but didn't increase volume - margin giveaway"

            recommendations.append(rec)
        
        # Sort by priority
        recommendations = sorted(recommendations, key=lambda x: x['implementation_priority'], reverse=True)

        print(f"\n✅ Generated {len(recommendations)} recommendations")
        print(f"🚩 Flagged {len(zone_1_to_0_flags)} Zone 1→0 moves (not recommended)")

        # ✅ SAVE FLAGS:
        self.zone_1_to_0_flags = zone_1_to_0_flags  # Store as instance variable

        return recommendations
    
    def _get_consensus_zone(self, combo: str, current_zone: int, 
                            historical_df: pd.DataFrame) -> int:
        """Find consensus zone from historical data for THIS SPECIFIC COMBO."""
        
        hist = historical_df[historical_df['Company_Combo_Key'] == combo].copy()
        
        if hist.empty:
            return current_zone
        
        hist = classify_customer_activity(hist, self.config)
        hist_consistency = calculate_purchase_consistency(hist, self.config)
        
        # Aggregate performance by zone
        zone_perf = hist.groupby(self.config.data_config.zone_column).agg({
            'Pounds_CY': 'sum',
            'Computer Margin $ Per LB CY': 'mean',  # ✅ ADD MARGIN
            'Customer_Status': [
                lambda x: (x == 'ACTIVE_BUYER').mean(),
                lambda x: (x == 'LAPSED_FROM_CATEGORY').mean()
            ],
            self.config.data_config.customer_id_column: 'nunique'
        }).reset_index()
        
        zone_perf.columns = ['Zone', 'Total_Volume', 'Avg_Margin', 'Active_Rate', 'Lapse_Rate', 'Customer_Count']
        
        # Require minimum sample size
        zone_perf = zone_perf[zone_perf['Customer_Count'] >= 3]
        
        if zone_perf.empty:
            return current_zone
        
        # Merge consistency
        zone_perf = zone_perf.merge(
            hist_consistency[['Zone', 'green_flag_rate']],
            on='Zone',
            how='left'
        ).fillna(0)
        
        # Normalize each metric to 0-1 scale
        zone_perf['Volume_Norm'] = zone_perf['Total_Volume'] / zone_perf['Total_Volume'].max()
        zone_perf['Margin_Norm'] = zone_perf['Avg_Margin'] / zone_perf['Avg_Margin'].max()  # ✅ ADD THIS
        zone_perf['Active_Norm'] = zone_perf['Active_Rate']
        zone_perf['GreenFlag_Norm'] = zone_perf['green_flag_rate']
        zone_perf['Retention_Norm'] = 1 - zone_perf['Lapse_Rate']

        # Weighted scoring: Volume 50%, Margin 25%, Quality 25%
        zone_perf['Score'] = (
            zone_perf['Volume_Norm'] * 0.50 +       # Volume: 50%
            zone_perf['Margin_Norm'] * 0.25 +       # Margin: 25% ✅ ADDED
            zone_perf['Active_Norm'] * 0.10 +       # Active: 10%
            zone_perf['GreenFlag_Norm'] * 0.10 +    # Green Flag: 10%
            zone_perf['Retention_Norm'] * 0.05      # Retention: 5%
        )
        
        best_zone = int(zone_perf.loc[zone_perf['Score'].idxmax(), 'Zone'])
        
        # Never recommend going UP from current zone
        if best_zone > current_zone:
            return current_zone
        
        return best_zone  # ✅ Return the best zone found
# ============================================================================
# DASHBOARD GENERATORS
# ============================================================================

def create_dashboard_guide() -> pd.DataFrame:
    """
    Create a user guide explaining how to read this dashboard.
    
    8th Grade Explanation:
    "This is your instruction manual. It explains what each tab means and 
    how to use this report to make pricing decisions."
    """
    
    guide_content = [
        {'Section': 'ABOUT THIS REPORT', 'Content': ''},
        {'Section': 'Purpose', 'Content': 'This dashboard analyzes historical sales data to find pricing zones that will grow volume, win back customers, and increase margin. Every recommendation is backed by your actual sales data.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'HOW TO USE THIS REPORT', 'Content': ''},
        {'Section': 'Step 1', 'Content': 'Start with Tab 1 (EXECUTIVE_SUMMARY) to see the big picture'},
        {'Section': 'Step 2', 'Content': 'Review Tab 2 (TOP_5_MOVES) - these are your quick wins to implement first'},
        {'Section': 'Step 3', 'Content': 'Check Tab 10 (FINANCIAL_IMPACT) to see the dollar value of each move'},
        {'Section': 'Step 4', 'Content': 'Use Tab 5 (TIMELINE) to plan your implementation schedule'},
        {'Section': 'Step 5', 'Content': 'Track progress using Tab 8 (YOY_CUSTOMERS) in your next run'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'TAB-BY-TAB GUIDE', 'Content': ''},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 1: EXECUTIVE_SUMMARY', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'High-level metrics: how many opportunities found, volume at stake, financial impact, year-over-year trends'},
        {'Section': 'Key metrics explained', 'Content': '"Distinct Customers This Year" = How many different restaurants bought this category in the last 8 weeks this year (each customer counted once, even if they have 100 transactions)'},
        {'Section': '', 'Content': '"Retention Rate" = What % of last year\'s customers are still buying (80%+ is healthy, <70% means we\'re losing customers)'},
        {'Section': '', 'Content': '"Lapsed Customers" = Restaurants that stopped buying THIS category but are still buying OTHER items from us (HIGH recovery opportunity!)'},
        {'Section': 'Action', 'Content': 'Share this page in executive meetings to show the scale of opportunity'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 2: TOP_5_MOVES', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'The 5 highest-priority zone changes ranked by impact and ease of implementation'},
        {'Section': 'Key columns explained', 'Content': '"Move_To_Zone" = The zone we recommend changing TO (always equals or lower than current - we never suggest price increases without approval)'},
        {'Section': '', 'Content': '"YoY_Customer_Change" = How many customers we gained/lost in the last 8 weeks vs same period last year (negative = losing customers)'},
        {'Section': '', 'Content': '"Problem" = Why this needs fixing in plain English'},
        {'Section': '', 'Content': '"Expected_Result" = What we predict will happen (conservative estimate based on historical patterns)'},
        {'Section': '', 'Content': '"Timeline" = How fast you\'ll see results (⚡ 2-4 weeks = quick wins, 📅 4-6 weeks = standard)'},
        {'Section': 'Action', 'Content': 'Implement these 5 moves first. They have the best chance of quick wins.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 3: RECOVERY_OPPORTUNITIES', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Every combo where customers stopped buying this category but are still active with the company'},
        {'Section': 'Key metrics', 'Content': '"Lapsed_Customers" = Count of restaurants that stopped buying but are still ordering other items (GOLD MINE for recovery!)'},
        {'Section': '', 'Content': '"Lapsed_Rate" = % of total customers who lapsed (30%+ is a red flag that pricing is too high)'},
        {'Section': '', 'Content': '"Expected_Recovery" = How many we predict will come back if we fix pricing (60% success rate assumption)'},
        {'Section': 'Action', 'Content': 'Sales team can call these customers with the new pricing. They already have a relationship!'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 4: REACTIVE_ALERTS', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Times when we overpriced, lost customers, then panic-dropped the zone. The lower zone looks good but only because we killed business at the higher zone first.'},
        {'Section': 'Why this matters', 'Content': 'If Detroit Bar & Grill is at Zone 2 and looks great, but we dropped from Zone 4 after losing customers, the REAL optimal is probably Zone 3 (not 2 or 4)'},
        {'Section': 'Key columns', 'Content': '"High_Zone_Used" = The zone where we lost customers. "Panic_Drop_To" = Where we moved it. "Likely_True_Optimal" = Where it should probably be.'},
        {'Section': 'Action', 'Content': 'Be skeptical of these zones. Test the middle ground between high/low.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 5: TIMELINE', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Week-by-week implementation plan'},
        {'Section': 'How to use', 'Content': 'Week 1-2: Make the changes. Week 3-4: Monitor early results. Week 5-8: Measure full impact. Week 8+: Run this analysis again to see what worked.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 6: LEARNING_TRACKER', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Tracks recommendations over time so the system learns what works'},
        {'Section': 'Note', 'Content': 'This will be empty on first run. After you implement recommendations and re-run the analysis, this tab shows which predictions were accurate.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 7: ALL_RECOMMENDATIONS', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Complete list of all recommended zone changes, sorted by priority'},
        {'Section': 'How to use', 'Content': 'After implementing Top 5, come here to see what to do next. Lower priority moves = smaller impact but still worth doing over time.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 8: YOY_CUSTOMERS', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Year-over-year customer counts by combo (NOT by zone). Shows which categories are growing or shrinking overall.'},
        {'Section': 'Important note', 'Content': 'This compares the LAST 8 WEEKS this year vs SAME 8 WEEKS last year. It\'s not the full year. Use this to spot trends.'},
        {'Section': 'Key columns', 'Content': '"Retained_Customers" = Bought in both years (good!). "New_Customers" = Bought this year but not last year (growth!). "Lost_Customers" = Bought last year but not this year (problem!).'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 9: VOLUME_DECLINES', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Biggest volume losses by combo AND zone. Shows exactly where pounds are bleeding.'},
        {'Section': 'Difference from Tab 8', 'Content': 'Tab 8 shows customers by COMBO only. Tab 9 shows volume by COMBO + ZONE. This is more tactical - tells you which specific zone is the problem.'},
        {'Section': 'Key insight', 'Content': 'If Detroit Bar & Grill_Z4 lost 33,000 lbs but Detroit Bar & Grill_Z2 gained 5,000 lbs, the problem is Zone 4 specifically (not the category overall).'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 10: FINANCIAL_IMPACT', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Dollar value of each recommendation using your actual Net Sales per LB and Margin per LB data'},
        {'Section': 'Three scenarios', 'Content': 'LOW = Conservative (if only 60% of predictions work). EXPECTED = Most likely outcome. HIGH = Optimistic (if 80%+ of predictions work).'},
        {'Section': 'Important', 'Content': '8-Week Impact = First cycle results. Annual Impact = Ongoing value if we keep the zone there (multiplied across 52 weeks).'},
        {'Section': 'Confidence levels', 'Content': 'HIGH = Lapsed customer recovery (proven pattern). MEDIUM = Reactive corrections or green flag zones. LOW = Standard adjustments.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Tab 11: ZONE_STICKINESS', 'Content': ''},
        {'Section': 'What it shows', 'Content': 'Which zones keep customers coming back week after week (the "green flag" metric)'},
        {'Section': 'Green Flag Rate', 'Content': '% of customers who buy consistently (75%+ of weeks). High rate = sticky zone that works. Low rate = customers try once and leave.'},
        {'Section': 'How to use', 'Content': 'If Zone 2 has 60% green flag rate and Zone 4 has 20%, that\'s proof Zone 2 works better at keeping customers engaged.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'KEY DEFINITIONS', 'Content': ''},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Distinct Customers', 'Content': 'Each customer counted ONCE, even if they have 100 transactions. This is the count of unique restaurants, not transaction count.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Lapsed Customer', 'Content': 'A restaurant that STOPPED buying this specific category (e.g. groundfish) BUT is still buying other items from us. HIGH recovery opportunity because relationship still exists!'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Lost Customer', 'Content': 'A restaurant that hasn\'t bought ANYTHING from us in 9+ weeks. Lower recovery potential - relationship may be gone.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Reactive Pricing', 'Content': 'When we overprice, lose customers, then panic and drop the zone. The lower zone looks good in data, but only because we killed business at the higher zone first. Be skeptical of these patterns.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Green Flag / Zone Stickiness', 'Content': 'Customers who buy week after week at a specific zone. High stickiness = zone works well. Low stickiness = customers try once and leave (sign of overpricing).'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Pounds CY vs PY', 'Content': 'CY = Current Year (the fiscal year you\'re analyzing). PY = Previous Year (same time period last year for comparison). Each row has BOTH years on it for easy comparison.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'COMMON QUESTIONS', 'Content': ''},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Q: Why 8 weeks for YoY?', 'Content': 'A: 8 weeks = 2 months, enough to smooth out weekly volatility but recent enough to spot trends. You can change this in the config (YOY_LOOKBACK_WEEKS).'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Q: Why never suggest Zone 1→0?', 'Content': 'A: Zone 0 gives away margin. If Zone 1 is too high and Zone 0 is optimal, we FLAG it for management approval (needs fractional zones like 0.5 or 0.75).'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Q: How often should we run this?', 'Content': 'A: First run: Implement top 20-30 moves. Week 4-6: Check early results. Week 8-12: Full results. Next run: Re-analyze with new data, implement next batch. Repeat quarterly.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Q: What if recommendations don\'t work?', 'Content': 'A: The system learns! Tab 6 (LEARNING_TRACKER) tracks predictions vs actuals. Over time, the system gets smarter about what works in your business.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'Q: Can I trust the financial projections?', 'Content': 'A: They use YOUR actual Net Sales and Margin per LB data (not guesses). The EXPECTED column is conservative. The LOW column is very conservative (60% success rate). Results may vary but are based on real historical patterns.'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'NEXT STEPS', 'Content': ''},
        {'Section': '1', 'Content': 'Review Tab 2 (TOP_5_MOVES) with your team'},
        {'Section': '2', 'Content': 'Check Tab 10 (FINANCIAL_IMPACT) to prioritize by ROI'},
        {'Section': '3', 'Content': 'Make the zone changes in your pricing system'},
        {'Section': '4', 'Content': 'Alert sales team about which customers may see new pricing'},
        {'Section': '5', 'Content': 'Track results weekly using your BI tools'},
        {'Section': '6', 'Content': 'Re-run this analysis in 8-12 weeks to measure actual impact'},
        {'Section': '', 'Content': ''},
        
        {'Section': 'SUPPORT', 'Content': ''},
        {'Section': 'Questions?', 'Content': 'Contact your pricing analyst or data team'},
        {'Section': 'Report Date', 'Content': datetime.now().strftime('%Y-%m-%d %H:%M')},
    ]
    
    return pd.DataFrame(guide_content)

def create_executive_summary(recommendations, yoy_metrics=None, reactive_flags=None, financial_summary=None, config=None):
    """One-page summary with company name."""
    
    high_priority = [r for r in recommendations if r['implementation_priority'] >= 70]
    recovery_opps = [r for r in recommendations if r['recommendation_type'] == 'HIGH_RECOVERY_POTENTIAL']
    fractional_needed = [r for r in recommendations if r.get('requires_fractional', False)]
    
    total_volume = sum(r['total_volume'] for r in high_priority)
    total_lapsed = sum(r['lapsed_customers'] for r in recovery_opps)
    
    summary = {
        'total_opportunities': len(recommendations),
        'high_priority_moves': len(high_priority),
        'easy_wins_lapsed_recovery': len(recovery_opps),
        'fractional_zones_needed': len(fractional_needed),
        'total_volume_at_stake': f"{total_volume:,.0f} lbs",
        'customers_to_win_back': int(total_lapsed),
        'expected_timeframe': '2-8 weeks',
        'first_action': high_priority[0]['stakeholder_message'] if high_priority else 'No immediate actions'
    }
    

    filter_info = []
        
    # Get company name from recommendations if filtered
    if config and config.FILTER_BY_company_number and recommendations:  # ✅ FIXED - attribute access
        company_name = recommendations[0].get('company_name', config.FILTER_BY_company_number)
        filter_info.append(f"Company: {company_name} (ID: {config.FILTER_BY_company_number})")
    elif config and config.FILTER_BY_company_number:  # ✅ Add safety check
        filter_info.append(f"Company Number: {config.FILTER_BY_company_number}")

    if config and config.FILTER_BY_company_region_id:  # ✅ Add safety check
        filter_info.append(f"Region: {config.FILTER_BY_company_region_id}")

    summary['scope'] = ', '.join(filter_info) if filter_info else 'All Companies & Regions'
    summary['yoy_lookback_window'] = f"{config.YOY_LOOKBACK_WEEKS} weeks" if config else "8 weeks"  # ✅ Add fallback
    
    # Add YoY metrics
    if yoy_metrics:
        summary.update({
            'yoy_comparison_period': yoy_metrics.get('week_range', 'N/A'),
            'distinct_customers_this_year': f"{yoy_metrics['distinct_customers_cy']:,}",
            'distinct_customers_last_year': f"{yoy_metrics['distinct_customers_py']:,}",
            'customer_change_count': f"{yoy_metrics['customer_change']:+,}",
            'customer_change_percent': f"{yoy_metrics['customer_change_pct']:+.1f}%",
            'customer_retention_rate': f"{yoy_metrics['retention_rate']:.1f}%",
            'new_customers_gained': f"{yoy_metrics['new_customers']:,}",
            'customers_lost': f"{yoy_metrics['lost_customers']:,}"
        })

        # ✅ ADD FINANCIAL SECTION
    if financial_summary:
        summary.update({
            '': '',  # Spacer
            'FINANCIAL IMPACT (TOP 10 MOVES)': '',
            'revenue_8wk_range': f"${financial_summary['top10_revenue_8wk_low']:,.0f} - ${financial_summary['top10_revenue_8wk_high']:,.0f}",
            'revenue_8wk_expected': f"${financial_summary['top10_revenue_8wk_expected']:,.0f}",
            'margin_8wk_range': f"${financial_summary['top10_margin_8wk_low']:,.0f} - ${financial_summary['top10_margin_8wk_high']:,.0f}",
            'margin_8wk_expected': f"${financial_summary['top10_margin_8wk_expected']:,.0f}",
            'annualized_revenue_all_moves': f"${financial_summary['all_revenue_annual_expected']:,.0f}",
            'annualized_margin_all_moves': f"${financial_summary['all_margin_annual_expected']:,.0f}",
            'confidence_breakdown': f"High: {financial_summary['high_confidence_count']}, Medium: {financial_summary['medium_confidence_count']}, Low: {financial_summary['low_confidence_count']}"
        })
    # ADD THESE DEBUG LINES:
    print(f"\n🐛 DEBUG create_executive_summary():")
    print(f"   financial_summary = {financial_summary}")
    print(f"   financial_summary is None? {financial_summary is None}")
    print(f"   financial_summary is empty? {not financial_summary if financial_summary else 'N/A'}")
    
    if financial_summary:
        print(f"   ✅ INSIDE financial_summary block!")
        print(f"   Keys: {list(financial_summary.keys()) if isinstance(financial_summary, dict) else 'Not a dict'}")
        summary.update({
            '': '',  # Spacer
            'FINANCIAL IMPACT (TOP 10 MOVES)': '',
            'revenue_8wk_range': f"${financial_summary['top10_revenue_8wk_low']:,.0f} - ${financial_summary['top10_revenue_8wk_high']:,.0f}",
            # ... rest of your financial code
        })
    else:
        print(f"   ❌ SKIPPING financial_summary block (it's None or empty)")
    
    return pd.DataFrame([summary])  


def create_quick_wins_dashboard(recommendations: List[Dict], 
                               yoy_metrics: Optional[Dict] = None) -> pd.DataFrame:
    """Top 5 moves in plain English with YoY context."""
    
    quick_wins = []
    
    # Get combo-level YoY data
    yoy_by_combo = {}
    if yoy_metrics and 'by_combo' in yoy_metrics:
        for _, row in yoy_metrics['by_combo'].iterrows():
            yoy_by_combo[row['Company_Combo_Key']] = row
    
    for i, rec in enumerate(recommendations[:5], 1):
        
        # Build problem description
        if rec['recommendation_type'] == 'HIGH_RECOVERY_POTENTIAL':
            problem = f"🎯 {rec['lapsed_customers']} restaurants stopped buying but still order other items"
        elif rec['recommendation_type'] == 'REACTIVE_CORRECTION':
            problem = f"⚠️ Overpriced, lost customers, then dropped price. Need sweet spot"
        elif rec['recommendation_type'] == 'NEEDS_FRACTIONAL_ZONES':
            problem = f"⚠️ Lowest zone but still losing {rec['lapsed_customers']} customers"
        else:
            problem = f"Zone {rec['current_zone']} higher than peer sites"
        
        # Add GREEN FLAG indicator
        green_flag_status = '🟢 HIGH' if rec.get('green_flag_rate', 0) >= 0.50 else '🟡 MEDIUM' if rec.get('green_flag_rate', 0) >= 0.30 else '🔴 LOW'
        
        # Get YoY customer data for this combo
        combo = rec['company_combo']
        yoy_data = yoy_by_combo.get(combo, {})
        
        cy_cust = yoy_data.get('distinct_customers_cy', 'N/A')
        py_cust = yoy_data.get('distinct_customers_py', 'N/A')
        change = yoy_data.get('customer_change', 'N/A')
        
        if isinstance(change, (int, float)):
            yoy_status = f"{change:+,} ({change/py_cust*100:+.1f}%)" if py_cust != 'N/A' and py_cust > 0 else f"{change:+,}"
        else:
            yoy_status = 'N/A'
        
        quick_wins.append({
            'Priority': i,
            'OpCo': rec['company_name'],
            'Category': f"{rec['cuisine']} - AG{rec['attribute_group']}",
            'Current_Zone': rec['current_zone'],
            'Move_To_Zone': rec['recommended_zone'],
            'Customers_This_Year': cy_cust if cy_cust != 'N/A' else 'N/A',
            'Customers_Last_Year': py_cust if py_cust != 'N/A' else 'N/A',
            'YoY_Customer_Change': yoy_status,
            'Problem': problem,
            'Expected_Result': rec['expected_result'],
            'Timeline': rec['timeline'],
            'Risk': rec['risk_level'],
            'Volume_at_Stake': f"{rec['total_volume']:,.0f} lbs",
            'Current_Zone_Stickiness': green_flag_status,
            'Consistent_Buyers': rec.get('consistent_buyers', 0)
        })
    
    return pd.DataFrame(quick_wins)

def create_yoy_customer_dashboard(yoy_metrics: Dict) -> pd.DataFrame:
    """
    Year-over-year customer tracking dashboard.
    
    Simple explanation:
    "Shows clearly: Are we gaining or losing customers? Which combos are growing?"
    """
    
    if not yoy_metrics or 'by_combo' not in yoy_metrics:
        return pd.DataFrame([{'Message': 'No YoY data available'}])
    
    combo_df = yoy_metrics['by_combo'].copy()
    
    # Sort by biggest customer losses first (need attention)
    combo_df = combo_df.sort_values('customer_change', ascending=True)
    
    # Add status indicators
    def status(change):
        if change > 10:
            return '🟢 GROWING'
        elif change > 0:
            return '🟡 SLIGHT GROWTH'
        elif change >= -10:
            return '🟠 SLIGHT DECLINE'
        else:
            return '🔴 MAJOR DECLINE'
    
    combo_df['Status'] = combo_df['customer_change'].apply(status)
    
    # Format for display
    display = combo_df[[
        'Company_Combo_Key', 'distinct_customers_cy', 'distinct_customers_py',
        'customer_change', 'customer_change_pct', 'retention_rate',
        'new_customers', 'lost_customers', 'Status'
    ]].copy()
    
    display.columns = [
        'Combo', 'Customers This Year', 'Customers Last Year',
        'Change', 'Change %', 'Retention Rate %',
        'New Customers', 'Lost Customers', 'Status'
    ]
    
    # Format percentages
    display['Change %'] = display['Change %'].apply(lambda x: f"{x:+.1f}%")
    display['Retention Rate %'] = display['Retention Rate %'].apply(lambda x: f"{x:.1f}%")
    
    return display

def create_zone_stickiness_report(consistency_analysis: pd.DataFrame) -> pd.DataFrame:
    """
    Show which zones have the best customer retention (GREEN FLAGS!).
    
    8th Grade Explanation:
    "This shows which zones keep customers coming back week after week. 
    High stickiness = GREEN FLAG! Low stickiness = customers try once and leave."
    """
    
    if consistency_analysis is None or consistency_analysis.empty:
        return pd.DataFrame([{'Message': 'No consistency data available'}])
    
    stickiness = consistency_analysis.copy()
    
    # Add color-coded flags
    def flag(rate):
        if rate >= 0.60:
            return '🟢 EXCELLENT'
        elif rate >= 0.40:
            return '🟡 GOOD'
        elif rate >= 0.20:
            return '🟠 FAIR'
        else:
            return '🔴 POOR'
    
    stickiness['Stickiness_Rating'] = stickiness['green_flag_rate'].apply(flag)
    
    # Sort by best stickiness
    stickiness = stickiness.sort_values('green_flag_rate', ascending=False)
    
    # Format for display
    display = stickiness[[
        'Company_Combo_Key', 'Zone', 'distinct_customers', 
        'consistent_buyers', 'green_flag_rate', 'avg_consistency_rate',
        'Stickiness_Rating'
    ]].copy()
    
    display.columns = [
        'Combo', 'Zone', 'Distinct Customers', 'Consistent Buyers',
        'GREEN FLAG Rate', 'Avg Consistency', 'Stickiness Rating'
    ]
    
    # Format percentages
    display['GREEN FLAG Rate'] = display['GREEN FLAG Rate'].apply(lambda x: f"{x:.1%}")
    display['Avg Consistency'] = display['Avg Consistency'].apply(lambda x: f"{x:.1%}")
    
    return display

def create_volume_decline_dashboard(df: pd.DataFrame, 
                                    config: FoodserviceConfig,
                                    top_n: int = 50) -> pd.DataFrame:
    """Show biggest volume losses by combo + zone with correct customer counts."""
    
    df = df.copy()
    
    print(f"\n📉 Analyzing volume declines by combo + zone...")
    
    # Ensure numeric
    df['Pounds_CY'] = pd.to_numeric(df['Pounds_CY'], errors='coerce').fillna(0)
    df['Pounds_PY'] = pd.to_numeric(df['Pounds_PY'], errors='coerce').fillna(0)
    df['Zone_Suffix_Numeric'] = pd.to_numeric(df['Zone_Suffix_Numeric'], errors='coerce')
    
    # Drop invalid zones
    df = df.dropna(subset=['Zone_Suffix_Numeric'])
    df = df[df['Zone_Suffix_Numeric'].between(0, 5)]
    
    # Group by combo + zone and count DISTINCT customers
    combo_zone = df.groupby(
        ['Company_Combo_Key', 'Zone_Suffix_Numeric'], 
        dropna=False
    ).agg({
        'Pounds_CY': 'sum',
        'Pounds_PY': 'sum',
        config.data_config.customer_id_column: 'nunique'  # Total distinct customers at this combo+zone
    }).reset_index()
    
    # Get CY customers only (where they actually bought CY)
    cy_only = df[df['Pounds_CY'] > 0].groupby(
        ['Company_Combo_Key', 'Zone_Suffix_Numeric'],
        dropna=False
    )[config.data_config.customer_id_column].nunique().reset_index()
    cy_only.columns = ['Company_Combo_Key', 'Zone_Suffix_Numeric', 'Customers_CY']
    
    # Get PY customers only (where they actually bought PY)
    py_only = df[df['Pounds_PY'] > 0].groupby(
        ['Company_Combo_Key', 'Zone_Suffix_Numeric'],
        dropna=False
    )[config.data_config.customer_id_column].nunique().reset_index()
    py_only.columns = ['Company_Combo_Key', 'Zone_Suffix_Numeric', 'Customers_PY']
    
    # Merge
    combo_zone = combo_zone.merge(cy_only, on=['Company_Combo_Key', 'Zone_Suffix_Numeric'], how='left')
    combo_zone = combo_zone.merge(py_only, on=['Company_Combo_Key', 'Zone_Suffix_Numeric'], how='left')
    
    combo_zone['Customers_CY'] = combo_zone['Customers_CY'].fillna(0).astype(int)
    combo_zone['Customers_PY'] = combo_zone['Customers_PY'].fillna(0).astype(int)
    
    # Drop the third customer column (total) - we only want CY and PY separately
    combo_zone = combo_zone.drop(config.data_config.customer_id_column, axis=1)
    
    # Calculate changes
    combo_zone['Volume_Change_Lbs'] = combo_zone['Pounds_CY'] - combo_zone['Pounds_PY']
    combo_zone['Volume_Change_Pct'] = (combo_zone['Volume_Change_Lbs'] / combo_zone['Pounds_PY'].replace(0, 1)) * 100
    combo_zone['Customer_Change'] = combo_zone['Customers_CY'] - combo_zone['Customers_PY']
    combo_zone['Customer_Change_Pct'] = (combo_zone['Customer_Change'] / combo_zone['Customers_PY'].replace(0, 1)) * 100
    
    # Filter to declines only
    declines = combo_zone[combo_zone['Volume_Change_Lbs'] < 0].copy()
    declines = declines.sort_values('Volume_Change_Lbs', ascending=True)
    
    if top_n:
        declines = declines.head(top_n)
    
    # Format display
    declines['Combo_Zone'] = declines['Company_Combo_Key'] + '_Z' + declines['Zone_Suffix_Numeric'].astype(int).astype(str)
    
    display = declines[[
        'Combo_Zone', 'Pounds_CY', 'Pounds_PY', 'Volume_Change_Lbs', 'Volume_Change_Pct',
        'Customers_CY', 'Customers_PY', 'Customer_Change', 'Customer_Change_Pct'
    ]].copy()
    
    display.columns = [
        'Combo + Zone', 'Pounds This Year', 'Pounds Last Year', 'Volume Change (lbs)', 'Volume Change %',
        'Customers This Year', 'Customers Last Year', 'Customer Change', 'Customer Change %'
    ]
    
    # Format
    display['Pounds This Year'] = display['Pounds This Year'].apply(lambda x: f"{x:,.0f}")
    display['Pounds Last Year'] = display['Pounds Last Year'].apply(lambda x: f"{x:,.0f}")
    display['Volume Change (lbs)'] = display['Volume Change (lbs)'].apply(lambda x: f"{x:+,.0f}")
    display['Volume Change %'] = display['Volume Change %'].apply(lambda x: f"{x:+.1f}%")
    display['Customer Change'] = display['Customer Change'].apply(lambda x: f"{x:+,}")
    display['Customer Change %'] = display['Customer Change %'].apply(lambda x: f"{x:+.1f}%")
    
    print(f"   ✅ Found {len(declines)} combo-zone pairs with volume declines")
    
    return display

def generate_high_value_leads(recommendations: List[Dict],
                              current_df: pd.DataFrame,
                              config: FoodserviceConfig,
                              output_path: str) -> pd.DataFrame:
    """
    Generate a leads file for FinalCat.py with high-value customers to target.
    
    8th Grade Explanation:
    "Create a list of specific customers the sales team should call about 
    the new pricing. Focus on customers most likely to buy more."
    
    Args:
        recommendations: List of zone recommendations
        historical_df: Full historical data
        current_df: Current period data (has Customer_Status)
        config: Configuration object
        output_path: Where to save the CSV
    
    Returns:
        DataFrame of leads
    """
    
    print(f"\n🎯 Generating high-value leads for FinalCat.py...")
    
    # Get config values
    min_revenue_threshold = config.LEADS_MIN_REVENUE_THRESHOLD
    
    # Check if filtering by company
    filter_company = getattr(config.data_config, 'FILTER_BY_COMPANY_NUMBER', None)
    
    if filter_company:
        # Filtering by specific company - use filtered lead count
        top_n = config.LEADS_N_FOR_FILTERED_COMPANY
        print(f"   🏢 Filtering to Company Number: {filter_company}")
        print(f"   📊 Generating {top_n} leads for this company only")
    else:
        # No company filter - use broad lead count across all companies
        top_n = config.LEADS_TOP_N_ALL_COMPANIES
        print(f"   🌐 Generating top {top_n} leads across ALL companies")
    
    # Filter recommendations by company if needed
    if filter_company:
        filtered_recs = [
            rec for rec in recommendations 
            if rec.get('company_number') == filter_company or 
               rec.get('company_name', '').startswith(filter_company)
        ]
        print(f"   📋 Found {len(filtered_recs)} recommendations for Company {filter_company}")
    else:
        filtered_recs = recommendations
    
    # Filter to high-priority recommendations only
    high_priority_recs = sorted(
        filtered_recs, 
        key=lambda x: x.get('implementation_priority', 0), 
        reverse=True
    )[:top_n]
    
    print(f"   ✅ Selected top {len(high_priority_recs)} recommendations")
    
    leads_data = []
    total_customers_identified = 0
    
    for rec in high_priority_recs:
        combo = rec['company_combo']
        current_zone = rec['current_zone']
        recommended_zone = rec['recommended_zone']
        
        # Get customers for this combo from CURRENT data (has Customer_Status)
        combo_data = current_df[current_df['Company_Combo_Key'] == combo].copy()
        
        if combo_data.empty:
            continue
        
        # Ensure numeric columns
        combo_data['Net_Sales_CY'] = pd.to_numeric(combo_data.get('Net_Sales_CY', 0), errors='coerce').fillna(0)
        combo_data['Pounds_CY'] = pd.to_numeric(combo_data['Pounds_CY'], errors='coerce').fillna(0)
        combo_data['Pounds_PY'] = pd.to_numeric(combo_data.get('Pounds_PY', 0), errors='coerce').fillna(0)
        
        # Check if Customer_Status exists
        has_customer_status = 'Customer_Status' in combo_data.columns
        has_last_invoice = 'Last_Invoice_Date' in combo_data.columns
        
        # Build aggregation dict dynamically
        agg_dict = {
            'Net_Sales_CY': 'sum',
            'Pounds_CY': 'sum',
            'Pounds_PY': 'sum',
            'Customer Name': 'first'
        }
        
        if has_customer_status:
            agg_dict['Customer_Status'] = 'first'
        
        if has_last_invoice:
            agg_dict['Last_Invoice_Date'] = 'max'
        
        # Get customer-level aggregates
        customer_summary = combo_data.groupby(config.data_config.customer_id_column).agg(agg_dict).reset_index()
        
        # If Customer_Status is missing, create a default
        if not has_customer_status:
            # Classify based on volume
            customer_summary['Customer_Status'] = customer_summary.apply(
                lambda row: 'LAPSED_FROM_CATEGORY' if row['Pounds_CY'] == 0 and row['Pounds_PY'] > 0 
                else 'ACTIVE_BUYER', 
                axis=1
            )
        
        # If Last_Invoice_Date is missing, use a placeholder
        if not has_last_invoice:
            customer_summary['Last_Invoice_Date'] = 'N/A'
        
        # Annualize revenue (multiply 8 weeks by 6.5 to get annual estimate)
        customer_summary['Estimated_Annual_Revenue'] = customer_summary['Net_Sales_CY'] * 6.5
        
        # Filter to high-value customers only
        high_value_customers = customer_summary[
            customer_summary['Estimated_Annual_Revenue'] >= min_revenue_threshold
        ].copy()
        
        if high_value_customers.empty:
            continue
        
        # Prioritize lapsed customers (highest recovery potential)
        lapsed = high_value_customers[
            high_value_customers['Customer_Status'] == 'LAPSED_FROM_CATEGORY'
        ].copy()
        
        active = high_value_customers[
            high_value_customers['Customer_Status'] == 'ACTIVE_BUYER'
        ].copy()
        
        # Create leads for lapsed customers (TOP PRIORITY)
        for _, cust in lapsed.iterrows():
            leads_data.append({
                'Lead_Type': 'LAPSED_RECOVERY',
                'Priority': 'HIGH',
                'Company_Combo': combo,
                'Company_Name': rec.get('company_name', ''),
                'Current_Zone': current_zone,
                'Recommended_Zone': recommended_zone,
                'Customer_ID': cust[config.data_config.customer_id_column],
                'Customer_Name': cust['Customer Name'],
                'Customer_Status': cust['Customer_Status'],
                'Last_Invoice_Date': cust['Last_Invoice_Date'],
                'Pounds_CY': cust['Pounds_CY'],
                'Pounds_PY': cust['Pounds_PY'],
                'Est_Annual_Revenue': cust['Estimated_Annual_Revenue'],
                'Reason': f"Stopped buying {combo} but still active. Win back at Zone {recommended_zone}.",
                'Sales_Talking_Point': f"We're adjusting pricing on {combo} to be more competitive. You stopped ordering this but we'd love to earn your business back.",
                'Expected_Action': 'Call customer, mention new zone pricing, ask for trial order'
            })
        
        # Create leads for active customers (MEDIUM PRIORITY - upsell opportunity)
        for _, cust in active.head(10).iterrows():  # Top 10 active per combo
            leads_data.append({
                'Lead_Type': 'UPSELL_OPPORTUNITY',
                'Priority': 'MEDIUM',
                'Company_Combo': combo,
                'Company_Name': rec.get('company_name', ''),
                'Current_Zone': current_zone,
                'Recommended_Zone': recommended_zone,
                'Customer_ID': cust[config.data_config.customer_id_column],
                'Customer_Name': cust['Customer Name'],
                'Customer_Status': cust['Customer_Status'],
                'Last_Invoice_Date': cust['Last_Invoice_Date'],
                'Pounds_CY': cust['Pounds_CY'],
                'Pounds_PY': cust['Pounds_PY'],
                'Est_Annual_Revenue': cust['Estimated_Annual_Revenue'],
                'Reason': f"Active buyer. New pricing at Zone {recommended_zone} could increase volume.",
                'Sales_Talking_Point': f"You're already buying {combo}. We've improved pricing to help you increase volume and save money.",
                'Expected_Action': 'Call customer, mention improved pricing, ask for larger or more frequent orders'
            })
        
        total_customers_identified += len(lapsed) + min(10, len(active))
    
    leads_df = pd.DataFrame(leads_data)
    
    if leads_df.empty:
        print("   ⚠️  No high-value leads identified")
        return leads_df
    
    # Sort by priority and revenue
    priority_map = {'HIGH': 3, 'MEDIUM': 2, 'LOW': 1}
    leads_df['Priority_Rank'] = leads_df['Priority'].map(priority_map)
    leads_df = leads_df.sort_values(
        ['Priority_Rank', 'Est_Annual_Revenue'], 
        ascending=[False, False]
    )
    leads_df = leads_df.drop('Priority_Rank', axis=1)
    
    # Format currency
    leads_df['Est_Annual_Revenue'] = leads_df['Est_Annual_Revenue'].apply(lambda x: f"${x:,.0f}")
    
    # Save to CSV
    leads_df.to_csv(output_path, index=False)
    
    print(f"   ✅ Generated {len(leads_df):,} high-value leads")
    print(f"   📊 Breakdown:")
    print(f"      • LAPSED_RECOVERY (High Priority): {len(leads_df[leads_df['Lead_Type']=='LAPSED_RECOVERY']):,}")
    print(f"      • UPSELL_OPPORTUNITY (Medium Priority): {len(leads_df[leads_df['Lead_Type']=='UPSELL_OPPORTUNITY']):,}")
    print(f"   💾 Saved to: {output_path}")
    print(f"   🎯 Total unique customers identified: {total_customers_identified:,}")
    
    return leads_df

def calculate_financial_impact(recommendations: List[Dict],
                              historical_df: pd.DataFrame,
                              config: FoodserviceConfig) -> Tuple[pd.DataFrame, Dict]:
    """
    Calculate financial impact using ACTUAL Net Sales and Margin data.
    Conservative estimates with confidence ranges.
    
    8th Grade Explanation:
    "If we do these zone changes, how much revenue and profit do we gain?
    We use your actual pricing data, not guesses. We show low/expected/high 
    scenarios so you know the range of outcomes."
    """
    
    print("\n💰 Calculating financial impact...")
    
    financial_data = []
    
    for rec in recommendations:
        combo = rec['company_combo']
        combo_data = historical_df[
            (historical_df['Company_Combo_Key'] == combo) & 
            (historical_df['Pounds_CY'] > 0)
        ]
        
        if combo_data.empty:
            continue
        
        # Use ACTUAL data from your trusted columns
        total_sales_cy = combo_data['Net_Sales_CY'].sum()
        total_pounds_cy = combo_data['Pounds_CY'].sum()
        total_margin_cy = combo_data['Margin_CY'].sum()
        
        # Calculate weighted averages
        avg_net_sales_per_lb = (total_sales_cy / total_pounds_cy) if total_pounds_cy > 0 else 0
        avg_margin_per_lb = (total_margin_cy / total_pounds_cy) if total_pounds_cy > 0 else 0

        # ==========================================
        # LAYER 1: Per-Pound Price Changes
        # ==========================================
        if config.has_per_lb_rates:
            per_lb_cy = combo_data[config.data_config.net_sales_per_lb_cy_column].replace(r'[\$,]', '', regex=True).astype(float).mean()
            per_lb_py = combo_data[config.data_config.net_sales_per_lb_py_column].replace(r'[\$,]', '', regex=True).astype(float).mean()
            per_lb_change_pct = ((per_lb_cy - per_lb_py) / per_lb_py * 100) if per_lb_py > 0 else 0
        else:
            per_lb_change_pct = 0
        
        # ==========================================
        # LAYER 2: Margin Percentage Changes
        # ==========================================
        if config.has_margin_pct:
            margin_pct_cy = combo_data[config.data_config.margin_pct_cy_column].str.replace(r'[\%,]', '', regex=True).astype(float).mean()
            margin_pct_py = combo_data[config.data_config.margin_pct_py_column].str.replace(r'[\%,]', '', regex=True).astype(float).mean()
            margin_pct_change = margin_pct_cy - margin_pct_py
        else:
            margin_pct_change = 0
        
        # ==========================================
        # LAYER 3: Build Diagnostic Flags
        # ==========================================
        customers_lost = rec.get('customer_change', 0) < 0
        price_increased = per_lb_change_pct > 5
        
        flags = []
        if customers_lost and price_increased:
            flags.append("🚨 OVERPRICED")
        if margin_pct_change < -2:
            flags.append("📉 MARGIN COMPRESSION")
        if margin_pct_change > 2 and not customers_lost:
            flags.append("✅ MARGIN EXPANSION")
        
        diagnostic_flag = " | ".join(flags) if flags else "—"
        
        # Conservative volume lift estimates based on recommendation type
        if rec['recommendation_type'] == 'HIGH_RECOVERY_POTENTIAL':
            # Lapsed customers coming back - high confidence
            volume_lift_low = rec['total_volume'] * 0.15      # 15% (conservative)
            volume_lift_expected = rec['total_volume'] * 0.25  # 25% (realistic)
            volume_lift_high = rec['total_volume'] * 0.35     # 35% (optimistic)
            confidence = 'HIGH'
            
        elif rec['recommendation_type'] == 'REACTIVE_CORRECTION':
            # Fix overpricing - medium confidence
            volume_lift_low = rec['total_volume'] * 0.12
            volume_lift_expected = rec['total_volume'] * 0.20
            volume_lift_high = rec['total_volume'] * 0.28
            confidence = 'MEDIUM'
            
        elif rec['recommendation_type'] == 'GREEN_FLAG_ZONE':
            # Match proven zones - medium confidence
            volume_lift_low = rec['total_volume'] * 0.10
            volume_lift_expected = rec['total_volume'] * 0.15
            volume_lift_high = rec['total_volume'] * 0.22
            confidence = 'MEDIUM'
            
        else:
            # Standard adjustment - lower confidence
            volume_lift_low = rec['total_volume'] * 0.08
            volume_lift_expected = rec['total_volume'] * 0.12
            volume_lift_high = rec['total_volume'] * 0.18
            confidence = 'LOW'
        
        # Calculate financial impact for each scenario
        revenue_low = volume_lift_low * avg_net_sales_per_lb
        revenue_expected = volume_lift_expected * avg_net_sales_per_lb
        revenue_high = volume_lift_high * avg_net_sales_per_lb
        
        margin_low = volume_lift_low * avg_margin_per_lb
        margin_expected = volume_lift_expected * avg_margin_per_lb
        margin_high = volume_lift_high * avg_margin_per_lb
        
        # Timeline adjustment (faster results = higher near-term value)
        weeks_to_result = 6  # Default
        if '2-4 weeks' in rec['timeline']:
            weeks_to_result = 3
        elif '4-6 weeks' in rec['timeline']:
            weeks_to_result = 5
        
        # Annualized impact (52 weeks)
        annualized_multiplier = 52 / weeks_to_result
        
        financial_data.append({
            'combo': rec['company_combo'],
            'company': rec['company_name'],
            'current_zone': rec['current_zone'],
            'recommended_zone': rec['recommended_zone'],
            'recommendation_type': rec['recommendation_type'],
            'priority': rec['implementation_priority'],
            
            # Volume data
            'current_volume_lbs': rec['total_volume'],
            'volume_lift_low_lbs': volume_lift_low,
            'volume_lift_expected_lbs': volume_lift_expected,
            'volume_lift_high_lbs': volume_lift_high,
            
            # Pricing data
            'avg_net_sales_per_lb': avg_net_sales_per_lb,
            'avg_margin_per_lb': avg_margin_per_lb,
            'price_per_lb_change_pct': per_lb_change_pct,
            'margin_pct_change': margin_pct_change,
            'diagnostic_flag': diagnostic_flag,
            
            # 8-Week Impact (first cycle)
            'revenue_gain_8wk_low': revenue_low,
            'revenue_gain_8wk_expected': revenue_expected,
            'revenue_gain_8wk_high': revenue_high,
            'margin_gain_8wk_low': margin_low,
            'margin_gain_8wk_expected': margin_expected,
            'margin_gain_8wk_high': margin_high,
            
            # Annualized Impact (ongoing value)
            'revenue_gain_annual_expected': revenue_expected * annualized_multiplier,
            'margin_gain_annual_expected': margin_expected * annualized_multiplier,
            
            'confidence': confidence,
            'timeline_weeks': weeks_to_result
        })
    
    df = pd.DataFrame(financial_data)
    
    if df.empty:
        return df, {}
    
    # Sort by expected revenue gain (biggest impact first)
    df = df.sort_values('revenue_gain_8wk_expected', ascending=False)
    
    # Calculate summary totals
    summary = {
        'total_recommendations': len(df),
        'high_confidence_count': len(df[df['confidence'] == 'HIGH']),
        'medium_confidence_count': len(df[df['confidence'] == 'MEDIUM']),
        'low_confidence_count': len(df[df['confidence'] == 'LOW']),
        
        # Top 10 moves (highest impact)
        'top10_revenue_8wk_low': df.head(10)['revenue_gain_8wk_low'].sum(),
        'top10_revenue_8wk_expected': df.head(10)['revenue_gain_8wk_expected'].sum(),
        'top10_revenue_8wk_high': df.head(10)['revenue_gain_8wk_high'].sum(),
        'top10_margin_8wk_low': df.head(10)['margin_gain_8wk_low'].sum(),
        'top10_margin_8wk_expected': df.head(10)['margin_gain_8wk_expected'].sum(),
        'top10_margin_8wk_high': df.head(10)['margin_gain_8wk_high'].sum(),
        
        # All moves
        'all_revenue_8wk_expected': df['revenue_gain_8wk_expected'].sum(),
        'all_margin_8wk_expected': df['margin_gain_8wk_expected'].sum(),
        'all_revenue_annual_expected': df['revenue_gain_annual_expected'].sum(),
        'all_margin_annual_expected': df['margin_gain_annual_expected'].sum(),
    }
    
    print(f"   ✅ Calculated financial impact for {len(df)} recommendations")
    print(f"   💰 Top 10 moves expected impact:")
    print(f"      Revenue (8 weeks): ${summary['top10_revenue_8wk_low']:,.0f} - ${summary['top10_revenue_8wk_high']:,.0f}")
    print(f"      Margin (8 weeks): ${summary['top10_margin_8wk_low']:,.0f} - ${summary['top10_margin_8wk_high']:,.0f}")
    
    return df, summary


def create_financial_impact_dashboard(financial_df: pd.DataFrame, 
                                      summary: Dict) -> pd.DataFrame:
    """Format financial impact for stakeholder dashboard."""
    
    if financial_df.empty:
        return pd.DataFrame([{'Message': 'No financial data calculated'}])
    
    display = financial_df[[
        'company',
        'current_zone',
        'recommended_zone',
        'recommendation_type',
        'current_volume_lbs',
        'volume_lift_expected_lbs',
        'avg_net_sales_per_lb',
        'avg_margin_per_lb',
        'revenue_gain_8wk_low',
        'revenue_gain_8wk_expected',
        'revenue_gain_8wk_high',
        'margin_gain_8wk_low',
        'margin_gain_8wk_expected',
        'margin_gain_8wk_high',
        'revenue_gain_annual_expected',
        'margin_gain_annual_expected',
        'confidence',
        'timeline_weeks'
    ]].copy()
    
    display.columns = [
        'Company',
        'Current Zone',
        'Move To Zone',
        'Type',
        'Current Volume (lbs)',
        'Expected Volume Gain (lbs)',
        'Avg Price/LB',
        'Avg Margin/LB',
        'Revenue Gain (8wk Low)',
        'Revenue Gain (8wk Expected)',
        'Revenue Gain (8wk High)',
        'Margin Gain (8wk Low)',
        'Margin Gain (8wk Expected)',
        'Margin Gain (8wk High)',
        'Revenue Gain (Annual)',
        'Margin Gain (Annual)',
        'Confidence',
        'Timeline (weeks)'
    ]
    
    # Format currency
    currency_cols = [
        'Avg Price/LB', 'Avg Margin/LB',
        'Revenue Gain (8wk Low)', 'Revenue Gain (8wk Expected)', 'Revenue Gain (8wk High)',
        'Margin Gain (8wk Low)', 'Margin Gain (8wk Expected)', 'Margin Gain (8wk High)',
        'Revenue Gain (Annual)', 'Margin Gain (Annual)'
    ]
    
    for col in currency_cols:
        display[col] = display[col].apply(lambda x: f"${x:,.0f}")
    
    # Format volume
    display['Current Volume (lbs)'] = display['Current Volume (lbs)'].apply(lambda x: f"{x:,.0f}")
    display['Expected Volume Gain (lbs)'] = display['Expected Volume Gain (lbs)'].apply(lambda x: f"{x:,.0f}")
    
    return display

def create_customer_recovery_tracker(recommendations: List[Dict]) -> pd.DataFrame:
    """Show recovery opportunity by combo."""
    
    recovery_data = []
    
    for rec in recommendations:
        if rec['lapsed_customers'] > 0:
            recovery_data.append({
                'OpCo': rec['company_name'],
                'Category': f"{rec['cuisine']} - AG{rec['attribute_group']}",
                'Current_Zone': rec['current_zone'],
                'Active_Customers': rec['active_customers'],
                'Lapsed_Customers': rec['lapsed_customers'],
                'Total_Customers': rec['total_customers'],
                'Lapsed_Rate': f"{rec['lapsed_pct']:.1%}",
                'Recovery_Potential': '🔥 HIGH' if rec['lapsed_pct'] >= 0.30 else '✅ MEDIUM' if rec['lapsed_pct'] >= 0.15 else '⚪ LOW',
                'Recommended_Action': f"Drop to Zone {rec['recommended_zone']}",
                'Expected_Recovery': f"{int(rec['lapsed_customers'] * 0.6)} customers"
            })
    
    df = pd.DataFrame(recovery_data)
    
    if not df.empty:
        df['_sort'] = df['Lapsed_Customers']
        df = df.sort_values('_sort', ascending=False).drop('_sort', axis=1)
    
    return df


def create_reactive_pricing_alerts(reactive_flags: Dict) -> pd.DataFrame:
    """Flag combos with reactive pricing."""
    
    alerts = []
    
    for combo, flag in reactive_flags.items():
        alerts.append({
            'Combo': combo,
            'What_Happened': flag['stakeholder_message'],
            'High_Zone_Used': flag['from_zone'],
            'Panic_Drop_To': flag['to_zone'],
            'Volume_Decline_Before': f"{flag['pre_decline_pct']:.1%}",
            'Volume_Recovery_After': f"{flag['post_recovery_pct']:.1%}",
            'Likely_True_Optimal': flag['likely_true_optimal'],
            'Trust_Level': 'MEDIUM - Based on volume trends'
        })
    
    return pd.DataFrame(alerts)


def create_implementation_timeline(recommendations: List[Dict]) -> pd.DataFrame:
    """Week-by-week plan."""
    
    timeline = []
    
    # Week 1-2
    week_1_2 = [r for r in recommendations if '2-4 weeks' in r['timeline']][:3]
    if week_1_2:
        timeline.append({
            'Timeframe': 'Week 1-2',
            'Action': 'Implement high-recovery moves',
            'Combos': len(week_1_2),
            'Expected_Impact': f"+{sum(r['total_volume'] * 0.3 for r in week_1_2):,.0f} lbs",
            'Focus': 'Easy wins - lapsed customer recovery'
        })
    
    # Week 3-4
    timeline.append({
        'Timeframe': 'Week 3-4',
        'Action': 'Monitor Week 1-2 moves',
        'Combos': len(week_1_2),
        'Expected_Impact': 'Confirmation of gains',
        'Focus': 'Track customer reactivation'
    })
    
    # Week 5-8
    week_5_8 = [r for r in recommendations if '4-6 weeks' in r['timeline']][:5]
    if week_5_8:
        timeline.append({
            'Timeframe': 'Week 5-8',
            'Action': 'Implement reactive corrections + peer consensus',
            'Combos': len(week_5_8),
            'Expected_Impact': f"+{sum(r['total_volume'] * 0.15 for r in week_5_8):,.0f} lbs",
            'Focus': 'Finding optimal zones'
        })
    
    return pd.DataFrame(timeline)


def create_learning_tracker_tab(learning_engine: LearningEngine) -> pd.DataFrame:
    """Show what we've learned."""
    
    completed = learning_engine.get_completed_recommendations()
    
    if not completed:
        return pd.DataFrame([{
            'Message': 'No completed recommendations yet',
            'Next_Steps': 'Implement recommendations, check back in 4-6 weeks'
        }])
    
    learning_data = []
    
    for rec in completed:
        learning_data.append({
            'Date_Recommended': rec.date_recommended,
            'Date_Implemented': rec.date_implemented,
            'OpCo': rec.company_name,
            'Category': rec.category_description,
            'Zone_Change': f"{rec.from_zone} → {rec.to_zone}",
            'Type': rec.recommendation_type,
            'Predicted_Volume_Lift': f"{rec.predicted_volume_lift:,.0f} lbs",
            'Actual_Volume_Lift': f"{rec.actual_volume_lift:,.0f} lbs" if rec.actual_volume_lift else 'N/A',
            'Outcome': rec.outcome_vs_prediction or 'Pending',
            'Weeks_to_Result': rec.weeks_to_result or 'N/A',
            'Lessons': '; '.join(rec.lessons_learned) if rec.lessons_learned else 'None yet'
        })
    
    return pd.DataFrame(learning_data)

# ============================================================================
# COLUMN NORMALIZATION
# ============================================================================

def _normalize_columns(df: pd.DataFrame, config: DataConfiguration) -> pd.DataFrame:
    """Normalize column names and create derived columns."""
    
    df = df.copy()
    
    # Rename for consistency
    rename_map = {}

    # Existing pound columns
    if config.pounds_cy_column in df.columns and config.pounds_cy_column != 'Pounds_CY':
        rename_map[config.pounds_cy_column] = 'Pounds_CY'

    if config.pounds_py_column in df.columns and config.pounds_py_column != 'Pounds_PY':
        rename_map[config.pounds_py_column] = 'Pounds_PY'

    # ADD LAST INVOICE DATE (NEW!)
    if config.last_invoice_date_column in df.columns and config.last_invoice_date_column != 'Last_Invoice_Date':
        rename_map[config.last_invoice_date_column] = 'Last_Invoice_Date'

    # ADD MARGIN COLUMNS (existing)
    if config.has_margin_data:
        if config.margin_cy_column in df.columns:
            rename_map[config.margin_cy_column] = 'Margin_CY'
        
        if config.margin_py_column in df.columns:
            rename_map[config.margin_py_column] = 'Margin_PY'
    
    # ADD NET SALES COLUMNS (NEW!)
    if config.has_net_sales_data:
        if config.net_sales_cy_column in df.columns:
            rename_map[config.net_sales_cy_column] = 'Net_Sales_CY'
        
        if config.net_sales_py_column in df.columns:
            rename_map[config.net_sales_py_column] = 'Net_Sales_PY'
    
    if rename_map:
        df = df.rename(columns=rename_map)
    
    # Ensure numeric columns
    for col in ['Pounds_CY', 'Pounds_PY']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # ADD: Ensure numeric for margin/sales (NEW!)
    if config.has_margin_data:
        for col in ['Margin_CY', 'Margin_PY']:
            if col in df.columns:
                # Remove $ signs and commas before converting
                df[col] = df[col].astype(str).str.replace('$', '').str.replace(',', '')
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    if config.has_net_sales_data:
        for col in ['Net_Sales_CY', 'Net_Sales_PY']:
            if col in df.columns:
                # Remove $ signs and commas before converting
                df[col] = df[col].astype(str).str.replace('$', '').str.replace(',', '')
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # ========================================
    # EXTRACT ZONE FROM PRICE ZONE ID
    # Format: "001-1" where 001=company, 1=zone
    # ========================================
    if 'Price Zone ID' in df.columns:
        print("   🔧 Extracting zone suffix from Price Zone ID...")
        
        # Extract everything after the dash
        df['Zone_Suffix'] = df['Price Zone ID'].astype(str).str.split('-').str[-1]
        
        # Convert to numeric
        df['Zone_Suffix_Numeric'] = pd.to_numeric(df['Zone_Suffix'], errors='coerce')
        
        # Keep only valid zones (0-5 typically)
        valid_zones = df['Zone_Suffix_Numeric'].between(0, 5, inclusive='both')
        invalid_count = (~valid_zones & df['Zone_Suffix_Numeric'].notna()).sum()
        
        if invalid_count > 0:
            print(f"      ⚠️  Found {invalid_count:,} rows with zones outside 0-5 range")
            print(f"      💡 Keeping all zones for now (you can filter later)")
        
        # Show distribution
        zone_counts = df['Zone_Suffix_Numeric'].value_counts().sort_index()
        print(f"      ✅ Zone distribution:")
        for zone, count in zone_counts.items():
            if pd.notna(zone):
                print(f"         Zone {int(zone)}: {count:,} rows")
        
    else:
        print("   ⚠️  No 'Price Zone ID' column found - cannot extract zone")
        df['Zone_Suffix_Numeric'] = 5  # Default fallback
    
    # Create combo key
    combo_parts = []
    
    if config.company_column in df.columns:
        combo_parts.append(df[config.company_column].astype(str))
    
    if config.use_cuisine and config.cuisine_column in df.columns:
        combo_parts.append(df[config.cuisine_column].astype(str))
    
    if config.use_attribute_group and config.attribute_group_column in df.columns:
        combo_parts.append(df[config.attribute_group_column].astype(str))
    
    if config.use_business_center and config.business_center_column in df.columns:
        combo_parts.append(df[config.business_center_column].astype(str))
    
    if config.use_item_group and config.item_group_column in df.columns:
        combo_parts.append(df[config.item_group_column].astype(str))
    
    if config.use_price_source and config.price_source_column in df.columns:
        combo_parts.append(df[config.price_source_column].astype(str))
    
    if combo_parts:
        df['Company_Combo_Key'] = combo_parts[0]
        for part in combo_parts[1:]:
            df['Company_Combo_Key'] = df['Company_Combo_Key'] + '_' + part
    else:
        df['Company_Combo_Key'] = 'DEFAULT'
    
    # ========================================
    # CREATE COMBINED FISCAL WEEK (YYYYWW)
    # ========================================
    if config.fiscal_week_column in df.columns:
        print("   📅 Creating combined fiscal week identifier...")
        
        df['Fiscal_Week'] = pd.to_numeric(df[config.fiscal_week_column], errors='coerce').fillna(0).astype(int)
        
        # If we have fiscal year, create combined identifier for year-boundary handling
        if 'Fiscal Year Key' in df.columns:
            df['Fiscal_Year'] = pd.to_numeric(df['Fiscal Year Key'], errors='coerce').fillna(0).astype(int)
            
            # Create YYYYWW format for proper sorting across years
            df['Fiscal_Week_Combined'] = (df['Fiscal_Year'] * 100 + df['Fiscal_Week']).astype(int)        
            fiscal_year = df['Fiscal_Year'].max()
            fiscal_week = df['Fiscal_Week'].max()
            combined = df['Fiscal_Week_Combined'].max()
            print(f"      ✅ Combined fiscal identifier: FY{fiscal_year} Week {fiscal_week} (YYYYWW: {combined})")
        else:
            df['Fiscal_Week_Combined'] = df['Fiscal_Week']
            print(f"      ✅ Using fiscal week only (no year column found)")
    else:
        print("   ⚠️  No Fiscal Week column found")
        df['Fiscal_Week'] = 0
        df['Fiscal_Week_Combined'] = 0
        
    return df


# ============================================================================
# BULLETPROOF EXCEL FORMATTING (NO STAKEHOLDER ACTION NEEDED)
# ============================================================================

def _format_excel_sheets(writer):
    """
    Format Excel so stakeholders can open and read immediately.
    NO column resizing, NO row adjusting, NO manual work needed.
    
    Simple explanation:
    "Make the Excel file look perfect the moment they open it. 
    They shouldn't have to click anything."
    """
    wb = writer.book
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    priority_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    alert_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    medium_fill = PatternFill(start_color='FFF4CC', end_color='FFF4CC', fill_type='solid')
    
    # Border for readability
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # ==========================================
        # STEP 1: Format Headers
        # ==========================================
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center', 
                wrap_text=True
            )
            cell.border = thin_border
        
        # ==========================================
        # STEP 2: Auto-fit ALL columns properly
        # ==========================================
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            
            # Calculate max length needed
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        # Handle multi-line content
                        lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in lines)
                        max_length = max(max_length, max_line_length)
                except:
                    pass
            
            # Set width with padding
            # Minimum 12 chars, maximum 80 chars for readability
            adjusted_width = min(max(max_length + 3, 12), 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # ==========================================
        # STEP 3: Set row heights and wrap text
        # ==========================================
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
            
            # Header row taller
            if row_idx == 1:
                ws.row_dimensions[row_idx].height = 30
            else:
                # Calculate height based on content
                max_lines = 1
                for cell in row:
                    if cell.value:
                        lines = str(cell.value).count('\n') + 1
                        max_lines = max(max_lines, lines)
                
                # 15 points per line + padding
                ws.row_dimensions[row_idx].height = max(15 * max_lines + 5, 20)
            
            # Format all cells in row
            for cell in row:
                if row_idx > 1:  # Not header
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
                    )
                    cell.border = thin_border
        
        # ==========================================
        # STEP 4: Freeze panes (header stays visible)
        # ==========================================
        ws.freeze_panes = 'A2'
        
        # ==========================================
        # STEP 5: Sheet-specific formatting
        # ==========================================
        
        # Executive Summary - Make it POP
        if 'EXECUTIVE_SUMMARY' in sheet_name:
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).font = Font(size=12, bold=True)
        
        # Top 5 Moves - Highlight priorities
        if 'TOP_5_MOVES' in sheet_name:
            priority_col = None
            
            # Find Priority column
            for col in range(1, ws.max_column + 1):
                if ws.cell(1, col).value == 'Priority':
                    priority_col = col
                    break
            
            if priority_col:
                for row in range(2, min(5, ws.max_row + 1)):  # Top 3 priorities
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = priority_fill
                        ws.cell(row, col).font = Font(bold=True, size=11)
        
        # Recovery Opportunities - Color code by potential
        if 'RECOVERY_OPPORTUNITIES' in sheet_name:
            potential_col = None
            
            # Find Recovery_Potential column
            for col in range(1, ws.max_column + 1):
                if 'Recovery_Potential' in str(ws.cell(1, col).value):
                    potential_col = col
                    break
            
            if potential_col:
                for row in range(2, ws.max_row + 1):
                    cell_value = str(ws.cell(row, potential_col).value)
                    
                    if 'HIGH' in cell_value:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row, col).fill = success_fill
                    elif 'MEDIUM' in cell_value:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row, col).fill = medium_fill
        
        # Reactive Alerts - Red highlight
        if 'REACTIVE_ALERTS' in sheet_name:
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = alert_fill
        
        # ==========================================
        # STEP 6: Add filters to data rows (except HOW_TO_USE)
        # ==========================================
        if ws.max_row > 1 and '0_HOW_TO_USE' not in sheet_name:
            ws.auto_filter.ref = ws.dimensions
        
        # ==========================================
        # STEP 7: Set print settings (if they print)
        # ==========================================
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically
        
        # Print titles (header repeats on each page)
        ws.print_title_rows = '1:1'
        
        # ==========================================
        # STEP 8: Zoom level for readability
        # ==========================================
        ws.sheet_view.zoomScale = 90  # Slightly zoomed out

        # ==========================================
        # HOW TO USE GUIDE TAB
        # ==========================================
        if 'HOW_TO_USE' in sheet_name:
            continue  # Skip HOW_TO_USE - it has custom formatting already!


def create_how_to_use_tab(writer, recommendations_count, high_priority_count, volume_at_stake):
    """
    Create a beautiful, engaging HOW_TO_USE guide tab as the FIRST tab.
    Uses colors, emojis, and clear sections to make it inviting and easy to understand.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = writer.book
    
    # Create or get the sheet (make it the FIRST tab)
    if "0_HOW_TO_USE" in wb.sheetnames:
        ws = wb["0_HOW_TO_USE"]
    else:
        ws = wb.create_sheet("0_HOW_TO_USE", 0)  # Insert as first sheet
    
    # ========================================
    # COLOR PALETTE (Vibrant & Professional)
    # ========================================
    SYSCO_BLUE = "0066CC"
    HEADER_BLUE = "4472C4"
    ACCENT_GREEN = "70AD47"
    ACCENT_ORANGE = "F4B183"
    ACCENT_RED = "E74C3C"
    LIGHT_BLUE = "D9E2F3"
    LIGHT_GREEN = "E2EFDA"
    LIGHT_YELLOW = "FFF4CC"

    
    # ========================================
    # FONT STYLES
    # ========================================
    body_font = Font(name="Calibri", size=11)
    bold_body = Font(name="Calibri", size=11, bold=True)
    
    # ========================================
    # FILL STYLES
    # ========================================
    header_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    light_green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    light_yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
    
    # ========================================
    # HELPER FUNCTIONS
    # ========================================
    def add_section_header(row, text, emoji=""):
        """Add a colored section header with WHITE font"""
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws.cell(row, 1, value=f"{emoji} {text}")
        cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")  # ← WHITE font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        return row + 1
        
    def add_text(row, text, indent=0, bold=False):
        """Add body text with optional indent"""
        indent_text = "   " * indent + text
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws.cell(row, 1, value=indent_text)
        cell.font = bold_body if bold else body_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 18
        return row + 1
    
    def add_callout_box(row, text, fill_color):
        """Add a highlighted callout box"""
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws.cell(row, 1, value=text)
        cell.font = bold_body
        cell.fill = fill_color
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 30
        return row + 1
    
    # ========================================
    # SET COLUMN WIDTHS
    # ========================================
    ws.column_dimensions['A'].width = 80
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 5
    
    # ========================================
    # BUILD THE GUIDE
    # ========================================
    row = 1
    
    # === MAIN TITLE (merged into one row with white font) ===
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row, 1, value="🎯 PRICING ZONE OPTIMIZATION GUIDE - Your roadmap to winning back customers and growing volume")
    cell.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")  # ← WHITE font
    cell.fill = PatternFill(start_color=SYSCO_BLUE, end_color=SYSCO_BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 50
    row += 2
    
    # === QUICK STATS BOX ===
    row = add_callout_box(row, 
        f"📊 THIS REPORT FOUND: {recommendations_count} opportunities | "
        f"{high_priority_count} high-priority moves | "
        f"{volume_at_stake:,.0f} lbs at stake",
        light_blue_fill)
    row += 2
    
    # === SECTION 1: WHAT IS THIS? ===
    row = add_section_header(row, "WHAT IS THIS REPORT?", "📖")
    row = add_text(row, "This report analyzes your historical sales to find pricing zones that are losing customers.", 1)
    row = add_text(row, "It recommends LOWER pricing zones (never higher) to win back customers who:", 1)
    row = add_text(row, "• Stopped buying this category but are still active Sysco customers", 2)
    row = add_text(row, "• Were overcharged, left, and we panicked and dropped the zone (reactive pricing)", 2)
    row = add_text(row, "• Are buying less than they used to because the price is too high", 2)
    row += 1
    row = add_callout_box(row, "💡 KEY INSIGHT: We only suggest LOWER prices. Never higher.", light_green_fill)
    row += 2
    
    # === SECTION 2: HOW TO READ THIS ===
    row = add_section_header(row, "HOW TO READ THIS DASHBOARD", "🔍")
    
    tabs = [
        ("Tab 1: Executive Summary", "Big picture - total opportunities, customer wins/losses, financial impact"),
        ("Tab 2: Top 5 Moves", "START HERE! The 5 most important zone changes to make right now"),
        ("Tab 3: Recovery Opportunities", "Customers who stopped buying but are still active - easiest wins!"),
        ("Tab 4: Reactive Alerts", "⚠️ Places where we already dropped zones reactively (data may be messy)"),
        ("Tab 5: Timeline", "Week-by-week schedule for implementing changes"),
        ("Tab 6: Learning Tracker", "(Empty first time) Tracks how accurate our predictions were"),
        ("Tab 7: All Recommendations", "Complete list of every suggested zone change"),
        ("Tab 8: YoY Customers", "How many customers we kept, gained, or lost compared to last year"),
        ("Tab 9: Volume Declines", "Biggest volume drops by zone - shows where we're bleeding"),
        ("Tab 10: Financial Impact", "Dollar projections - revenue and margin estimates for changes"),
        ("Tab 11: Zone Stickiness", "Which zones keep customers engaged week after week")
    ]
    
    for tab_name, description in tabs:
        row = add_text(row, f"{tab_name}:", 1, bold=True)
        row = add_text(row, description, 2)
    row += 2
    
    # === SECTION 3: KEY TERMS ===
    row = add_section_header(row, "KEY TERMS EXPLAINED (8TH GRADE LEVEL)", "📚")
    
    terms = [
        ("Pricing Zone", "A number (0-4) that determines the price customers pay. Lower number = lower price."),
        ("Combo", "A unique combination like 'Detroit + Seafood + CPA'. We recommend zones for each combo."),
        ("Lapsed Customer", "Someone who USED TO buy this category but stopped. Still buying other stuff from us though!"),
        ("YoY (Year-over-Year)", "Comparing this year to last year. Are we growing or shrinking?"),
        ("Reactive Pricing", "❌ BAD: We charged too much → lost customers → panicked and dropped the zone."),
        ("Green Flag Zone", "✅ GOOD: A zone where customers keep coming back week after week (75%+ consistency)."),
        ("Volume at Stake", "Total pounds we could win back if we fix the pricing zones."),
        ("Pounds CY/PY", "Current Year pounds vs Previous Year pounds. Shows if we're growing or losing volume."),
        ("Last Invoice Date", "Last time this customer bought ANYTHING from us (not just this category)."),
        ("Implementation Priority", "A score (1-100) that tells you which changes to make first.")
    ]
    
    for term, definition in terms:
        row = add_text(row, f"• {term}:", 1, bold=True)
        row = add_text(row, definition, 2)
    row += 2
    
    # === SECTION 4: WHAT TO DO NEXT ===
    row = add_section_header(row, "WHAT TO DO NEXT (ACTION STEPS)", "✅")
    
    steps = [
        ("STEP 1: Look at Tab 2 (Top 5 Moves)", "These are your highest-impact changes. Focus here first."),
        ("STEP 2: Check Tab 3 (Recovery Opportunities)", "These are the easiest wins - customers who just need a better price to come back."),
        ("STEP 3: Review Tab 10 (Financial Impact)", "See how much revenue and margin you'll gain from these changes."),
        ("STEP 4: Make the changes in your system", "Implement the recommended zone changes for the top combos."),
        ("STEP 5: Run this report again in 8 weeks", "See if the changes worked! Tab 6 will track your accuracy over time.")
    ]
    
    for step, description in steps:
        row = add_text(row, step, 1, bold=True)
        row = add_text(row, description, 2)
    row += 2
    
    row = add_callout_box(row, "⏱️ TIMELINE: Expect to see results in 4-8 weeks after making changes", light_yellow_fill)
    row += 2
    
    # === SECTION 5: TRAFFIC LIGHT GUIDE ===
    row = add_section_header(row, "TRAFFIC LIGHT GUIDE (WHAT THE COLORS MEAN)", "🚦")
    
    row = add_text(row, "Throughout this dashboard, we use colors to help you prioritize:", 1)
    row += 1
    
    # Red box
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row, 1, value="🔴 RED = HIGH PRIORITY / URGENT - Do these first!")
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Yellow box
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row, 1, value="🟡 YELLOW = MEDIUM PRIORITY - Important but not urgent")
    cell.font = Font(name="Calibri", size=12, bold=True, color="000000")  # ← BLACK text for yellow background
    cell.fill = PatternFill(start_color=ACCENT_ORANGE, end_color=ACCENT_ORANGE, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Green box
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row, 1, value="🟢 GREEN = GOOD / SUCCESS - Keep doing this!")
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 2
    
    # === SECTION 6: IMPORTANT RULES ===
    row = add_section_header(row, "IMPORTANT RULES WE FOLLOW", "⚖️")
    
    rules = [
        "✅ We NEVER suggest raising prices (going to a higher zone number)",
        "✅ We NEVER suggest Zone 1 → Zone 0 without flagging it for fractional zones",
        "✅ We only recommend changes that have strong historical data backing them",
        "✅ We prioritize winning back LAPSED customers (stopped buying but still active)",
        "✅ We flag REACTIVE pricing situations where data might be contaminated",
        "✅ We focus on volume AND margin together (not just one or the other)"
    ]
    
    for rule in rules:
        row = add_text(row, rule, 1, bold=True)
    row += 2
    
    # === SECTION 7: QUESTIONS? ===
    row = add_section_header(row, "QUESTIONS?", "❓")
    row = add_text(row, "If you have questions about this report or need help implementing changes:", 1)
    row = add_text(row, "• Contact your Analytics team", 2)
    row = add_text(row, "• Review the detailed tabs for more context", 2)
    row = add_text(row, "• Run the report again after making changes to track progress", 2)
    row += 2
    
    # === FINAL CALLOUT ===
    row = add_callout_box(row, 
        "🎉 YOU'RE READY! Go to Tab 2 (Top 5 Moves) and start winning back customers!",
        light_green_fill)
    
    # ========================================
    # IMPORTANT: Remove auto-filter from this sheet
    # ========================================
    if ws.auto_filter.ref:
        ws.auto_filter.ref = None
    
    print("   ✅ HOW_TO_USE tab created with beautiful formatting!")
    return ws

def write_stakeholder_excel(recommendations: List[Dict], 
                           reactive_flags: Dict,
                           output_path: str,
                           learning_engine: Optional[LearningEngine] = None,
                           optimization_engine: Optional['FoodserviceZoneEngine'] = None,
                           historical_df: Optional[pd.DataFrame] = None,
                           current_df: Optional[pd.DataFrame] = None):  
    
    from pandas import ExcelWriter
    
    print("\n📊 Creating stakeholder dashboard...")
    print("   ⚙️  Formatting for immediate readability...")
    
    # ADD THESE DEBUG LINES:
    print(f"\n🐛 DEBUG write_stakeholder_excel():")
    print(f"   optimization_engine = {optimization_engine}")
    print(f"   optimization_engine is None? {optimization_engine is None}")
    if optimization_engine:
        print(f"   Has current_df? {hasattr(optimization_engine, 'current_df')}")
        if hasattr(optimization_engine, 'current_df'):
            print(f"   current_df shape: {optimization_engine.current_df.shape}")
    
    
    print("\n📊 Creating stakeholder dashboard...")
    print("   ⚙️  Formatting for immediate readability...")
    
    # Get YoY metrics from engine
    yoy_metrics = None
    if optimization_engine and hasattr(optimization_engine, 'yoy_customer_metrics'):
        yoy_metrics = optimization_engine.yoy_customer_metrics
    
    # Calculate summary stats for the guide
    recommendations_count = len(recommendations)
    high_priority_count = sum(1 for r in recommendations if r.get('recommendation_type') == 'HIGH_RECOVERY_POTENTIAL')
    volume_at_stake = sum(r.get('total_volume', 0) for r in recommendations)
    
    with ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # ==========================================
        # CALCULATE FINANCIAL IMPACT FIRST (using current_df which has CY/PY data!)
        # ==========================================
        financial_summary = None
        financial_df = None
        
        # Use current_df from the optimization_engine 
        if current_df is not None and not current_df.empty and optimization_engine:
            print("   💰 Calculating financial impact using current data (has CY/PY columns)...")
            try:
                financial_df, financial_summary = calculate_financial_impact(
                    recommendations,
                    current_df,  # ← NEW - use the parameter!
                    optimization_engine.config
                )
                print(f"      ✅ Financial calculations complete!")
                if financial_summary:
                    print(f"      📊 Top 10 Expected Revenue (8wk): ${financial_summary.get('top10_revenue_8wk_expected', 0):,.0f}")
                    print(f"      📊 Top 10 Expected Margin (8wk): ${financial_summary.get('top10_margin_8wk_expected', 0):,.0f}")
            except Exception as e:
                print(f"      ⚠️ Financial calculations failed: {e}")
                import traceback
                traceback.print_exc()
                financial_summary = None
                financial_df = None
        else:
            print("   ⚠️ Skipping financial calculations (no current_df)")  
        
        # ==========================================
        # TAB 0: HOW_TO_USE (FIRST TAB - BEAUTIFUL!)
        # ==========================================
        print("   📄 Tab 0: HOW_TO_USE (Beautiful Guide)")
        create_how_to_use_tab(writer, recommendations_count, high_priority_count, volume_at_stake)
        
        # ==========================================
        # TAB 1: Executive Summary (WITH FINANCIAL DATA!)
        # ==========================================
        print("   📄 Tab 1: Executive Summary")
        if financial_summary:
            print(f"      💰 Including financial projections in Executive Summary")
        exec_summary = create_executive_summary(recommendations, yoy_metrics, reactive_flags, financial_summary, optimization_engine.config)

        # Convert to vertical layout
        exec_rows = [
            {'Metric': 'Total Opportunities Found', 'Value': exec_summary['total_opportunities'].iloc[0]},
            {'Metric': 'High Priority Moves (Do This Week)', 'Value': exec_summary['high_priority_moves'].iloc[0]},
            {'Metric': 'Easy Wins - Lapsed Customer Recovery', 'Value': exec_summary['easy_wins_lapsed_recovery'].iloc[0]},
            {'Metric': 'Fractional Zones Needed', 'Value': exec_summary['fractional_zones_needed'].iloc[0]},
            {'Metric': 'Total Volume at Stake', 'Value': exec_summary['total_volume_at_stake'].iloc[0]},
            {'Metric': 'Customers We Can Win Back', 'Value': exec_summary['customers_to_win_back'].iloc[0]},
            {'Metric': '', 'Value': ''},  # Spacer
            {'Metric': 'YEAR-OVER-YEAR CUSTOMER METRICS', 'Value': ''},
            {'Metric': 'Distinct Customers This Year', 'Value': exec_summary.get('distinct_customers_this_year', pd.Series(['N/A'])).iloc[0]},
            {'Metric': 'Distinct Customers Last Year', 'Value': exec_summary.get('distinct_customers_last_year', pd.Series(['N/A'])).iloc[0]},
            {'Metric': 'Customers Retained', 'Value': exec_summary.get('customers_retained', pd.Series(['N/A'])).iloc[0]},
            {'Metric': 'New Customers Gained', 'Value': exec_summary.get('new_customers_gained', pd.Series(['N/A'])).iloc[0]},
            {'Metric': 'Customers Lost', 'Value': exec_summary.get('customers_lost', pd.Series(['N/A'])).iloc[0]},
            {'Metric': 'Customer Change', 'Value': exec_summary.get('customer_change_count', pd.Series(['N/A'])).iloc[0]},
            {'Metric': 'Customer Change %', 'Value': exec_summary.get('customer_change_percent', pd.Series(['N/A'])).iloc[0]},
            {'Metric': 'Customer Retention Rate', 'Value': exec_summary.get('customer_retention_rate', pd.Series(['N/A'])).iloc[0]},
            {'Metric': '', 'Value': ''},  # Spacer
            {'Metric': 'Scope', 'Value': exec_summary['scope'].iloc[0]},
            {'Metric': 'YoY Lookback Window', 'Value': exec_summary['yoy_lookback_window'].iloc[0]},
            {'Metric': 'Comparison Period', 'Value': exec_summary['yoy_comparison_period'].iloc[0]},
            {'Metric': '', 'Value': ''},  # Spacer
            {'Metric': 'Expected Timeframe for Results', 'Value': exec_summary['expected_timeframe'].iloc[0]},
            {'Metric': 'First Action to Take', 'Value': exec_summary['first_action'].iloc[0]}
        ]
        exec_df = pd.DataFrame(exec_rows)
        exec_df.to_excel(writer, sheet_name='1_EXECUTIVE_SUMMARY', index=False)
        
        # ==========================================
        # TAB 2: Top 5 Quick Wins (with YoY!)
        # ==========================================
        print("   📄 Tab 2: Top 5 Quick Wins")
        quick_wins = create_quick_wins_dashboard(recommendations, yoy_metrics)
        if not quick_wins.empty:
            quick_wins.to_excel(writer, sheet_name='2_TOP_5_MOVES', index=False)
        
        # ==========================================
        # TAB 3: Customer Recovery Tracker
        # ==========================================
        print("   📄 Tab 3: Recovery Opportunities")
        recovery = create_customer_recovery_tracker(recommendations)
        
        if not recovery.empty:
            recovery.to_excel(writer, sheet_name='3_RECOVERY_OPPORTUNITIES', index=False)
        
        # ==========================================
        # TAB 4: Reactive Pricing Alerts
        # ==========================================
        if reactive_flags:
            print("   📄 Tab 4: Reactive Pricing Alerts")
            alerts = create_reactive_pricing_alerts(reactive_flags)
            if not alerts.empty:
                alerts.to_excel(writer, sheet_name='4_REACTIVE_ALERTS', index=False)
        
        
        # ==========================================
        # TAB 5: Implementation Timeline
        # ==========================================
        print("   📄 Tab 5: Implementation Timeline")
        timeline = create_implementation_timeline(recommendations)
        if not timeline.empty:
            timeline.to_excel(writer, sheet_name='5_TIMELINE', index=False)
        
        # ==========================================
        # TAB 6: Learning Tracker (if available)
        # ==========================================
        if learning_engine:
            print("   📄 Tab 6: Learning Tracker")
            learning_df = create_learning_tracker_tab(learning_engine)
            if not learning_df.empty:
                learning_df.to_excel(writer, sheet_name='6_LEARNING_TRACKER', index=False)
        
        # ==========================================
        # TAB 7: All Recommendations (detailed backup)
        # ==========================================
        print("   📄 Tab 7: All Recommendations (detailed)")
        all_recs = pd.DataFrame(recommendations)

        key_cols = [
            'company_name', 'cuisine', 'attribute_group', 'current_zone',
            'recommended_zone', 'recommendation_type', 'stakeholder_message',
            'expected_result', 'timeline', 'risk_level', 'total_volume',
            'lapsed_customers', 'implementation_priority'
        ]

        display_cols = [col for col in key_cols if col in all_recs.columns]
        if display_cols:
            all_recs_display = all_recs[display_cols].copy()
            
            all_recs_display.columns = [
                'OpCo', 'Cuisine', 'Attribute Group', 'Current Zone',
                'Recommended Zone', 'Type', 'Explanation',
                'Expected Result', 'Timeline', 'Risk', 'Volume (lbs)',
                'Lapsed Customers', 'Priority Score'
            ]
            
            # Sort by Volume descending
            all_recs_display = all_recs_display.sort_values('Volume (lbs)', ascending=False)
            
            all_recs_display.to_excel(writer, sheet_name='7_ALL_RECOMMENDATIONS', index=False)
            
            # ADD NOTE about Priority Score
            ws = writer.sheets['7_ALL_RECOMMENDATIONS']
            note_col = 14  # Column N (next to M which is Priority Score)
            
            ws.cell(1, note_col, value="ℹ️ HOW PRIORITY SCORE IS CALCULATED:")
            ws.cell(1, note_col).font = Font(bold=True, size=10, color="0066CC")
            
            ws.cell(2, note_col, value="Score (0-100) combines:")
            ws.cell(3, note_col, value="• Volume at stake (40%)")
            ws.cell(4, note_col, value="• Lapsed customers (30%)")
            ws.cell(5, note_col, value="• Recovery potential (20%)")
            ws.cell(6, note_col, value="• Risk level (10%)")
            ws.cell(7, note_col, value="Higher score = Do this first!")
            
            for i in range(2, 8):
                ws.cell(i, note_col).font = Font(size=9)
                ws.cell(i, note_col).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Make column N wider
            ws.column_dimensions['N'].width = 35
                    
        # ==========================================
        # TAB 8: YoY Customer Tracking 
        # ==========================================
        if yoy_metrics:
            print("   📄 Tab 8: Year-over-Year Customer Tracking")
            yoy_dashboard = create_yoy_customer_dashboard(yoy_metrics)
            if not yoy_dashboard.empty:
                # ✅ SORT by Lost Customers descending
                if 'Customers_Lost' in yoy_dashboard.columns:
                    yoy_dashboard = yoy_dashboard.sort_values('Customers_Lost', ascending=False)
                yoy_dashboard.to_excel(writer, sheet_name='8_YOY_CUSTOMERS', index=False)
                
        # ==========================================
        # TAB 9: Zone Stickiness Report
        # ==========================================
        if optimization_engine and hasattr(optimization_engine, 'consistency_analysis'):
            if optimization_engine.consistency_analysis is not None:
                print("   📄 Tab 9: Zone Stickiness Report")
                stickiness = create_zone_stickiness_report(optimization_engine.consistency_analysis)
                if not stickiness.empty:
                    # Debug: print column names to see exact spelling
                    print(f"      Stickiness columns: {list(stickiness.columns)}")
                    
                    # Try to find the Consistent Buyers column (might have different name)
                    consistent_col = None
                    for col in stickiness.columns:
                        if 'consistent' in col.lower() and 'buyer' in col.lower():
                            consistent_col = col
                            break
                    
                    # Sort by Consistent Buyers descending
                    if consistent_col:
                        stickiness = stickiness.sort_values(consistent_col, ascending=False)
                        print(f"      ✅ Sorted by '{consistent_col}' descending")
                    else:
                        print(f"      ⚠️ Could not find Consistent Buyers column")
                    
                    stickiness.to_excel(writer, sheet_name='9_ZONE_STICKINESS', index=False)
                    
                    # ADD NOTE about what Stickiness means
                    ws = writer.sheets['9_ZONE_STICKINESS']
                    note_col = stickiness.shape[1] + 2  # Two columns after last data column
                    
                    ws.cell(1, note_col, value="ℹ️ WHAT IS ZONE STICKINESS?")
                    ws.cell(1, note_col).font = Font(bold=True, size=11, color="0066CC")
                    
                    ws.cell(2, note_col, value="'Consistent Buyers' = Customers who buy")
                    ws.cell(3, note_col, value="in 75%+ of weeks")
                    ws.cell(4, note_col, value="")
                    ws.cell(5, note_col, value="HIGH (>75%) = GOOD!")
                    ws.cell(6, note_col, value="Price is right, customers loyal")
                    ws.cell(7, note_col, value="")
                    ws.cell(8, note_col, value="LOW (<50%) = WARNING!")
                    ws.cell(9, note_col, value="Customers sporadic, may shop around")
                    ws.cell(10, note_col, value="")
                    ws.cell(11, note_col, value="Use this to validate GREEN FLAG zones")
                    
                    for i in range(2, 12):
                        ws.cell(i, note_col).font = Font(size=10)
                        ws.cell(i, note_col).alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Make note column wider
                    col_letter = ws.cell(1, note_col).column_letter
                    ws.column_dimensions[col_letter].width = 40
        # ==========================================
        # TAB 10: Volume Decline Analysis 
        # ==========================================
        print("   📄 Tab 10: Volume Decline by Combo + Zone")
        volume_declines = create_volume_decline_dashboard(
            historical_df if historical_df is not None else current_df, 
            optimization_engine.config,
            top_n=50  # Show top 50 worst performers
        )
        if not volume_declines.empty:
            volume_declines.to_excel(writer, sheet_name='10_VOLUME_DECLINES', index=False)

        # ==========================================
        # TAB 10: Financial Impact (NEEDS historical_df!)
        # ==========================================
        if financial_df is not None and not financial_df.empty:
            print("   📄 Tab 1: Financial Impact Projections")
            financial_display = create_financial_impact_dashboard(financial_df, financial_summary)
            financial_display.to_excel(writer, sheet_name='11_FINANCIAL_IMPACT', index=False)
        else:
            print("   ⚠️ Skipping Tab 1: No financial data available")
        
        # ==========================================
        # Apply formatting to ALL sheets
        # ==========================================
        print("   🎨 Applying formatting...")
        _format_excel_sheets(writer)
    
    print(f"   ✅ Dashboard complete and ready to open!")
    print(f"   📁 Saved to: {output_path}")


# ========================================
# NOTE: You'll also need to update create_executive_summary_tab()
# ========================================
# Make sure your create_executive_summary_tab function signature looks like this:
#
# def create_executive_summary_tab(recommendations, yoy_metrics, reactive_flags, financial_summary=None):
#     """Create executive summary with optional financial projections."""
#     # ... your existing code ...
#     
#     # Add section for financial summary if available:
#     if financial_summary:
#         summary_dict['Expected_Revenue_Lift_8wk'] = financial_summary.get('expected_revenue_8wk', 0)
#         summary_dict['Expected_Margin_Lift_8wk'] = financial_summary.get('expected_margin_8wk', 0)
#         summary_dict['Expected_Revenue_Annual'] = financial_summary.get('expected_revenue_annual', 0)
#         summary_dict['Expected_Margin_Annual'] = financial_summary.get('expected_margin_annual', 0)
#     
#     return pd.DataFrame([summary_dict])


# ========================================
# IMPORTANT: HOW TO CALL THIS FUNCTION
# ========================================
# In your main run_foodservice_analysis() function:
#
# write_stakeholder_excel(
#     recommendations, 
#     reactive_flags, 
#     excel_path, 
#     learning_engine, 
#     engine,
#     historical_df=historical_df  # ← KEEP THIS! Will be None on first run, that's OK!
# )
#
# On FIRST RUN: historical_df is None, but financial calcs still work (uses current_df)
# On FUTURE RUNS: historical_df has data, learning tracker works


# ========================================
# TROUBLESHOOTING: If financial data isn't showing
# ========================================
# 1. Check console output - it should print:
#    "💰 Calculating financial impact using current data (has CY/PY columns)..."
#    "✅ Financial calculations complete!"
#    "💰 Including financial projections in Executive Summary"
#
# 2. If you see "⚠️ Skipping financial calculations", then optimization_engine.current_df doesn't exist
#
# 3. Make sure your current_df has these columns (normalized names):
#    - Net_Sales_CY and Net_Sales_PY
#    - Margin_CY and Margin_PY
#    - Pounds_CY and Pounds_PY
#
# 4. Make sure the keys match what you're expecting in create_executive_summary_tab():
#    financial_summary should have keys like:
#    - 'top10_revenue_8wk_expected'
#    - 'top10_margin_8wk_expected'
#    - 'all_revenue_annual_expected'
#    etc.
#
# 5. The calculate_financial_impact() function should use the CY/PY columns
#    from current_df (which has both years of data in one file!)

def create_comparison_export(recommendations: List[Dict], 
                             output_path: str,
                             learning_engine: Optional[LearningEngine] = None,
                             yoy_metrics: Optional[Dict] = None):
    """Create CSV with YoY customer baseline."""
    
    comparison_data = []
    
    # Get combo-level YoY data
    yoy_by_combo = {}
    if yoy_metrics and 'by_combo' in yoy_metrics:
        for _, row in yoy_metrics['by_combo'].iterrows():
            yoy_by_combo[row['Company_Combo_Key']] = row
    
    for rec in recommendations:
        combo = rec['company_combo']
        yoy_data = yoy_by_combo.get(combo, {})
        
        comparison_data.append({
            'run_date': datetime.now().strftime('%Y-%m-%d'),
            'combo_key': combo,
            'company': rec['company_name'],
            'cuisine': rec.get('cuisine', 'N/A'),
            'attribute_group': rec.get('attribute_group', 'N/A'),
            'current_zone': rec['current_zone'],
            'recommended_zone': rec['recommended_zone'],
            'recommendation_type': rec['recommendation_type'],
            
            # Volume baselines
            'total_volume_baseline': rec['total_volume'],
            'active_customers_baseline': rec['active_customers'],
            'lapsed_customers_baseline': rec['lapsed_customers'],
            
            # YoY customer baselines (NEW!)
            'distinct_customers_cy_baseline': yoy_data.get('distinct_customers_cy', None),
            'distinct_customers_py_baseline': yoy_data.get('distinct_customers_py', None),
            'customer_retention_rate_baseline': yoy_data.get('retention_rate', None),
            
            # Prediction
            'predicted_volume_lift': rec.get('expected_result', 'N/A'),
            'implementation_priority': rec['implementation_priority'],
            
            # For next run
            'was_implemented': False,
            'actual_volume_after': None,
            'actual_customers_after': None,
            'actual_distinct_customers_cy_after': None  
        })
    
    comparison_df = pd.DataFrame(comparison_data)
    comparison_df.to_csv(output_path, index=False)
    
    print(f"   💾 Comparison file saved: {output_path}")
    print(f"   📝 Use this file to track results in your next run")


# ============================================================================
# MAIN EXECUTION
# ============================================================================
CATEGORY_NAME = "groundfish"
def run_foodservice_zone_optimization(
    current_data_path: str,
    historical_data_paths: Optional[List[str]] = None,
    output_name: str = 'zone_optimization',
    data_config: Optional[DataConfiguration] = None,
    enable_learning: bool = True,
    yoy_lookback_weeks: int = 8,  
    filter_company_number: Optional[str] = None,  
    filter_company_region_id: Optional[str] = None  
):
    """
    Main function with configurable lookback and filtering.
    
    Args:
        current_data_path: Path to current week's data CSV
        historical_data_paths: List of paths to historical data CSVs
        output_name: Base name for output files
        data_config: Configuration for which columns to use
        enable_learning: Whether to use learning system
        yoy_lookback_weeks: How many weeks to compare for YoY (default 8)
        filter_company_number: Optional - filter to specific Company Number
        filter_company_region_id: Optional - filter to specific Company Region ID
    """
    
    print("🚀 Starting Foodservice Zone Optimization Analysis")
    print("=" * 60)
    
    # Setup configuration with new parameters
    if data_config is None:
        data_config = DataConfiguration()
    
    input_config = InputConfiguration(
        current_data_path=current_data_path,
        historical_data_paths=historical_data_paths or []
    )
    
    config = FoodserviceConfig(
        data_config, 
        input_config,
        yoy_lookback_weeks=yoy_lookback_weeks,  
        filter_company_number=filter_company_number,    
        filter_company_region_id=filter_company_region_id       
    )
    
    # Show filter settings
    if filter_company_number or filter_company_region_id:
        print("\n🔍 FILTERING ENABLED:")
        if filter_company_number:
            print(f"   • Company Number: {filter_company_number}")
        if filter_company_region_id:
            print(f"   • Company Region ID: {filter_company_region_id}")
    
    # Validate paths
    valid, issues = input_config.validate_paths()
    if not valid:
        print("❌ Configuration issues:")
        for issue in issues:
            print(f"   • {issue}")
        return None, None
    
    # Load current data
    print("\n📂 Loading current data...")
    current_df = pd.read_csv(current_data_path, low_memory=False)
    print(f"   ✅ Loaded {len(current_df):,} rows")
    
    # Validate columns
    valid, missing = data_config.validate_dataframe(current_df)
    if not valid:
        print("❌ Missing required columns:")
        for col in missing:
            print(f"   • {col}")
        return None, None
    
    # Normalize
    current_df = _normalize_columns(current_df, data_config)
    
    # Apply filters (NEW!)
    current_df = apply_filters(current_df, config)
    if current_df.empty:
        print("❌ No data after filtering!")
        return None, None
    
    # Load historical data
    historical_df = None
    if historical_data_paths:
        print("\n📂 Loading historical data...")
        hist_dfs = []
        for path in historical_data_paths:
            if os.path.exists(path):
                df = pd.read_csv(path, low_memory=False)
                df = _normalize_columns(df, data_config)
                hist_dfs.append(df)
                print(f"   ✅ Loaded {os.path.basename(path)}: {len(df):,} rows")
        
        if hist_dfs:
            historical_df = pd.concat(hist_dfs, ignore_index=True)
            print(f"   ✅ Combined historical: {len(historical_df):,} rows")
            
            # Apply filters to historical too (NEW!)
            historical_df = apply_filters(historical_df, config)
            print(f"   ✅ After filters: {len(historical_df):,} rows")
    
    # Initialize learning engine
    learning_engine = None
    if enable_learning:
        print("\n🧠 Initializing learning system...")
        learning_engine = LearningEngine(input_config.learning_file_path)
        print(f"   ✅ Loaded {len(learning_engine.recommendations)} past recommendations")
    
    # Initialize optimization engine
    print("\n🔧 Initializing optimization engine...")
    engine = FoodserviceZoneEngine(config, learning_engine)
    
    # Generate recommendations
    print("\n🎯 Generating recommendations...")
    recommendations = engine.generate_recommendations(current_df, historical_df)
    print(f"   ✅ Generated {len(recommendations)} recommendations")
    
    # Get reactive flags
    reactive_flags = engine.reactive_flags
    print(f"   ⚠️  Found {len(reactive_flags)} reactive pricing patterns")
    # ==========================================
    # FILTER TO HIGH-IMPACT RECOMMENDATIONS ONLY
    # ==========================================
    print("\n🎯 Filtering to high-impact recommendations...")

    original_count = len(recommendations)

    # Define impact thresholds
    MIN_VOLUME = 50  
    MIN_CUSTOMERS = 3  
    MIN_PRIORITY = 60  
    MIN_VOLUME_PER_CUSTOMER = 150  # High-value customer threshold

    high_impact = []

    for rec in recommendations:
        volume_potential = rec['total_volume']
        if rec['lapsed_customers'] > 0:
            volume_potential += rec['total_volume'] * 0.3  
        
        # Calculate volume per customer
        total_customers = rec['total_customers']
        volume_per_customer = volume_potential / total_customers if total_customers > 0 else 0
        
        # Check all criteria
        meets_volume = volume_potential >= MIN_VOLUME
        meets_priority = rec['implementation_priority'] >= MIN_PRIORITY
        
        # Need 3+ customers AND EITHER 5+ customers OR high volume per customer
        meets_customers = (
            rec['total_customers'] >= MIN_CUSTOMERS and
            (rec['total_customers'] >= 5 or volume_per_customer >= MIN_VOLUME_PER_CUSTOMER)
        )
        
        if meets_volume and meets_customers and meets_priority:
            high_impact.append(rec)

    print(f"   📊 Filtered from {original_count} → {len(high_impact)} recommendations")
    print(f"   🎯 Focus: Volume ≥{MIN_VOLUME} lbs, Priority ≥{MIN_PRIORITY}, Customers ≥3 (with 5+ OR {MIN_VOLUME_PER_CUSTOMER}+ lbs/customer)")

    # Replace recommendations with filtered list
    recommendations = high_impact
    # Save to learning system
    if learning_engine:
        print("\n💾 Saving recommendations to learning system...")
        for rec in recommendations:
            predicted_outcomes = {
                'volume_lift': rec['total_volume'] * 0.20,  # Conservative estimate
                'customer_recovery': rec.get('lapsed_customers', 0) * 0.6,
                'timeline_weeks': 6
            }
            learning_engine.save_recommendation(rec, predicted_outcomes)
        print(f"   ✅ Saved {len(recommendations)} recommendations for future tracking")
    
    # Create output paths
    timestamp = config.get_timestamp()
    excel_path = os.path.join(
        input_config.output_directory,
        f"{output_name}_{timestamp}.xlsx"
    )
    comparison_path = os.path.join(
        input_config.output_directory,
        f"{output_name}_comparison_{timestamp}.csv"
    )
    
    # Write Excel (with perfect formatting)
    write_stakeholder_excel(recommendations, reactive_flags, excel_path, learning_engine, engine, historical_df, current_df)                                                                                                        

    # ==========================================
    # GENERATE LEADS FILE (if enabled)
    # ==========================================
    if config.GENERATE_LEADS_FILE:
        print("\n" + "="*50)
        print("🎯 GENERATING SALES LEADS FILE")
        print("="*50)
        
        leads_df = generate_high_value_leads(
            recommendations=recommendations,
            current_df=current_df,
            config=config,
            output_path=config.LEADS_OUTPUT_PATH
        )
        
        print(f"\n✅ Leads file ready for FinalCat.py: {config.LEADS_OUTPUT_PATH}")
    
    # Create comparison export
    create_comparison_export(recommendations, comparison_path, learning_engine)
    
    # Print summary
    print("\n" + "=" * 60)
    print("✅ ANALYSIS COMPLETE")
    print("=" * 60)
    
    exec_summary = create_executive_summary(recommendations)
    print(f"\n📈 KEY FINDINGS:")
    print(f"   • Total Opportunities: {exec_summary['total_opportunities'].iloc[0]}")
    print(f"   • High Priority Moves: {exec_summary['high_priority_moves'].iloc[0]}")
    print(f"   • Easy Wins (Lapsed Recovery): {exec_summary['easy_wins_lapsed_recovery'].iloc[0]}")
    print(f"   • Volume at Stake: {exec_summary['total_volume_at_stake'].iloc[0]}")
    print(f"   • Customers to Win Back: {exec_summary['customers_to_win_back'].iloc[0]:,}")
    print(f"\n🎯 FIRST ACTION:")
    print(f"   {exec_summary['first_action'].iloc[0]:}")
    print(f"\n📅 EXPECTED TIMEFRAME: {exec_summary['expected_timeframe'].iloc[0]:}")
    
    print(f"\n📁 FILES CREATED:")
    print(f"   • Dashboard: {excel_path}")
    print(f"   • Comparison: {comparison_path}")
    if learning_engine:
        print(f"   • Learning State: {learning_engine.learning_file_path}")
    
    print(f"\n💡 NEXT STEPS:")
    print(f"   1. Open {os.path.basename(excel_path)} - NO RESIZING NEEDED!")
    print(f"   2. Review Tab 2 (Top 5 Moves) with stakeholders")
    print(f"   3. Implement recommendations")
    print(f"   4. Re-run this analysis in 4-6 weeks to track results")
    
    return recommendations, reactive_flags


if __name__ == "__main__":
    
    # ==========================================
    # CONFIGURATION
    # ==========================================
    
    # Set to specific company number to filter, or None for all companies
    TARGET_COMPANY = None  # Change to '13' to filter to Company 13
    
    # Enable/disable learning system persistence
    ENABLE_LEARNING_PERSISTENCE = False  # Set to True to track recommendations over time
    
    # ==========================================
    # STEP 1: Configure your data columns
    # ==========================================
    
    data_config = DataConfiguration(
    # Required columns 
        company_column='Company Name',
        customer_id_column='Company Customer Number',
        last_invoice_date_column='Last Invoice Date',
        fiscal_week_column='Fiscal Week Number',
        pounds_cy_column='Pounds CY',
        pounds_py_column='Pounds PY',
        zone_column='Zone_Suffix_Numeric',
        
        # Filtering columns
        company_number_column='Company Number', 
        company_region_id_column='Company Region ID',  
        
        # Toggle grouping columns
        use_cuisine=True,
        cuisine_column='NPD Cuisine Type',
        
        use_attribute_group=True,
        attribute_group_column='Attribute Group ID',
        
        use_business_center=False,
        business_center_column='Business Center ID',
        
        use_item_group=False,
        item_group_column='Item Group ID',
        
        use_price_source=True,
        price_source_column='Price Source Type'
    )
    
    HISTORICAL_DATA = [
        r"C:\Users\kmor6669\OneDrive - Sysco Corporation\Desktop\Pricing\groundfish_1325_1326.csv",
        r"C:\Users\kmor6669\OneDrive - Sysco Corporation\Desktop\Pricing\groundfish_1324_1225.csv"
    ]
    
    # ==========================================
    # STEP 2: Load and extract current baseline
    # ==========================================

    print("🔍 Extracting current baseline from historical data...")
    all_hist = []
    for path in HISTORICAL_DATA:
        if os.path.exists(path):
            df = pd.read_csv(path, low_memory=False)
            all_hist.append(df)
            print(f"   ✅ Loaded {os.path.basename(path)}: {len(df):,} rows")

    combined = pd.concat(all_hist, ignore_index=True)
    # Find latest week using COMBINED format (YYYYWW)
    combined['Fiscal_Year'] = pd.to_numeric(combined['Fiscal Year Key'], errors='coerce').fillna(0).astype(int)
    combined['Fiscal_Week_Num'] = pd.to_numeric(combined['Fiscal Week Number'], errors='coerce').fillna(0).astype(int)
    combined['Fiscal_Week_Combined'] = (combined['Fiscal_Year'] * 100 + combined['Fiscal_Week_Num']).astype(int)
    
    latest_week_combined = combined['Fiscal_Week_Combined'].max()
    latest_week = latest_week_combined % 100  # Keep this for compatibility
    fiscal_year = latest_week_combined // 100
    
    print(f"\n   📅 Latest week in data: {latest_week_combined} (FY{fiscal_year} Week {latest_week})")
    
    # Extract current - LAST 8 WEEKS using combined identifier
    lookback_weeks = 8
    current_week_start = latest_week_combined - lookback_weeks + 1
    
    current_df = combined[
        combined['Fiscal_Week_Combined'].between(current_week_start, latest_week_combined)
    ].copy()
    
    print(f"   📊 Using weeks {current_week_start}-{latest_week_combined} as 'current' baseline (last {lookback_weeks} weeks)")
    print(f"   ✅ Current baseline: {len(current_df):,} rows")

    # Save to temp WITH ORIGINAL COLUMN NAMES
    temp_current_path = os.path.join(
        r"C:\Users\kmor6669\OneDrive - Sysco Corporation\Desktop\Pricing",
        f"temp_current_{latest_week}.csv"
    )
    current_df.to_csv(temp_current_path, index=False)
    print(f"   💾 Saved current baseline: {temp_current_path}")

    # NOW normalize for filtering
    combined = _normalize_columns(combined, data_config)

    # TEMP: Save normalized historical for audit
    normalized_hist_path = os.path.join(
        r"C:\Users\kmor6669\OneDrive - Sysco Corporation\Desktop\Pricing",
        "temp_normalized_historical.csv"
    )
    combined.to_csv(normalized_hist_path, index=False)
    print(f"   💾 Saved normalized historical for audit: {normalized_hist_path}")
    
# ==========================================
# STEP 3: Generate report
# ==========================================

    if TARGET_COMPANY:
        print(f"\n{'='*60}")
        print(f"Processing Company Number: {TARGET_COMPANY}")
        print('='*60)
        
        # Look up company name
        filtered_data = combined[
            combined[data_config.company_number_column].astype(str).str.strip().str.lstrip('0') == str(TARGET_COMPANY).lstrip('0')
        ]
        
        if filtered_data.empty:
            print(f"   ❌ No data found for Company Number '{TARGET_COMPANY}'")
            exit()
        
        company_name = filtered_data[data_config.company_column].iloc[0]
        safe_company_name = str(company_name).replace(' ', '_').replace('/', '_')
        output_name = f'{safe_company_name}_{TARGET_COMPANY}_{CATEGORY_NAME}_zones'
        
        print(f"   📍 Company Name: {company_name}")
    else:
        print(f"\n{'='*60}")
        print(f"Processing ALL Companies - Aggregated Opportunities")
        print('='*60)
        output_name = f'ALL_COMPANIES_{CATEGORY_NAME}_zones'

    # Run analysis
    recommendations, reactive_flags = run_foodservice_zone_optimization(
        current_data_path=temp_current_path,
        historical_data_paths=HISTORICAL_DATA,
        output_name=output_name,
        data_config=data_config,
        enable_learning=ENABLE_LEARNING_PERSISTENCE,  
        filter_company_number=TARGET_COMPANY,
        filter_company_region_id=None
    )

    if TARGET_COMPANY:
        print(f"\n✅ Report complete for {company_name} (Company #: {TARGET_COMPANY})")
    else:
        print(f"\n✅ Aggregated report complete for ALL companies")
    
    # Cleanup (KEEP FILES FOR AUDIT)
    print("\n📋 FILES SAVED FOR LOGIC AUDIT:")
    print(f"   • Historical: {normalized_hist_path}")
    print(f"   • Current: {temp_current_path}")
    print("\n🎉 ALL DONE!")
    print("\n💡 Run audit with:")
    print(f"   python logic_audit.py {normalized_hist_path} {temp_current_path}")
    
