#!/usr/bin/env python3
# category_site_benchmark.py
# Single-site performance vs company benchmark

import argparse
import pandas as pd
import numpy as np
from datetime import datetime
from category_utils import clean_numeric_column, load_and_prepare, compute_windows

# python category_site_benchmark.py --source "C:\Users\kmor6669\Sysco Corporation\SBC APP - Pictures\Seafood\Shrimp_Prawns\Inputs\550.csv" --alignment "C:\Users\kmor6669\Sysco Corporation\SBC APP - Pictures\Seafood\Shrimp_Prawns\Alignment550.csv" --company "55" --outdir "C:\Users\kmor6669\Sysco Corporation\SBC APP - Pictures\Seafood\Shrimp_Prawns\Outputs"


def calculate_site_benchmarks(site_data, company_data):
    """
    Compare one site's performance to overall company.
    Shows overall + breakdown by account type + Sysco brand split.
    """
    
    rows = []
    
    def add_benchmark_row(account_type, site_subset, company_subset):
        """Helper to calculate one benchmark row."""
        site_cy = site_subset["Pounds_CY"].sum()
        site_py = site_subset["Pounds_PY"].sum()
        site_delta = site_cy - site_py
        site_yoy_pct = (site_delta / site_py) if site_py > 0 else 0
        
        company_cy = company_subset["Pounds_CY"].sum()
        company_py = company_subset["Pounds_PY"].sum()
        company_delta = company_cy - company_py
        company_yoy_pct = (company_delta / company_py) if company_py > 0 else 0
        
        site_share_cy = (site_cy / company_cy) if company_cy > 0 else 0
        site_share_py = (site_py / company_py) if company_py > 0 else 0
        share_change = site_share_cy - site_share_py
        performance_gap = site_yoy_pct - company_yoy_pct
        
        rows.append({
            "Account_Type": account_type,
            "Site_CY": site_cy,
            "Site_PY": site_py,
            "Site_YoY_%": site_yoy_pct,
            "Company_YoY_%": company_yoy_pct,
            "Gap_vs_Company": performance_gap,
            "Site_Market_Share_CY": site_share_cy,
            "Site_Market_Share_PY": site_share_py,
            "Share_Change": share_change
        })
    
    # === OVERALL ===
    add_benchmark_row("All Accounts", site_data, company_data)
    
    # === BY ACCOUNT TYPE ===
    if "Customer Account Type Code" in site_data.columns:
        for acct_type in ["TRS", "CMU", "LCC"]:
            site_acct = site_data[site_data["Customer Account Type Code"] == acct_type]
            company_acct = company_data[company_data["Customer Account Type Code"] == acct_type]
            
            if not site_acct.empty and not company_acct.empty:
                add_benchmark_row(f"  {acct_type}", site_acct, company_acct)
                
                # === SYSCO BRAND SPLIT ===
                if "Sysco Brand Indicator" in site_data.columns:
                    for brand in ["Y", "N"]:
                        site_brand = site_acct[site_acct["Sysco Brand Indicator"] == brand]
                        company_brand = company_acct[company_acct["Sysco Brand Indicator"] == brand]
                        
                        if not site_brand.empty and not company_brand.empty:
                            brand_label = "Sysco" if brand == "Y" else "Non-Sysco"
                            add_benchmark_row(f"    {acct_type} - {brand_label}", site_brand, company_brand)
    
    return pd.DataFrame(rows)

def calculate_site_ranking(site_name, all_sites_df):
    """Show where this site ranks among all company sites."""
    
    # Aggregate by site
    site_summary = all_sites_df.groupby("Company Name", dropna=False).agg(
        Pounds_CY=("Pounds_CY", "sum"),
        Pounds_PY=("Pounds_PY", "sum"),
        Delta_YoY_Lbs=("Delta_YoY_Lbs", "sum"),
        Customers_CY=("Distinct_Customers_CY", "sum"),
        Customers_PY=("Distinct_Customers_PY", "sum")
    ).reset_index()
    
    site_summary["YoY_Pct"] = np.where(
        site_summary["Pounds_PY"] > 0,
        site_summary["Delta_YoY_Lbs"] / site_summary["Pounds_PY"],
        np.nan
    )
    
    # Sort by Delta YoY (ascending = worst first)
    site_summary = site_summary.sort_values("Delta_YoY_Lbs", ascending=True).reset_index(drop=True)
    site_summary["Rank"] = range(1, len(site_summary) + 1)
    
    # Find this site's rank
    your_rank = site_summary[site_summary["Company Name"] == site_name]
    
    return site_summary, your_rank

def _overall_yoy_site(df):
    """Calculate overall YoY metrics for site."""
    row = {
        "Pounds CY": df["Pounds_CY"].sum(),
        "Pounds PY": df["Pounds_PY"].sum(),
        "Delta Pounds YoY": df["Delta_YoY_Lbs"].sum(),
        "Customers CY": df["Distinct_Customers_CY"].sum(),
        "Customers PY": df["Distinct_Customers_PY"].sum(),
        "Delta Customers": df["Delta_Customers"].sum(),
    }
    row["YoY %"] = (row["Delta Pounds YoY"]/row["Pounds PY"]) if row["Pounds PY"] else np.nan
    row["Customer Retention %"] = (row["Customers CY"]/row["Customers PY"]) if row["Customers PY"] > 0 else np.nan
    row["Avg Lbs/Customer CY"] = (row["Pounds CY"]/row["Customers CY"]) if row["Customers CY"] > 0 else 0
    row["Avg Lbs/Customer PY"] = (row["Pounds PY"]/row["Customers PY"]) if row["Customers PY"] > 0 else 0
    return pd.DataFrame([row])

def _brand_split_site(df):
    """Calculate Sysco vs Non-Sysco split."""
    if "Sysco Brand Indicator" not in df.columns:
        return pd.DataFrame(columns=["Sysco Brand Indicator","Pounds_CY","Pounds_PY","Delta_YoY_Lbs","YoY_Pct"])
    
    g = df.groupby("Sysco Brand Indicator", dropna=False).agg(
        Pounds_CY=("Pounds_CY","sum"),
        Pounds_PY=("Pounds_PY","sum"),
        Delta_YoY_Lbs=("Delta_YoY_Lbs","sum")
    ).reset_index()
    g["YoY_Pct"] = np.where(g["Pounds_PY"]>0, g["Delta_YoY_Lbs"]/g["Pounds_PY"], np.nan)
    return g

def _items_rank_site(df):
    """Rank items by Delta YoY."""
    keys = []
    for k in ["Item Number", "Item Description", "Brand ID", "Brand"]:
        if k in df.columns:
            keys.append(k)
    if not keys:
        return pd.DataFrame(columns=["Item","Item Description","Brand ID","Pounds_CY","Pounds_PY","Delta_YoY_Lbs","YoY_Pct"])

    g = df.groupby(keys, dropna=False).agg(
        Pounds_CY=("Pounds_CY","sum"),
        Pounds_PY=("Pounds_PY","sum"),
        Delta_YoY_Lbs=("Delta_YoY_Lbs","sum")
    ).reset_index()
    g["YoY_Pct"] = np.where(g["Pounds_PY"]>0, g["Delta_YoY_Lbs"]/g["Pounds_PY"], np.nan)

    if "Item Number" in g.columns:
        g.rename(columns={"Item Number":"Item"}, inplace=True)

    return g.sort_values("Delta_YoY_Lbs", ascending=True)

def _format_table_at_site(ws, header_row_0idx, n_rows, number_headers=None, percent_headers=None):
    """Format table with proper number and percentage formatting."""
    from openpyxl.styles import Alignment
    
    number_headers = set(number_headers or [])
    percent_headers = set(percent_headers or [])
    header_row_1idx = header_row_0idx + 1
    first_data = header_row_1idx + 1
    last_data = header_row_1idx + n_rows
    
    if n_rows <= 0:
        return

    # Map header text to column index
    header_to_col = {str(c.value).strip(): c.col_idx
                     for c in ws[header_row_1idx] if c.value is not None}

    def _coerce(cell, as_percent):
        v = cell.value
        if v is None or isinstance(v, (int, float)):
            return
        s = str(v).replace(",", "").strip()
        if not s:
            return
        try:
            if as_percent and s.endswith("%"):
                cell.value = float(s[:-1]) / 100.0
            else:
                cell.value = float(s)
        except Exception:
            pass

    # Format numbers
    for h in number_headers:
        col = header_to_col.get(h)
        if not col:
            continue
        for row in ws.iter_rows(min_row=first_data, max_row=last_data, min_col=col, max_col=col):
            _coerce(row[0], as_percent=False)
            row[0].number_format = "#,##0"

    # Format percents
    for h in percent_headers:
        col = header_to_col.get(h)
        if not col:
            continue
        for row in ws.iter_rows(min_row=first_data, max_row=last_data, min_col=col, max_col=col):
            _coerce(row[0], as_percent=True)
            row[0].number_format = "0.0%"

def _write_site_account_tab(writer, tab_name, df, site_name):
    """Write account type tab with metrics."""
    start = 0
    
    # Add site name header
    ws = writer.book.create_sheet(tab_name)
    ws.cell(1, 1, value=f"{site_name} - {tab_name.replace('_', ' ')}").font = Font(bold=True, size=14, color="0066CC")
    start = 2
    
    # 1) OVERALL
    overall = _overall_yoy_site(df)
    overall.to_excel(writer, sheet_name=tab_name, index=False, startrow=start)
    _format_table_at_site(
        ws, header_row_0idx=start, n_rows=overall.shape[0],
        number_headers={"Pounds CY","Pounds PY","Delta Pounds YoY","Customers CY","Customers PY","Delta Customers","Avg Lbs/Customer CY","Avg Lbs/Customer PY"},
        percent_headers={"YoY %","Customer Retention %"}
    )
    start += overall.shape[0] + 2

    # 2) BRAND SPLIT
    brand = _brand_split_site(df)
    if not brand.empty:
        brand.to_excel(writer, sheet_name=tab_name, index=False, startrow=start)
        _format_table_at_site(
            ws, header_row_0idx=start, n_rows=brand.shape[0],
            number_headers={"Pounds_CY","Pounds_PY","Delta_YoY_Lbs"},
            percent_headers={"YoY_Pct"}
        )
        start += brand.shape[0] + 2

    # 3) ITEMS (Top 20)
    items = _items_rank_site(df).head(20)
    if not items.empty:
        ws.cell(start + 1, 1, value="Top 20 Items by Delta YoY").font = Font(bold=True, size=11)
        items.to_excel(writer, sheet_name=tab_name, index=False, startrow=start + 1)
        _format_table_at_site(
            ws, header_row_0idx=start + 1, n_rows=items.shape[0],
            number_headers={"Pounds_CY","Pounds_PY","Delta_YoY_Lbs"},
            percent_headers={"YoY_Pct"}
        )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True)
    parser.add_argument("--alignment", required=True)
    parser.add_argument("--company", required=True, help="Site number or name to benchmark")
    parser.add_argument("--outdir", default=".")
    args = parser.parse_args()
    
    print(f"\n🎯 SITE BENCHMARK REPORT")
    print(f"Site: {args.company}")
    print("="*60)
    
    # Load ALL sites first (for company benchmark)
    company_raw, _ = load_and_prepare(args.source, args.alignment)
    company_status, current_week = compute_windows(company_raw)
    
    # Load FILTERED site data
    site_raw, _ = load_and_prepare(args.source, args.alignment, company=args.company)
    site_status, _ = compute_windows(site_raw)
    
    if site_status.empty:
        print(f"❌ No data found for site '{args.company}'")
        return
    
    # Calculate benchmarks
    benchmarks = calculate_site_benchmarks(site_status, company_status)

    # === DSM SUMMARY FOR THIS SITE ===
    if not site_status.empty:
        from final import build_dsm_opportunity_scorecard
        
        # Get source path for win-back calculations
        dsm_summary, territory_detail, winback_targets, conversion_targets = build_dsm_opportunity_scorecard(
            site_status, site_raw, args.source, active_weeks=6, min_ytd=min_ytd
        )
    
    # Save to Excel
    output_path = f"{args.outdir}/Site_{args.company}_Report_{datetime.now():%Y%m%d}.xlsx"
    
    from openpyxl.styles import PatternFill, Font
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        
        # TAB 1: Benchmark Summary
        benchmarks.to_excel(writer, sheet_name="Benchmark", index=False)
        
        ws_bench = writer.book["Benchmark"]
        
        # Add title and explanation
        ws_bench.insert_rows(1, 3)
        ws_bench.cell(1, 1, value=f"SITE PERFORMANCE BENCHMARK: {site_name}").font = Font(bold=True, size=14, color="0066CC")
        ws_bench.cell(2, 1, value=f"How this site compares to company-wide average | FY Week {current_week}").font = Font(size=10, italic=True, color="666666")
        
        # === ADD EXPLAINER TEXT ===
        explain_row = ws_bench.max_row + 3
        
        ws_bench.cell(explain_row, 1, value="HOW TO READ THIS REPORT:").font = Font(bold=True, size=12, color="0066CC")
        explain_row += 1
        
        ws_bench.cell(explain_row, 1, value="Gap vs Company: Negative (red) = site underperforming company average | Positive (green) = outperforming")
        explain_row += 1
        
        ws_bench.cell(explain_row, 1, value="Market Share Change: Negative = losing share of company volume | Positive = gaining share")
        explain_row += 1
        
        ws_bench.cell(explain_row, 1, value="Hierarchy: Indented rows show account type breakdowns and Sysco vs Non-Sysco brand splits")
        
        # TAB 2: Site Ranking
        site_ranking.to_excel(writer, sheet_name="Site_Ranking", index=False)
        
        # Highlight your site in yellow
        ws_rank = writer.book["Site_Ranking"]
        if not your_rank.empty:
            your_row = your_rank.iloc[0]["Rank"] + 1  # +1 for header
            for col in range(1, 8):
                ws_rank.cell(your_row, col).fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                ws_rank.cell(your_row, col).font = Font(bold=True)
        
        # Add title
        ws_rank.insert_rows(1, 2)
        ws_rank.cell(1, 1, value=f"COMPANY-WIDE SITE RANKING (Your Site: {site_name})").font = Font(bold=True, size=14, color="0066CC")
        ws_rank.cell(2, 1, value=f"All sites ranked by Delta YoY Pounds | Your site highlighted in yellow").font = Font(size=10, italic=True, color="666666")
        
        # === ACCOUNT TYPE TABS ===
        _write_site_account_tab(writer, "03_All_Accounts", site_status.copy(), site_name)
        
        # TRS
        site_trs_full = site_status[site_status["Customer Account Type Code"] == "TRS"].copy() if "Customer Account Type Code" in site_status.columns else pd.DataFrame()
        if not site_trs_full.empty:
            _write_site_account_tab(writer, "04_TRS", site_trs_full, site_name)
        
        # LCC
        site_lcc = site_status[site_status["Customer Account Type Code"] == "LCC"].copy() if "Customer Account Type Code" in site_status.columns else pd.DataFrame()
        if not site_lcc.empty:
            _write_site_account_tab(writer, "05_LCC", site_lcc, site_name)
        
        # CMU
        site_cmu = site_status[site_status["Customer Account Type Code"] == "CMU"].copy() if "Customer Account Type Code" in site_status.columns else pd.DataFrame()
        if not site_cmu.empty:
            _write_site_account_tab(writer, "06_CMU", site_cmu, site_name)
        
        # TAB: DSM Opportunities (if exists)
        if not dsm_summary.empty:
            dsm_summary.to_excel(writer, sheet_name="07_DSM_Opportunities", index=False)
            
            ws_dsm = writer.book["07_DSM_Opportunities"]
            ws_dsm.insert_rows(1, 2)
            ws_dsm.cell(1, 1, value=f"DSM OPPORTUNITIES AT {site_name}").font = Font(bold=True, size=14, color="0066CC")
            ws_dsm.cell(2, 1, value="Win-Back = Active customers declining | Conversion = Buying but not aligned").font = Font(size=10, italic=True, color="666666")
        
        # TAB: TRS Leads
        if not trs_leads.empty:
            trs_leads.to_excel(writer, sheet_name="08_Site_Leads", index=False)
            
            ws_leads = writer.book["08_Site_Leads"]
            ws_leads.insert_rows(1, 3)
            ws_leads.cell(1, 1, value=f"TRS SALES LEADS FOR {site_name}").font = Font(bold=True, size=14, color="0066CC")
            ws_leads.cell(2, 1, value=f"{len(trs_leads)} opportunities | TRS accounts only | Minimum {min_ytd:,.0f} lbs PY").font = Font(size=10, italic=True)
            ws_leads.cell(3, 1, value="'Needs Both' = highest priority (neither item nor vendor aligned)")
            
            print(f"\n📋 TRS Leads: {len(trs_leads)} opportunities")
        
        # === GENERATE VENDOR SPLITS (separate CSV files) ===
        import os
        vendor_dir = os.path.join(args.outdir, "Vendor_Leads")
        os.makedirs(vendor_dir, exist_ok=True)

        if not trs_leads.empty and "Aligned Vendor (SUVC)" in trs_leads.columns:
            for vendor_code in trs_leads["Aligned Vendor (SUVC)"].dropna().unique():
                vendor_leads = trs_leads[trs_leads["Aligned Vendor (SUVC)"] == vendor_code]
                vendor_name = vendor_leads["Aligned Supplier Name"].iloc[0] if "Aligned Supplier Name" in vendor_leads.columns else "Unknown"
                safe_name = "".join(c for c in str(vendor_name) if c.isalnum() or c in (' ', '-', '_')).strip()[:30]
                
                csv_path = os.path.join(vendor_dir, f"Site{args.company}_Vendor_{vendor_code}_{safe_name}.csv")
                vendor_leads.to_csv(csv_path, index=False)
            
            print(f"📦 Created vendor CSV files in {vendor_dir}")

    print(f"\n✅ Report saved: {output_path}")
    print("\n📊 Performance Summary:")
    print(benchmarks.to_string(index=False))    

if __name__ == "__main__":
    main()
